from ctypes import alignment
from urllib import request
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from .forms import UsuarioRegistroForm
from .models import Documento, Usuario, Semillero,SemilleroUsuario, Archivo, Aprendiz, ProyectoAprendiz, Proyecto, UsuarioProyecto, SemilleroProyecto, Entregable, SemilleroDocumento, Evento
from django.utils import timezone 
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import force_str, force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.contrib.auth.tokens import default_token_generator
from django.conf import settings
from django.template.loader import render_to_string
from functools import wraps
from django.utils.timezone import now
from django.db.models import Q, Avg
from datetime import datetime, timedelta
from django.db.models import Case, When, Value, IntegerField
from django.http import JsonResponse
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from datetime import date
from openpyxl.styles import Border, Side, Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from io import BytesIO
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from .forms import AprendizForm
from cryptography.fernet import Fernet
from django.conf import settings
import base64
import hashlib
from django.http import JsonResponse
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from itertools import chain
from operator import attrgetter
from . import models
from django.db.models import Value, CharField, F, Q, DateTimeField
from django.db.models.functions import Cast
from itertools import chain
from django.urls import reverse
from django.utils.dateparse import parse_date

# Funciones de cifrado/descifrado
def cifrar_numero(numero):
    if not numero:
        return numero
    try:
        key = hashlib.sha256(settings.SECRET_KEY.encode()).digest()
        cipher = Fernet(base64.urlsafe_b64encode(key))
        return cipher.encrypt(str(numero).encode()).decode()
    except:
        return numero

def descifrar_numero(numero_cifrado):
    if not numero_cifrado:
        return numero_cifrado
    try:
        key = hashlib.sha256(settings.SECRET_KEY.encode()).digest()
        cipher = Fernet(base64.urlsafe_b64encode(key))
        return cipher.decrypt(numero_cifrado.encode()).decode()
    except:
        return "****"  # Mostrar asteriscos si falla

# NOTIFICACIONES
def obtener_notificaciones(request):
    notificaciones = []
    ahora = timezone.now()

    # 1. VALIDAR SESIÓN
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        return []

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        return []

    # 2. EVENTOS PRÓXIMOS (24h)
    eventos = Evento.objects.filter(
        fecha_eve__gte=ahora.date(),
        fecha_eve__lte=(ahora + timedelta(days=1)).date(),
        estado_eve__in=['Próximo', 'Programado']
    ).order_by('fecha_eve', 'hora_inicio')[:3]

    for evento in eventos:
        dt_evento = datetime.combine(evento.fecha_eve, evento.hora_inicio)

        if timezone.is_naive(dt_evento):
            dt_evento = timezone.make_aware(dt_evento, timezone.get_current_timezone())

        minutos = (dt_evento - ahora).total_seconds() / 60
        if minutos < 0:
            continue

        horas = int(minutos // 60)

        notificaciones.append({
            'tipo': 'evento',
            'icono': 'fa-calendar-check',
            'clase_icono': 'warning',
            'titulo': f'Evento Próximo: {evento.nom_eve}',
            'mensaje': f'Comienza en {horas} hora(s)' if horas > 0 else f'Comienza en {int(minutos)} minuto(s)',
            'tiempo': evento.fecha_eve.strftime('%d/%m/%Y'),
            'url': reverse('eventos'),
            'leida': False
        })

    # 3. ENTREGABLES PENDIENTES (próximos 7 días)
    proyectos_usuario = UsuarioProyecto.objects.filter(
        cedula=usuario
    ).values_list('cod_pro', flat=True)

    entregables_pendientes = Entregable.objects.filter(
        cod_pro__in=proyectos_usuario,
        estado='Pendiente',
        fecha_fin__gte=ahora.date(),
        fecha_fin__lte=(ahora + timedelta(days=7)).date()
    ).select_related('cod_pro').order_by('fecha_fin')[:3]

    for entregable in entregables_pendientes:
        sempro = entregable.cod_pro.semilleroproyecto_set.first()

        if sempro:
            url = (
                reverse('detalle-proyecto', kwargs={
                    'id_sem': sempro.id_sem.id_sem,
                    'cod_pro': entregable.cod_pro.cod_pro
                }) + '?tab=entregables'
            )
        else:
            url = reverse('proyectos')

        dias_restantes = (entregable.fecha_fin - ahora.date()).days

        notificaciones.append({
            'tipo': 'entregable',
            'icono': 'fa-file-alt',
            'clase_icono': 'warning' if dias_restantes <= 2 else 'success',
            'titulo': f'Entregable: {entregable.nom_entre}',
            'mensaje': f'Vence en {dias_restantes} día(s) - {entregable.cod_pro.nom_pro}',
            'tiempo': entregable.fecha_fin.strftime('%d/%m/%Y'),
            'url': url,
            'leida': False
        })

    # 4. PROYECTOS CON BAJO PROGRESO (<30%)
    proyectos_estancados = Proyecto.objects.filter(
        usuarioproyecto__cedula=usuario,
        progreso__lt=30,
        estado_pro__in=['planeacion', 'ejecucion']
    ).order_by('progreso')[:2]

    for proyecto in proyectos_estancados:
        sempro = proyecto.semilleroproyecto_set.first()

        if sempro:
            url = reverse('detalle-proyecto', kwargs={
                'id_sem': sempro.id_sem.id_sem,
                'cod_pro': proyecto.cod_pro
            })
        else:
            url = reverse('proyectos')

        notificaciones.append({
            'tipo': 'proyecto',
            'icono': 'fa-exclamation-triangle',
            'clase_icono': 'error',
            'titulo': 'Proyecto con bajo progreso',
            'mensaje': f'{proyecto.nom_pro} - {proyecto.progreso}% completado',
            'tiempo': 'Requiere atención',
            'url': url,
            'leida': False
        })

    # 5. NUEVOS MIEMBROS (solo si es líder)
    mis_semilleros = SemilleroUsuario.objects.filter(
        cedula=usuario,
        es_lider=True
    ).values_list('id_sem', flat=True)

    nuevos_miembros = SemilleroUsuario.objects.filter(
        id_sem__in=mis_semilleros
    ).select_related('cedula', 'id_sem').order_by('-semusu_id')[:2]

    for miembro in nuevos_miembros:
        if miembro.cedula.cedula == usuario.cedula:
            continue

        url = reverse('resu-miembros', kwargs={'id_sem': miembro.id_sem.id_sem})

        notificaciones.append({
            'tipo': 'miembro',
            'icono': 'fa-user-plus',
            'clase_icono': 'success',
            'titulo': 'Nuevo miembro en semillero',
            'mensaje': f'{miembro.cedula.nom_usu} se unió a {miembro.id_sem.nombre}',
            'tiempo': 'Reciente',
            'url': url,
            'leida': False
        })

    # 6. ENTREGABLES RETRASADOS
    entregables_retrasados = Entregable.objects.filter(
        cod_pro__in=proyectos_usuario,
        estado='Retrasado'
    ).select_related('cod_pro')[:2]

    for entregable in entregables_retrasados:
        sempro = entregable.cod_pro.semilleroproyecto_set.first()

        if sempro:
            url = (
                reverse('detalle-proyecto', kwargs={
                    'id_sem': sempro.id_sem.id_sem,
                    'cod_pro': entregable.cod_pro.cod_pro
                }) + '?tab=entregables'
            )
        else:
            url = reverse('proyectos')

        notificaciones.append({
            'tipo': 'retrasado',
            'icono': 'fa-clock',
            'clase_icono': 'error',
            'titulo': 'Entregable retrasado',
            'mensaje': f'{entregable.nom_entre} - {entregable.cod_pro.nom_pro}',
            'tiempo': f'Venció el {entregable.fecha_fin.strftime("%d/%m/%Y")}',
            'url': url,
            'leida': False
        })

    return notificaciones

def api_notificaciones(request):
    from django.http import JsonResponse
    try:
        notificaciones = obtener_notificaciones(request)
        return JsonResponse({
            'notificaciones': notificaciones,
            'count': len(notificaciones)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()

        return JsonResponse({
            'notificaciones': [],
            'count': 0,
            'error': str(e)
        }, status=500)

# VISTA BIENVENIDO
def bienvenido(request):

    # CONTADORES DINÁMICOS
    investigadores = Usuario.objects.filter(rol="Investigador").count()
    instructores = Usuario.objects.filter(rol="Instructor").count()
    semilleros = Semillero.objects.count()
    proyectos = Proyecto.objects.count()
    capacidad_instalada = Proyecto.objects.filter(tipo="capacidadinstalada").count()  
    proyectos_formativos = Proyecto.objects.filter(tipo="Formativo").count()
    sennova = Proyecto.objects.filter(tipo="Sennova").count()  

    context = {
        'investigadores': investigadores,
        'instructores': instructores,
        'semilleros': semilleros,
        'proyectos': proyectos,
        'capacidad_instalada': capacidad_instalada,
        'proyectos_formativos': proyectos_formativos,
        'sennova': sennova,
    }

    return render(request, 'paginas/bienvenido.html', context)

# VISTAS DE LOGIN
def registro(request):
    if request.method == 'POST':
        form = UsuarioRegistroForm(request.POST)
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                
                # Valores por defecto explícitos
                usuario.email_verificado = False
                usuario.is_active = True       
                usuario.estado = 'Activo'     
                
                usuario.save()

                # Enviar correo de verificación
                if hasattr(usuario, 'generar_token_verificacion'):
                    usuario.generar_token_verificacion()
                if hasattr(usuario, 'enviar_email_verificacion'):
                    usuario.enviar_email_verificacion(request)

                messages.success(request, "¡Registro exitoso! Verifica tu correo electrónico.")
                return redirect('iniciarsesion')
            except Exception as e:
                messages.error(request, f"Error al registrar usuario: {e}")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = UsuarioRegistroForm()

    return render(request, 'paginas/registro.html', {
        'form': form,
        'current_page_name': 'registro',
        'show_register': True  
    })

# Vista para verificar el correo electrónico
def verificar_email(request, token):
    try:
        # Buscar el administrador con este token
        usuario = Usuario.objects.get(token_verificacion=token)
        
        # Verificar si el token ha expirado
        if usuario.token_expira and usuario.token_expira < timezone.now():
            messages.error(request, "El enlace de verificación ha expirado. Por favor, solicita uno nuevo.")
            return redirect('iniciarsesion')
        
        # Marcar como verificado
        usuario.email_verificado = True
        usuario.token_verificacion = None
        usuario.token_expira = None
        usuario.save()
        
        messages.success(request, "¡Tu correo electrónico ha sido verificado correctamente! Ahora puedes iniciar sesión.")
        return redirect('iniciarsesion')
        
    except Usuario.DoesNotExist:
        messages.error(request, "El enlace de verificación no es válido.")
        return redirect('iniciarsesion') 

# VISTA DE INICIAR SESION
def iniciarsesion(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        password = request.POST.get('password')
        rol = request.POST.get('rol')

        errores = {}

        # Validaciones básicas
        if not rol:
            errores['error_rol'] = "Debe seleccionar un rol."
        if not cedula:
            errores['error_user'] = "La cédula es obligatoria."
        if not password:
            errores['error_password'] = "La contraseña es obligatoria."

        if errores:
            return render(request, 'paginas/registro.html', {
                **errores,
                'cedula': cedula,
                'rol': rol,
                'current_page_name': 'Iniciar Sesión'
            })

        # Buscar usuario por cédula
        try:
            usuario = Usuario.objects.get(cedula=cedula)
        except Usuario.DoesNotExist:
            return render(request, 'paginas/registro.html', {
                'error_user': 'Usuario no registrado.',
                'cedula': cedula,
                'rol': rol,
                'current_page_name': 'Iniciar Sesión'
            })

        # Verificar contraseña
        if not usuario.check_password(password):
            return render(request, 'paginas/registro.html', {
                'error_password': 'Contraseña incorrecta.',
                'cedula': cedula,
                'rol': rol,
                'current_page_name': 'Iniciar Sesión'
            })

        # Verificar rol
        if usuario.rol != rol:
            return render(request, 'paginas/registro.html', {
                'error_rol': 'El rol seleccionado no coincide con tu usuario.',
                'cedula': cedula,
                'rol': rol,
                'current_page_name': 'Iniciar Sesión'
            })

        # Verificar correo electrónico
        if not usuario.email_verificado:
            return render(request, 'paginas/registro.html', {
                'error_user': 'Debes verificar tu correo antes de iniciar sesión.',
                'cedula': cedula,
                'rol': rol,
                'current_page_name': 'Iniciar Sesión'
            })

        # ✅ TODO CORRECTO - Crear sesión personalizada
        request.session['cedula'] = usuario.cedula
        request.session['nom_usu'] = usuario.nom_usu
        request.session['ape_usu'] = usuario.ape_usu
        request.session['rol'] = usuario.rol
        request.session['correo_ins'] = usuario.correo_ins
        
        # Configurar tiempo de expiración de sesión (opcional)
        request.session.set_expiry(3600)  # 1 hora
        
        # Actualizar último acceso
        usuario.last_login = timezone.now()
        usuario.save(update_fields=['last_login'])

        messages.success(request, f"¡Bienvenido, {usuario.nom_usu}!")
        return redirect('home')

    return render(request, 'paginas/registro.html', {
        'current_page_name': 'Iniciar Sesión'
    })

def mostrar_recuperar_contrasena(request):

    return render(request, 'paginas/registro.html', {
        'mostrar_modal': True,
        'show_login': True,  # <- para que cargue login directamente
        'current_page_name': 'Recuperar Contraseña'
    })

def recuperar_contrasena(request):
    """
    Esta vista procesa el formulario de recuperación de contraseña
    """
    if request.method == 'POST':
        email = request.POST.get('email')
        
        if not email:
            return render(request, 'iniciarsesion', {
                'mostrar_modal': True,
                'email_error': 'El correo electrónico es obligatorio',
                'current_page_name': 'Recuperar Contraseña'
            })
        
        try:
            # Verificar si existe un administrador con ese correo
            admin = Usuario.objects.filter(correo_ins=email).first()
            
            if admin:
                # Generar el token y el uid codificado para el enlace de restablecimiento
                uid = urlsafe_base64_encode(force_bytes(admin.pk))
                token = default_token_generator.make_token(admin)
                
                # Construir el enlace de restablecimiento
                reset_link = f"{request.scheme}://{request.get_host()}/reset-password/{uid}/{token}/"
                
                try:
                    # Preparar y enviar el correo
                    subject = "Restablecimiento de contraseña"
                    message = render_to_string('paginas/reset_password_email.html', {
                        'user': admin,
                        'reset_link': reset_link,
                        'site_name': 'GC'
                    })
                    
                    # Cambiado fail_silently a False para que muestre errores
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [admin.correo_ins],
                        html_message=message,
                        fail_silently=False
                    )
                    print(f"Correo enviado exitosamente a {admin.correo_ins}")
                except Exception as email_error:
                    print(f"Error al enviar correo: {email_error}")
                    # Ahora mostramos el error al usuario para facilitar la depuración
                    messages.error(request, f"Problema al enviar el correo: {email_error}")
                    return redirect('iniciarsesion')
                
            # Por seguridad, mostramos un mensaje genérico independientemente de si el correo existe o no
            messages.success(request, "Si el correo está asociado a una cuenta, recibirás instrucciones para restablecer tu contraseña.")
            return redirect('iniciarsesion')
            
        except Exception as e:
            print(f"Error durante recuperación de contraseña: {e}")
            messages.error(request, f"{str(e)}")
            return redirect('iniciarsesion')
    
    # Si no es POST, redirigir a la página de inicio de sesión
    return redirect('iniciarsesion')

def reset_password(request, uidb64, token):
    """
    Vista que muestra el formulario de restablecimiento de contraseña
    cuando el usuario hace clic en el enlace del correo.
    """
    try:
        # 1️⃣ Decodificar el UID
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(pk=uid)

        # 2️⃣ Validar el token
        if default_token_generator.check_token(usuario, token):
            print(f"✅ Token válido para el usuario: {usuario.nom_usu}")
            return render(request, 'paginas/reset_password.html', {
                'valid': True,
                'uidb64': uidb64,
                'token': token,
                'current_page_name': 'Restablecer Contraseña'
            })
        else:
            print("❌ Token inválido o expirado")
            messages.error(request, "El enlace de restablecimiento no es válido o ha expirado.")
            return redirect('iniciarsesion')

    except Exception as e:
        print(f"⚠️ Error en reset_password: {e}")
        messages.error(request, "Error al procesar el enlace de restablecimiento.")
        return redirect('iniciarsesion')

def reset_password_confirm(request):
    """
    Vista que procesa el formulario de restablecimiento de contraseña.
    """
    if request.method == 'POST':
        uidb64 = request.POST.get('uidb64')
        token = request.POST.get('token')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        # Validar coincidencia de contraseñas
        if password1 != password2:
            return render(request, 'paginas/reset_password.html', {
                'valid': True,
                'uidb64': uidb64,
                'token': token,
                'error': 'Las contraseñas no coinciden.',
                'current_page_name': 'Restablecer Contraseña'
            })

        try:
            # Decodificar el UID y obtener el usuario
            uid = force_str(urlsafe_base64_decode(uidb64))
            usuario = Usuario.objects.get(pk=uid)

            # Verificar token válido
            if default_token_generator.check_token(usuario, token):

                # Guardar contraseña de forma segura
                usuario.set_password(password1)
                usuario.save()

                messages.success(request, "Tu contraseña ha sido restablecida con éxito. Ahora puedes iniciar sesión.")
                return redirect('iniciarsesion')
            else:
                messages.error(request, "El enlace de restablecimiento no es válido o ha expirado.")
                return redirect('iniciarsesion')

        except (TypeError, ValueError, OverflowError, Usuario.DoesNotExist) as e:
            print(f"⚠️ Error al restablecer contraseña: {e}")
            messages.error(request, "El enlace de restablecimiento no es válido.")
            return redirect('iniciarsesion')
    
    # Si no es POST, redirigir a la página de inicio de sesión
    return redirect('iniciarsesion')

def login_required(view_func):
    """
    Decorador personalizado que verifica si el usuario ha iniciado sesión
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
    
        cedula = request.session.get('cedula')
        
        if not cedula:
            messages.error(request, "Debes iniciar sesión para acceder a esta página")
            return redirect('iniciarsesion')
        
        # Verificar que el usuario existe en la base de datos
        try:
            from .models import Usuario
            usuario = Usuario.objects.get(cedula=cedula)
    
            # Verificar estado con manejo seguro de None
            estado_usuario = getattr(usuario, 'estado', None)
            if estado_usuario and estado_usuario != 'Activo':
                messages.error(request, "Tu cuenta no está activa")
                return redirect('iniciarsesion')
                
        except Usuario.DoesNotExist:
            # Si el usuario no existe, limpiar la sesión
            request.session.flush()
            messages.error(request, "Usuario no encontrado. Por favor inicia sesión nuevamente")
            return redirect('iniciarsesion')
        
        # Todo OK, ejecutar la vista
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view

# VISTA DE PRIVACIDAD
@login_required
def privacidad(request):
    usuario_id = request.session.get('cedula')
    
    try:
        usuario = Usuario.objects.get(pk=usuario_id)
        
        if request.method == 'POST':
            contraseña_actual = request.POST.get('contraseña_actual')
            nueva_contraseña = request.POST.get('nueva_contraseña')
            confirmar_contraseña = request.POST.get('confirmar_contraseña')
            
            #  Verificar la contraseña actual usando el campo correcto
            if not check_password(contraseña_actual, usuario.password):
                messages.error(request, "La contraseña actual es incorrecta.")
                return redirect('privacidad')
            
            #  Verificar coincidencia
            if nueva_contraseña != confirmar_contraseña:
                messages.error(request, "Las contraseñas nuevas no coinciden.")
                return redirect('privacidad')
            
            # Validar longitud mínima
            if len(nueva_contraseña) < 8:
                messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
                return redirect('privacidad')
            
            # Guardar correctamente usando set_password()
            usuario.set_password(nueva_contraseña)
            usuario.save()

            messages.success(request, "Contraseña actualizada correctamente.")
            return redirect('home')
            
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado. Por favor, inicie sesión nuevamente.")
        return redirect('iniciarsesion')
    except Exception as e:
        messages.error(request, f"Error al actualizar la contraseña: {str(e)}")
        return redirect('privacidad')
    
    return render(request, 'paginas/home.html', {
        'current_page': 'privacidad',
        'current_page_name': 'Privacidad'
    })

# VISTAS DE HOME 
@login_required
def home(request):
    cedula = request.session.get('cedula')

    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    # Contar semilleros donde participa el usuario
    total_semilleros = usuario.semilleros.count()
    total_proyectos = usuario.proyectos.count()
    total_aprendices = Aprendiz.objects.filter(
        proyectos__in=usuario.proyectos.all()
    ).distinct().count()

    # ========== ACTIVIDADES RECIENTES ==========
    # Obtener semilleros donde el usuario participa
    mis_semilleros_ids = SemilleroUsuario.objects.filter(
        cedula=usuario
    ).values_list('id_sem', flat=True)
    
    # Obtener proyectos donde el usuario participa (directamente)
    mis_proyectos_ids = UsuarioProyecto.objects.filter(
        cedula=usuario
    ).values_list('cod_pro', flat=True)
    
    # Fecha límite (última semana)
    fecha_limite = timezone.now() - timedelta(days=7)
    
    # ===== SEMILLEROS RECIENTES =====
    # Mostrar semilleros creados o modificados donde participo
    semilleros_recientes = Semillero.objects.filter(
        id_sem__in=mis_semilleros_ids,
        fecha_creacion__gte=fecha_limite
    ).annotate(
        tipo_actividad=Value('semillero', output_field=CharField()),
        fecha_actividad=F('fecha_creacion')
    )
    
    # ===== PROYECTOS RECIENTES =====
    # Opción 1: Proyectos donde participo directamente
    proyectos_directos = Proyecto.objects.filter(
        cod_pro__in=mis_proyectos_ids,
        fecha_creacion__gte=fecha_limite
    )
    
    # Opción 2: Proyectos de mis semilleros
    proyectos_semilleros = Proyecto.objects.filter(
        semilleroproyecto__id_sem__in=mis_semilleros_ids,
        fecha_creacion__gte=fecha_limite
    )
    
    # Combinar ambos (sin duplicados)
    proyectos_recientes = (proyectos_directos | proyectos_semilleros).distinct().annotate(
        tipo_actividad=Value('proyecto', output_field=CharField()),
        fecha_actividad=F('fecha_creacion')
    )
    
    # ===== ENTREGABLES RECIENTES =====
    # De todos los proyectos donde participo (directos + de semilleros)
    todos_mis_proyectos = (
        Proyecto.objects.filter(cod_pro__in=mis_proyectos_ids) | 
        Proyecto.objects.filter(semilleroproyecto__id_sem__in=mis_semilleros_ids)
    ).distinct().values_list('cod_pro', flat=True)
    
    entregables_recientes = Entregable.objects.filter(
        cod_pro__in=todos_mis_proyectos,
        fecha_inicio__gte=fecha_limite.date()
    ).annotate(
        tipo_actividad=Value('entregable', output_field=CharField()),
        fecha_actividad=Cast(F('fecha_inicio'), output_field=DateTimeField())
    )
    
    # ===== EVENTOS RECIENTES =====
    # Mostrar TODOS los eventos recientes (incluidos los que yo creé)
    eventos_recientes = Evento.objects.filter(
        fecha_eve__gte=fecha_limite.date()
    ).annotate(
        tipo_actividad=Value('evento', output_field=CharField()),
        
        fecha_actividad=Cast(F('fecha_eve'), output_field=DateTimeField())
    )
    
    # ===== COMBINAR Y ORDENAR =====
    actividades = sorted(
        chain(semilleros_recientes, proyectos_recientes, entregables_recientes, eventos_recientes),
        key=lambda x: x.fecha_actividad if hasattr(x, 'fecha_actividad') else timezone.now(),
        reverse=True
    )[:10]  # Mostrar las últimas 10 actividades

    return render(request, 'paginas/home.html', {
        'current_page': 'home',
        'current_page_name': 'Inicio',
        'usuario': usuario,
        'total_semilleros': total_semilleros,
        'total_proyectos': total_proyectos,
        'total_aprendices': total_aprendices,
        'actividades': actividades,
        
    })

# VISTAS PERFIL
@login_required
def perfil(request):
    cedula = request.session.get('cedula')
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    # Calcular último acceso
    ultimo_acceso = "Sin registro"
    if usuario.last_login:
        tiempo = timezone.now() - usuario.last_login
        dias, segundos = tiempo.days, tiempo.seconds
        if dias == 0:
            if segundos < 60:
                ultimo_acceso = "Hace unos segundos"
            elif segundos < 3600:
                minutos = segundos // 60
                ultimo_acceso = f"Hace {minutos} minuto{'s' if minutos != 1 else ''}"
            else:
                horas = segundos // 3600
                ultimo_acceso = f"Hace {horas} hora{'s' if horas != 1 else ''}"
        elif dias == 1:
            ultimo_acceso = "Ayer"
        elif dias < 7:
            ultimo_acceso = f"Hace {dias} día{'s' if dias != 1 else ''}"
        elif dias < 30:
            semanas = dias // 7
            ultimo_acceso = f"Hace {semanas} semana{'s' if semanas != 1 else ''}"
        elif dias < 365:
            meses = dias // 30
            ultimo_acceso = f"Hace {meses} mes{'es' if meses != 1 else ''}"
        else:
            años = dias // 365
            ultimo_acceso = f"Hace {años} año{'s' if años != 1 else ''}"

    # Actualización de datos del perfil
    if request.method == 'POST':
        usuario.nom_usu = request.POST.get('nombre1')
        usuario.ape_usu = request.POST.get('nombre2')
        usuario.correo_per = request.POST.get('correo1')
        usuario.telefono = request.POST.get('celular')
        usuario.fecha_nacimiento = request.POST.get('fecha')
        usuario.genero = request.POST.get('genero')
        usuario.correo_ins = request.POST.get('correo2')
        usuario.vinculacion_laboral = request.POST.get('vinculacion')
        usuario.dependencia = request.POST.get('dependencia')
        usuario.rol = request.POST.get('rol')
        usuario.save()
        
        # Actualizar también la sesión
        request.session['nom_usu'] = usuario.nom_usu
        request.session['ape_usu'] = usuario.ape_usu
        request.session['rol'] = usuario.rol
        
        messages.success(request, "Cambios guardados correctamente.")
        return redirect('perfil')

    semilleros = usuario.semilleros.all()
    total_semilleros = semilleros.count()

    proyectos = SemilleroProyecto.objects.filter(
        cod_pro__usuarios=usuario
    ).select_related('id_sem', 'cod_pro')
    total_proyectos = proyectos.count()

    context = {
        'usuario': usuario,
        'ultimo_acceso': ultimo_acceso,
        'semilleros': semilleros,
        'total_semilleros': total_semilleros,
        'proyectos': proyectos,
        'total_proyectos': total_proyectos,
    }

    return render(request, 'paginas/perfil.html', context)

# VISTA ACTUALIZAR FOTO PERFIL
@login_required
def actualizar_foto(request):
    if request.method == 'POST':
        usuario_id = request.session.get('cedula')
        usuario = Usuario.objects.get(cedula=usuario_id)

        if 'imagen_perfil' in request.FILES:
            # Eliminar la imagen anterior (opcional pero recomendado)
            if usuario.imagen_perfil:
                usuario.imagen_perfil.delete(save=False)
            
            # Guardar la nueva
            usuario.imagen_perfil = request.FILES['imagen_perfil']
            usuario.save()
            messages.success(request, "Imagen actualizada correctamente.")
        else:
            messages.error(request, "No se seleccionó ninguna imagen.")
        
        return redirect('perfil')  # o a la vista donde se muestra el perfil

# VISTAS SEMILLEROS
@login_required
def semilleros(request):
    # OBTENER USUARIO DESDE LA SESIÓN
    cedula = request.session.get('cedula')  
    
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    semilleros = usuario.semilleros.all()

    # Búsqueda por texto
    buscar = request.GET.get("buscar")
    if buscar:
        semilleros = semilleros.filter(
            Q(cod_sem__icontains=buscar) |
            Q(sigla__icontains=buscar) |
            Q(nombre__icontains=buscar)
        )

    # Filtro por rango de fechas
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        fecha_inicio_parsed = parse_date(fecha_inicio)
        if fecha_inicio_parsed:
            semilleros = semilleros.filter(fecha_creacion__gte=fecha_inicio_parsed)
    
    if fecha_fin:
        fecha_fin_parsed = parse_date(fecha_fin)
        if fecha_fin_parsed:
            semilleros = semilleros.filter(fecha_creacion__lte=fecha_fin_parsed)

    # Ordenamiento
    ordenar = request.GET.get("ordenar", "")
    if ordenar == "recientes":
        semilleros = semilleros.order_by("-fecha_creacion")
    elif ordenar == "antiguos":
        semilleros = semilleros.order_by("fecha_creacion")
    elif ordenar == "az":
        semilleros = semilleros.order_by("nombre")
    elif ordenar == "za":
        semilleros = semilleros.order_by("-nombre")
    
    # Cálculos adicionales para cada semillero
    for semillero in semilleros:
        # Calcular y actualizar progreso
        actualizar_progreso_semillero(semillero)
        
        # Resto del código existente...
        cedulas = SemilleroUsuario.objects.filter(
            id_sem=semillero
        ).values_list('cedula', flat=True)
        
        total_usuarios = Usuario.objects.filter(cedula__in=cedulas).count()
        total_aprendices = Aprendiz.objects.filter(id_sem=semillero).count()
        semillero.total_miembros = total_usuarios + total_aprendices

        proyectos = SemilleroProyecto.objects.filter(id_sem=semillero)
        total_proyectos = proyectos.count()
        semillero.total_proyectos = total_proyectos

        total_entregables = Entregable.objects.filter(
            cod_pro__in=proyectos.values('cod_pro')
        ).count()
        semillero.total_entregables = total_entregables
        
    return render(request, 'paginas/semilleros.html', {
        'semilleros': semilleros,
        'usuario': usuario
    })

@login_required
def crear_semillero(request):
    cedula = request.session.get('cedula')
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    if request.method == 'POST':
        cod_sem = request.POST.get('cod_sem')  
        sigla = request.POST.get('sigla')
        nombre = request.POST.get('nombre')
        desc_sem = request.POST.get('desc_sem')
        objetivo = request.POST.getlist('objetivo')

        objetivo_texto = "\n".join(objetivo)

        if not all([cod_sem, sigla, nombre, desc_sem, objetivo_texto]):
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('semilleros')

        try:
            # Crear el semillero
            semillero = Semillero(
                cod_sem=cod_sem,
                sigla=sigla,
                nombre=nombre,
                desc_sem=desc_sem,
                objetivo=objetivo_texto,
                estado='Activo',
                progreso_sem=0
            )
            semillero.save()

            # Registrar al usuario creador como miembro y líder
            SemilleroUsuario.objects.create(
                id_sem=semillero,
                cedula=usuario,
                es_lider=True
            )

            messages.success(request, f'Semillero "{sigla}" creado exitosamente')
            return redirect('semilleros')

        except Exception as e:
            messages.error(request, f'Error al crear semillero: {str(e)}')
            return redirect('semilleros')

    return redirect('semilleros')

@login_required
def eliminar_semilleros(request):
    cedula = request.session.get('cedula')
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    if request.method == 'POST':
        semilleros_ids = request.POST.getlist('semilleros_eliminar')
        
        if not semilleros_ids:
            messages.warning(request, 'No se seleccionó ningún semillero para eliminar')
            return redirect('semilleros')
        
        try:
            eliminados = 0
            errores = []
            
            for id_sem in semilleros_ids:
                try:
                    semillero = Semillero.objects.get(id_sem=id_sem)
                    
                    # 1. Restaurar roles de líderes de semillero
                    lideres_sem = SemilleroUsuario.objects.filter(
                        id_sem=semillero, 
                        es_lider=True
                    ).select_related('cedula')
                    
                    for relacion_lider in lideres_sem:
                        usuario_lider = relacion_lider.cedula
                        if usuario_lider.rol == 'Líder de Semillero':
                            # Verificar si es líder en otros semilleros
                            otros_semilleros_lider = SemilleroUsuario.objects.filter(
                                cedula=usuario_lider, 
                                es_lider=True
                            ).exclude(id_sem=semillero).exists()
                            
                            if not otros_semilleros_lider:
                                # Restaurar rol original
                                if hasattr(usuario_lider, 'rol_original') and usuario_lider.rol_original:
                                    usuario_lider.rol = usuario_lider.rol_original
                                elif usuario_lider.vinculacion_laboral and 'instructor' in usuario_lider.vinculacion_laboral.lower():
                                    usuario_lider.rol = 'Instructor'
                                else:
                                    usuario_lider.rol = 'Investigador'
                                usuario_lider.save()
                    
                    # 2. Obtener todos los proyectos del semillero
                    proyectos = Proyecto.objects.filter(
                        semilleroproyecto__id_sem=semillero
                    )
                    
                    for proyecto in proyectos:
                        # Restaurar roles de líderes de proyecto
                        lideres_proyecto = UsuarioProyecto.objects.filter(
                            cod_pro=proyecto, 
                            es_lider_pro=True
                        ).select_related('cedula')
                        
                        for relacion_lider in lideres_proyecto:
                            usuario_lider = relacion_lider.cedula
                            if usuario_lider.rol == 'Líder de Proyecto':
                                # Verificar si es líder en otros proyectos
                                otros_proyectos_lider = UsuarioProyecto.objects.filter(
                                    cedula=usuario_lider, 
                                    es_lider_pro=True
                                ).exclude(cod_pro=proyecto).exists()
                                
                                if not otros_proyectos_lider:
                                    # Restaurar rol
                                    if hasattr(usuario_lider, 'rol_original') and usuario_lider.rol_original:
                                        usuario_lider.rol = usuario_lider.rol_original
                                    else:
                                        es_lider_sem = SemilleroUsuario.objects.filter(
                                            cedula=usuario_lider, 
                                            es_lider=True
                                        ).exists()
                                        if es_lider_sem:
                                            usuario_lider.rol = 'Líder de Semillero'
                                        elif usuario_lider.vinculacion_laboral and 'instructor' in usuario_lider.vinculacion_laboral.lower():
                                            usuario_lider.rol = 'Instructor'
                                        else:
                                            usuario_lider.rol = 'Investigador'
                                    usuario_lider.save()
                        
                        # Eliminar entregables y archivos
                        entregables = Entregable.objects.filter(cod_pro=proyecto)
                        for entregable in entregables:
                            archivos = Archivo.objects.filter(entregable=entregable)
                            for archivo in archivos:
                                if archivo.archivo:
                                    archivo.archivo.delete(save=False)
                            archivos.delete()
                        entregables.delete()
                        
                        # Eliminar relaciones con usuarios y aprendices
                        UsuarioProyecto.objects.filter(cod_pro=proyecto).delete()
                        ProyectoAprendiz.objects.filter(cod_pro=proyecto).delete()
                        
                        # Eliminar relación proyecto-semillero
                        SemilleroProyecto.objects.filter(cod_pro=proyecto).delete()
                        
                        # Eliminar el proyecto
                        proyecto.delete()
                    
                    # 3. Eliminar documentos del semillero
                    documentos_semillero = SemilleroDocumento.objects.filter(id_sem=semillero)
                    for rel_doc in documentos_semillero:
                        documento = rel_doc.cod_doc
                        if documento.archivo:
                            documento.archivo.delete(save=False)
                        documento.delete()
                    documentos_semillero.delete()
                    
                    # 4. Eliminar relaciones con usuarios
                    SemilleroUsuario.objects.filter(id_sem=semillero).delete()
                    
                    # 5. Actualizar aprendices (quitar referencia al semillero)
                    Aprendiz.objects.filter(id_sem=semillero).delete()
                    
                    # 6. Eliminar el semillero
                    semillero.delete()
                    
                    eliminados += 1
                    
                except Semillero.DoesNotExist:
                    errores.append(f'Semillero con ID {id_sem} no encontrado')
                except Exception as e:
                    errores.append(f'Error al eliminar semillero {id_sem}: {str(e)}')
            
            # Mostrar mensajes
            if eliminados > 0:
                messages.success(
                    request, 
                    f'✅ {eliminados} semillero(s) eliminado(s) correctamente y roles restaurados.'
                )
            
            if errores:
                for error in errores:
                    messages.error(request, f'⚠️ {error}')
            
            return redirect('semilleros')
            
        except Exception as e:
            messages.error(request, f'⚠️ Error general al eliminar semilleros: {str(e)}')
            return redirect('semilleros')
    
    # Si no es POST, redirigir
    return redirect('semilleros')

def actualizar_progreso_semillero(semillero):
    # Obtener proyectos del semillero
    proyectos = Proyecto.objects.filter(semilleroproyecto__id_sem=semillero)
    total_proyectos = proyectos.count()

    if total_proyectos == 0:
        semillero.progreso_sem = 0
    else:
        promedio = proyectos.aggregate(Avg('progreso'))['progreso__avg']
        semillero.progreso_sem = round(promedio) if promedio else 0

    semillero.save(update_fields=['progreso_sem'])
    return semillero.progreso_sem

@login_required
def resumen(request, id_sem):
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión.")
        return redirect('iniciarsesion')

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    semillero = get_object_or_404(Semillero, id_sem=id_sem)

    # Objetivos en lista
    objetivos_lista = []
    if semillero.objetivo:
        objetivos_lista = [o.strip() for o in semillero.objetivo.split("\n") if o.strip()]

    # Conteos
    cedulas = SemilleroUsuario.objects.filter(id_sem=semillero).values_list('cedula', flat=True)
    total_usuarios = Usuario.objects.filter(cedula__in=cedulas).count()
    total_aprendices = Aprendiz.objects.filter(id_sem=semillero).count()
    total_miembros = total_usuarios + total_aprendices

    proyectos = SemilleroProyecto.objects.filter(id_sem=semillero)
    total_proyectos = proyectos.count()

    total_entregables = Entregable.objects.filter(
        cod_pro__in=proyectos.values('cod_pro')
    ).count()

    # ACTUALIZAR progreso — con request
    actualizar_progreso_semillero(semillero)

    return render(request, 'paginas/resumen.html', {
        'current_page': 'resumen',
        'current_page_name': 'Semilleros',
        'semillero': semillero,
        'objetivos_lista': objetivos_lista,
        'total_miembros': total_miembros,
        'total_proyectos': total_proyectos,
        'total_entregables': total_entregables,
        'usuario': usuario
    })

@login_required
def resu_miembros(request, id_sem):
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión.")
        return redirect('iniciarsesion')

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion') 
    
    semillero = get_object_or_404(Semillero, id_sem=id_sem)

    cedulas_en_semillero = SemilleroUsuario.objects.filter(
        id_sem=semillero
    ).values_list('cedula', flat=True)
    
    # Instructores
    usuarios = Usuario.objects.filter(cedula__in=cedulas_en_semillero)
    usuarios_disponibles = Usuario.objects.exclude(cedula__in=cedulas_en_semillero)
    
    # Aprendices
    aprendices = Aprendiz.objects.filter(id_sem=semillero)

    total_miembros = usuarios.count() + aprendices.count()

    # Ordenar miembros: primero el líder de semillero, luego líderes de proyecto, luego los demás
    miembros = SemilleroUsuario.objects.filter(
        id_sem=semillero
    ).select_related('cedula').order_by('-es_lider', 'cedula__nom_usu')

    # Obtener proyectos del semillero
    proyectos = Proyecto.objects.filter(semilleroproyecto__id_sem=semillero)
    codigos_proyectos = proyectos.values_list('cod_pro', flat=True)

    # CRÍTICO: Agregar verificación de liderazgo de proyecto a TODOS los miembros
    for miembro in miembros:
        # Verificar si este usuario es líder de algún proyecto del semillero
        miembro.es_lider_proyecto = UsuarioProyecto.objects.filter(
            cedula=miembro.cedula,
            cod_pro__in=codigos_proyectos,
            es_lider_pro=True
        ).exists()
        
        # Obtener nombres de los proyectos que lidera (para mostrar tooltip)
        if miembro.es_lider_proyecto:
            proyectos_liderados = UsuarioProyecto.objects.filter(
                cedula=miembro.cedula,
                cod_pro__in=codigos_proyectos,
                es_lider_pro=True
            ).select_related('cod_pro').values_list('cod_pro__nom_pro', flat=True)
            miembro.proyectos_liderados = list(proyectos_liderados)

    # Crear lista de instructores filtrada Y con la misma lógica de liderazgo
    instructores = SemilleroUsuario.objects.filter(
        id_sem=semillero
    ).filter(
        Q(cedula__rol__icontains='instructor') |
        Q(cedula__rol__icontains='investigador') |
        Q(cedula__rol__icontains='líder') |
        Q(cedula__rol__icontains='lider')
    ).select_related('cedula').order_by('-es_lider', 'cedula__nom_usu')

    # CRÍTICO: Agregar la misma verificación a la lista de instructores
    for instructor in instructores:
        instructor.es_lider_proyecto = UsuarioProyecto.objects.filter(
            cedula=instructor.cedula,
            cod_pro__in=codigos_proyectos,
            es_lider_pro=True
        ).exists()
        
        if instructor.es_lider_proyecto:
            proyectos_liderados = UsuarioProyecto.objects.filter(
                cedula=instructor.cedula,
                cod_pro__in=codigos_proyectos,
                es_lider_pro=True
            ).select_related('cod_pro').values_list('cod_pro__nom_pro', flat=True)
            instructor.proyectos_liderados = list(proyectos_liderados)

    # Verificar si hay instructores
    tiene_instructores = any(
        m.cedula.rol.lower() in ['instructor', 'investigador', 'líder', 'lider'] 
        for m in miembros
    )

    total_proyectos = proyectos.count()

    # Entregables asociados a esos proyectos
    total_entregables = Entregable.objects.filter(
        cod_pro__in=proyectos.values('cod_pro')
    ).count()

    context = {
        'current_page': 'resu_miembros',
        'current_page_name': 'Semilleros',
        'semillero': semillero,
        'usuarios': usuarios,
        'aprendices': aprendices,
        'miembros': miembros, 
        'Usuarios': usuarios_disponibles,
        'total_miembros': total_miembros,
        'proyectos': proyectos,
        'tiene_instructores': tiene_instructores,
        'total_proyectos': total_proyectos,
        'total_entregables': total_entregables,
        'instructores': instructores, 
        'usuario' : usuario,
    }

    return render(request, 'paginas/resu-miembros.html', context)

@login_required
def agregar_miembros(request, id_sem):
    if request.method == 'POST':
        semillero = get_object_or_404(Semillero, id_sem=id_sem)
        miembros_seleccionados = request.POST.getlist('miembros_seleccionados')
        
        if not miembros_seleccionados:
            messages.warning(request, 'No se seleccionó ningún miembro')
            return redirect('resu-miembros', id_sem=id_sem)
        
        try:
            agregados = 0
            ya_existentes = 0
            
            for cedula in miembros_seleccionados:
                usuario = get_object_or_404(Usuario, cedula=cedula)
                
                existe = SemilleroUsuario.objects.filter(
                    id_sem=semillero, 
                    cedula=usuario
                ).exists()
                
                if not existe:
                    SemilleroUsuario.objects.create(
                        id_sem=semillero, 
                        cedula=usuario
                    )
                    agregados += 1
                else:
                    ya_existentes += 1
            
            if agregados > 0:
                messages.success(request, f'{agregados} miembro(s) agregado(s) exitosamente')
            if ya_existentes > 0:
                messages.info(request, f'{ya_existentes} miembro(s) ya estaban en el semillero')
            
            return redirect('resu-miembros', id_sem=id_sem)
            
        except Exception as e:
            messages.error(request, f'Error al agregar miembros: {str(e)}')
            return redirect('resu-miembros', id_sem=id_sem)
    
    return redirect('resu-miembros', id_sem=id_sem)

@login_required
def asignar_lider_semillero(request, id_sem):
    if request.method == "POST":
        semillero = get_object_or_404(Semillero, id_sem=id_sem)
        id_relacion = request.POST.get("lider_semillero")

        # Verificar existencia segura
        try:
            nueva_relacion_lider = SemilleroUsuario.objects.get(semusu_id=id_relacion, id_sem=semillero)
        except SemilleroUsuario.DoesNotExist:
            messages.error(request, "El usuario seleccionado no pertenece a este semillero.")
            return redirect("resu-miembros", semillero.id_sem)

        # Paso 1: Obtener el líder anterior (si existe)
        try:
            relacion_lider_anterior = SemilleroUsuario.objects.get(id_sem=semillero, es_lider=True)
            usuario_anterior = relacion_lider_anterior.cedula
            
            # Cambiar rol del antiguo líder a "Miembro" o "Instructor"
            if usuario_anterior.rol == 'Líder de Semillero':
                usuario_anterior.rol = 'Instructor'  # o 'Miembro', según tu necesidad
                usuario_anterior.save()
            
            # Quitar liderazgo en la tabla intermedia
            relacion_lider_anterior.es_lider = False
            relacion_lider_anterior.save()
            
        except SemilleroUsuario.DoesNotExist:
            # No había líder anterior, continuar normalmente
            pass

        # Paso 2: Asignar nuevo líder en la tabla intermedia
        nueva_relacion_lider.es_lider = True
        nueva_relacion_lider.save()

        # Paso 3: Cambiar el rol del nuevo líder en la tabla Usuario
        nuevo_usuario_lider = nueva_relacion_lider.cedula
        nuevo_usuario_lider.rol = 'Líder de Semillero'
        nuevo_usuario_lider.save()

        messages.success(
            request, 
            f"{nuevo_usuario_lider.nom_usu} {nuevo_usuario_lider.ape_usu} ha sido asignado como líder del semillero."
        )
        return redirect("resu-miembros", semillero.id_sem)
    
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    
    if request.method == "POST":
        cedula_instructor = request.POST.get("instructor_seleccionado")
        cod_proyecto = request.POST.get("proyecto_seleccionado")

        # Validar que se seleccionaron ambos
        if not cedula_instructor or not cod_proyecto:
            messages.error(request, "Debes seleccionar un instructor y un proyecto.")
            return redirect("resu-miembros", id_sem=id_sem)

        try:
            # Obtener el instructor y el proyecto
            instructor = Usuario.objects.get(cedula=cedula_instructor)
            proyecto = Proyecto.objects.get(cod_pro=cod_proyecto)

            # Verificar que el proyecto pertenece al semillero
            if not SemilleroProyecto.objects.filter(id_sem=semillero, cod_pro=proyecto).exists():
                messages.error(request, "El proyecto seleccionado no pertenece a este semillero.")
                return redirect("resu-miembros", id_sem=id_sem)

            # Verificar que el instructor pertenece al semillero
            if not SemilleroUsuario.objects.filter(id_sem=semillero, cedula=instructor).exists():
                messages.error(request, "El instructor seleccionado no pertenece a este semillero.")
                return redirect("resu-miembros", id_sem=id_sem)

            # Paso 1: Obtener y actualizar el líder anterior del proyecto
            try:
                relacion_anterior = UsuarioProyecto.objects.get(cod_pro=proyecto, es_lider=True)
                usuario_anterior = relacion_anterior.cedula
                
                # Cambiar rol si era "Líder de Proyecto"
                if usuario_anterior.rol == 'Líder de Proyecto':
                    usuario_anterior.rol = 'Instructor'  
                    usuario_anterior.save()
                
                # Quitar liderazgo
                relacion_anterior.es_lider_pro = False
                relacion_anterior.save()
                
            except UsuarioProyecto.DoesNotExist:
                # No había líder anterior
                pass

            # Paso 2: Buscar si ya existe la relación usuario-proyecto
            try:
                relacion = UsuarioProyecto.objects.get(
                    cedula=instructor,
                    cod_pro=proyecto
                )
                # Si existe, actualizar el liderazgo
                relacion.es_lider_pro = True
                relacion.save()
                creada = False
            except UsuarioProyecto.DoesNotExist:
                # Si no existe, crear nueva relación
                relacion = UsuarioProyecto.objects.create(
                    cedula=instructor,
                    cod_pro=proyecto,
                    es_lider=True
                )
                creada = True

            # Paso 3: Actualizar el rol del nuevo líder
            instructor.rol = 'Líder de Proyecto'  
            instructor.save()

            mensaje = "asignado" if creada else "actualizado como"
            messages.success(
                request, 
                f"{instructor.nom_usu} {instructor.ape_usu} ha sido {mensaje} líder del proyecto '{proyecto.nom_pro}'."
            )
            return redirect("resu-miembros", id_sem=id_sem)

        except Usuario.DoesNotExist:
            messages.error(request, "El instructor seleccionado no existe.")
            return redirect("resu-miembros", id_sem=id_sem)
        
        except Proyecto.DoesNotExist:
            messages.error(request, "El proyecto seleccionado no existe.")
            return redirect("resu-miembros", id_sem=id_sem)
        
        except Exception as e:
            messages.error(request, f"Error al asignar líder: {e}")
            return redirect("resu-miembros", id_sem=id_sem)

    # Si es GET, redirigir a resu-miembros (el modal se abre desde allí)
    return redirect("resu-miembros", id_sem=id_sem)

@login_required
def resu_proyectos(request, id_sem, cod_pro=None):
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión.")
        return redirect('iniciarsesion')

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    semillero = get_object_or_404(Semillero, id_sem=id_sem)

    proyecto_editar = None
    mostrar_modal_editar = False
    mostrar_modal_gestionar = False
    lineas_tec_lista = []
    lineas_inv_lista = []
    lineas_sem_lista = []
    miembros_proyecto_actual = []

    mostrar_gestionar = request.GET.get('gestionar_equipo')

    # GUARDAR CAMBIOS DEL PROYECTO
    if request.method == 'POST' and cod_pro:
        proyecto_editar = get_object_or_404(Proyecto, cod_pro=cod_pro)

        try:
            proyecto_editar.nom_pro = request.POST.get('nom_pro', '').strip()
            proyecto_editar.tipo = request.POST.get('tipo', '').strip().lower()
            proyecto_editar.desc_pro = request.POST.get('desc_pro', '').strip()
            proyecto_editar.programa_formacion = request.POST.get('programa_formacion', '').strip() if proyecto_editar.tipo == 'formativo' else None
            # LÍNEAS DESDE EL FORMULARIO
            lineas_tec = request.POST.getlist('lineastec[]')
            lineas_inv = request.POST.getlist('lineasinv[]')
            lineas_sem = request.POST.getlist('lineassem[]')

            # ACTUALIZAR SOLO SI EL FORMULARIO ENVÍA LÍNEAS
            if lineas_tec:
                proyecto_editar.linea_tec = "\n".join([l.strip() for l in lineas_tec if l.strip()])

            if lineas_inv:
                proyecto_editar.linea_inv = "\n".join([l.strip() for l in lineas_inv if l.strip()])

            if lineas_sem:
                proyecto_editar.linea_sem = "\n".join([l.strip() for l in lineas_sem if l.strip()])

            nueva_nota = request.POST.get('notas', '').strip()
            if nueva_nota:
                if proyecto_editar.notas:
                    proyecto_editar.notas += f"\n{nueva_nota}"
                else:
                    proyecto_editar.notas = nueva_nota

            proyecto_editar.save()

            miembros_seleccionados = request.POST.getlist('miembros_proyecto[]')

            for cedula in miembros_seleccionados:

                # Primero verificar si es usuario
                usuario = Usuario.objects.filter(cedula=cedula).first()
                if usuario:
                    ya_existe = UsuarioProyecto.objects.filter(
                        cedula=usuario, 
                        cod_pro=proyecto_editar
                    ).exists()
                    if not ya_existe:
                        UsuarioProyecto.objects.create(
                            cedula=usuario, 
                            cod_pro=proyecto_editar, 
                            estado="activo"
                        )
                    continue
                
                # Sino, intentar como aprendiz
                aprendiz = Aprendiz.objects.filter(cedula_apre=cedula).first()
                if aprendiz:
                    ya_existe = ProyectoAprendiz.objects.filter(
                        cedula_apre=aprendiz, 
                        cod_pro=proyecto_editar
                    ).exists()
                    if not ya_existe:
                        ProyectoAprendiz.objects.create(
                            cedula_apre=aprendiz, 
                            cod_pro=proyecto_editar, 
                            estado="activo"
                        )

            messages.success(request, f'Proyecto "{proyecto_editar.nom_pro}" actualizado correctamente.')
            return redirect('resu-proyectos', id_sem=id_sem)

        except Exception as e:
            messages.error(request, f'Error al actualizar proyecto: {str(e)}')
            return redirect('resu-proyectos', id_sem=id_sem, cod_pro=cod_pro)

    # MODAL GESTIONAR EQUIPO
    if mostrar_gestionar and cod_pro and request.method == 'GET':
        proyecto_editar = get_object_or_404(Proyecto, cod_pro=cod_pro)
        mostrar_modal_gestionar = True

        if not proyecto_editar.estado_pro:
            messages.error(request, '❌ Este proyecto está desactivado y no puede ser editado.')
            return redirect('resu-proyectos', id_sem=id_sem)

        usuarios_proyecto = UsuarioProyecto.objects.filter(cod_pro=proyecto_editar).select_related('cedula')
        aprendices_proyecto = ProyectoAprendiz.objects.filter(cod_pro=proyecto_editar).select_related('cedula_apre')

        miembros_equipo = []

        for up in usuarios_proyecto:
            # Traer si este usuario es líder de semillero
            su = SemilleroUsuario.objects.filter(cedula=up.cedula, id_sem=semillero).first()
            es_lider_sem = su.es_lider if su else False

            miembros_equipo.append({
                'cedula': up.cedula.cedula,
                'nombre_completo': f"{up.cedula.nom_usu} {up.cedula.ape_usu}",
                'nom_usu': up.cedula.nom_usu,
                'ape_usu': up.cedula.ape_usu,
                'email': up.cedula.correo_ins if up.cedula.correo_ins else up.cedula.correo_per,
                'iniciales': up.cedula.get_iniciales,
                'tipo': 'Usuario',
                'rol': up.cedula.rol,
                'estado': up.estado,
                'es_lider': up.es_lider_pro,
                'es_lider_sem': es_lider_sem,
                'usuario': usuario
            })

        for ap in aprendices_proyecto:
            miembros_equipo.append({
                'cedula': ap.cedula_apre.cedula_apre,
                'nombre_completo': f"{ap.cedula_apre.nombre} {ap.cedula_apre.apellido}",
                'nom_usu': ap.cedula_apre.nombre,
                'ape_usu': ap.cedula_apre.apellido,
                'email': ap.cedula_apre.correo_ins if hasattr(ap.cedula_apre, 'correo_ins') and ap.cedula_apre.correo_ins else (ap.cedula_apre.correo_per if hasattr(ap.cedula_apre, 'correo_per') else ''),
                'iniciales': ap.cedula_apre.get_iniciales,
                'tipo': 'Aprendiz',
                'rol': 'Aprendiz',
                'estado': ap.estado,
                'es_lider': False,
            })

        # FILTROS
        busqueda = request.GET.get('busqueda_usuario', '').strip().lower()
        filtro_rol = request.GET.get('filtro_rol', '').strip().lower()
        filtro_estado = request.GET.get('filtro_estado', '').strip().lower()

        if busqueda:
            miembros_equipo = [
                m for m in miembros_equipo
                if busqueda in m['nombre_completo'].lower() or busqueda in m['email'].lower()
            ]

        if filtro_rol:
            if filtro_rol == 'es_lider':
                miembros_equipo = [m for m in miembros_equipo if m.get('es_lider', False)]
            else:
                miembros_equipo = [m for m in miembros_equipo if m['rol'].lower() == filtro_rol]

        if filtro_estado:
            miembros_equipo = [m for m in miembros_equipo if m['estado'].lower() == filtro_estado]

        miembros_proyecto_actual = miembros_equipo

    # MODAL EDITAR
    elif cod_pro and request.method == 'GET' and not mostrar_gestionar:
        proyecto_editar = get_object_or_404(Proyecto, cod_pro=cod_pro)
        mostrar_modal_editar = True

        if not proyecto_editar.estado_pro:
            messages.error(request, ' Este proyecto está desactivado y no puede ser editado.')
            return redirect('resu-proyectos', id_sem=id_sem)

        lineas_tec_lista = []
        lineas_inv_lista = []
        lineas_sem_lista = []

        if proyecto_editar.linea_tec:
            lineas_tec_lista = [l.strip() for l in proyecto_editar.linea_tec.split('\n') if l.strip()]

        if proyecto_editar.linea_inv:
            lineas_inv_lista = [l.strip() for l in proyecto_editar.linea_inv.split('\n') if l.strip()]

        if proyecto_editar.linea_sem:
            lineas_sem_lista = [l.strip() for l in proyecto_editar.linea_sem.split('\n') if l.strip()]

        usuarios_proyecto = UsuarioProyecto.objects.filter(cod_pro=proyecto_editar).select_related('cedula')
        aprendices_proyecto = ProyectoAprendiz.objects.filter(cod_pro=proyecto_editar).select_related('cedula_apre')

        for up in usuarios_proyecto:
            miembros_proyecto_actual.append({
                'cedula': up.cedula.cedula,
                'nombre_completo': f"{up.cedula.nom_usu} {up.cedula.ape_usu}",
                'iniciales': up.cedula.get_iniciales,
                'tipo': 'Usuario',
                'rol': up.cedula.rol,
                'estado': up.estado
            })

        for ap in aprendices_proyecto:
            miembros_proyecto_actual.append({
                'cedula': ap.cedula_apre.cedula_apre,
                'nombre_completo': f"{ap.cedula_apre.nombre} {ap.cedula_apre.apellido}",
                'iniciales': ap.cedula_apre.get_iniciales,
                'tipo': 'Aprendiz',
                'rol': 'Aprendiz',
                'estado': ap.estado
            })

    proyectos = Proyecto.objects.filter(semilleroproyecto__id_sem=semillero)

    for proyecto in proyectos:
        verificar_y_actualizar_estados_entregables(proyecto)

    tipo_seleccionado = None
    if cod_pro:
        proyectos = proyectos.order_by(
            Case(
                When(cod_pro=cod_pro, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )

        proyecto_sel = proyectos.filter(cod_pro=cod_pro).first()
        if proyecto_sel:
            tipo_seleccionado = proyecto_sel.tipo.lower()

    proyectos_sennova = list(proyectos.filter(tipo__iexact="sennova"))
    proyectos_capacidad = list(proyectos.filter(tipo__iexact="capacidad instalada"))
    proyectos_formativos = list(proyectos.filter(tipo__iexact="formativo"))

    for proyecto in proyectos_sennova + proyectos_capacidad + proyectos_formativos:
        usuarios_proyecto = UsuarioProyecto.objects.filter(cod_pro=proyecto).select_related('cedula')
        aprendices_proyecto = ProyectoAprendiz.objects.filter(cod_pro=proyecto).select_related('cedula_apre')

        proyecto.lineas_tec_lista = [l.strip() for l in proyecto.linea_tec.split('\n') if l.strip()] if proyecto.linea_tec else []
        proyecto.lineas_inv_lista = [l.strip() for l in proyecto.linea_inv.split('\n') if l.strip()] if proyecto.linea_inv else []
        proyecto.lineas_sem_lista = [l.strip() for l in proyecto.linea_sem.split('\n') if l.strip()] if proyecto.linea_sem else []
        
        miembros_lista = []

        for up in usuarios_proyecto:
            miembros_lista.append({
                'cedula': up.cedula.cedula,
                'nombre_completo': f"{up.cedula.nom_usu} {up.cedula.ape_usu}",
                'iniciales': up.cedula.get_iniciales,
                'tipo': 'Usuario',
                'rol': up.cedula.rol,
                'estado': up.estado
            })

        for ap in aprendices_proyecto:
            miembros_lista.append({
                'cedula': ap.cedula_apre.cedula_apre,
                'nombre_completo': f"{ap.cedula_apre.nombre} {ap.cedula_apre.apellido}",
                'iniciales': ap.cedula_apre.get_iniciales,
                'tipo': 'Aprendiz',
                'rol': 'Aprendiz',
                'estado': ap.estado
            })

        proyecto.miembros_lista = miembros_lista
        proyecto.miembros_activos = [m for m in miembros_lista if m['estado'] == 'activo']
        
        proyecto.lineas_tec_lista = [l.strip() for l in proyecto.linea_tec.split('\n') if l.strip()] if proyecto.linea_tec else []
        proyecto.lineas_inv_lista = [l.strip() for l in proyecto.linea_inv.split('\n') if l.strip()] if proyecto.linea_inv else []
        proyecto.lineas_sem_lista = [l.strip() for l in proyecto.linea_sem.split('\n') if l.strip()] if proyecto.linea_sem else []

    proyectos_count = SemilleroProyecto.objects.filter(id_sem=semillero)
    total_proyectos = proyectos_count.count()

    usuarios = SemilleroUsuario.objects.filter(id_sem=semillero)
    aprendices = Aprendiz.objects.filter(id_sem=semillero)
    total_miembros = usuarios.count() + aprendices.count()
    miembros = usuarios.select_related('cedula')

    total_entregables = Entregable.objects.filter(
        cod_pro__in=proyectos_count.values('cod_pro')
    ).count()

    proyectos_semillero = proyectos_count.values_list('cod_pro', flat=True)
    usuarios_asignados = UsuarioProyecto.objects.filter(
        cod_pro__in=proyectos_semillero
    ).values_list('cedula__cedula', flat=True)
    aprendices_asignados = ProyectoAprendiz.objects.filter(
        cod_pro__in=proyectos_semillero
    ).values_list('cedula_apre__cedula_apre', flat=True)

    cedulas_asignadas = set(str(c) for c in usuarios_asignados) | set(str(c) for c in aprendices_asignados)

    usuarios_semillero = SemilleroUsuario.objects.filter(id_sem=semillero).select_related('cedula')
    usuarios_disponibles = [u for u in usuarios_semillero if str(u.cedula.cedula) not in cedulas_asignadas]

    aprendices_semillero = Aprendiz.objects.filter(id_sem=semillero)
    aprendices_disponibles = [a for a in aprendices_semillero if str(a.cedula_apre) not in cedulas_asignadas]

    miembros_semillero = []
    for u in usuarios_disponibles:
        miembros_semillero.append({
            'cedula': u.cedula.cedula,
            'nombre_completo': f"{u.cedula.nom_usu} {u.cedula.ape_usu}",
            'iniciales': u.cedula.get_iniciales,
            'tipo': 'Usuario',
            'rol': u.cedula.rol
        })

    for a in aprendices_disponibles:
        miembros_semillero.append({
            'cedula': a.cedula_apre,
            'nombre_completo': f"{a.nombre} {a.apellido}",
            'iniciales': a.get_iniciales,
            'tipo': 'Aprendiz',
            'rol': 'Aprendiz'
        })

    context = {
        'current_page': 'resu_proyectos',
        'current_page_name': 'Semilleros',
        'semillero': semillero,
        'proyectos_sennova': proyectos_sennova,
        'proyectos_capacidad': proyectos_capacidad,
        'proyectos_formativos': proyectos_formativos,
        'total_proyectos': total_proyectos,
        'total_miembros': total_miembros,
        'miembros': miembros,
        'total_entregables': total_entregables,
        'miembros_semillero': miembros_semillero,
        'proyecto_seleccionado': cod_pro,
        'tipo_seleccionado': tipo_seleccionado,
        'miembros_proyecto_actual': miembros_proyecto_actual,
        'mostrar_modal_editar': mostrar_modal_editar,
        'mostrar_modal_gestionar': mostrar_modal_gestionar,
        'proyecto_editar': proyecto_editar,
        'lineas_tec_lista': lineas_tec_lista,
        'lineas_inv_lista': lineas_inv_lista,
        'lineas_sem_lista': lineas_sem_lista,
        'busqueda_usuario': request.GET.get('busqueda_usuario', ''),
        'filtro_rol': request.GET.get('filtro_rol', ''),
        'filtro_estado': request.GET.get('filtro_estado', ''),
        'usuario' : usuario
    }

    return render(request, 'paginas/resu-proyectos.html', context)

@login_required
def asignar_lider_proyecto_ajax(request, id_sem, cod_pro):
    """
    Asigna o quita el rol de líder de proyecto mediante AJAX
    """
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    proyecto = get_object_or_404(Proyecto, cod_pro=cod_pro)
    cedula_miembro = request.POST.get('cedula_miembro')
    
    if not cedula_miembro:
        return JsonResponse({'success': False, 'error': 'No se especificó el miembro'})
    
    try:
        # Verificar que el proyecto pertenece al semillero
        if not SemilleroProyecto.objects.filter(id_sem=semillero, cod_pro=proyecto).exists():
            return JsonResponse({'success': False, 'error': 'El proyecto no pertenece a este semillero'})
        
        # Intentar obtener el miembro (solo usuarios pueden ser líderes)
        try:
            miembro = Usuario.objects.get(cedula=cedula_miembro)
        except Usuario.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Solo los usuarios pueden ser líderes de proyecto'})
        
        # Verificar que el usuario pertenece al semillero
        if not SemilleroUsuario.objects.filter(id_sem=semillero, cedula=miembro).exists():
            return JsonResponse({'success': False, 'error': 'El usuario no pertenece a este semillero'})
        
        # Verificar si el usuario ya está en el proyecto
        try:
            relacion_actual = UsuarioProyecto.objects.get(cedula=miembro, cod_pro=proyecto)
            es_lider_actual = relacion_actual.es_lider_pro
        except UsuarioProyecto.DoesNotExist:
            # Si no existe la relación, crearla
            relacion_actual = UsuarioProyecto.objects.create(
                cedula=miembro,
                cod_pro=proyecto,
                es_lider_pro=False,
                estado='activo'
            )
            es_lider_actual = False
        
        # Si ya es líder, informar
        if es_lider_actual:
            return JsonResponse({
                'success': True,
                'es_lider_pro': True,
                'mensaje': f'{miembro.nom_usu} {miembro.ape_usu} ya es líder de este proyecto'
            })
        
        # Quitar liderazgo al líder anterior y restaurar su rol
        try:
            relacion_anterior = UsuarioProyecto.objects.get(cod_pro=proyecto, es_lider_pro=True)
            usuario_anterior = relacion_anterior.cedula
            
            # Guardar el rol original ANTES de cambiar a líder (si no está guardado)
            if not hasattr(usuario_anterior, 'rol_original') or not usuario_anterior.rol_original:
                # Si no existe el campo, buscar el rol en el perfil base
                rol_original = usuario_anterior.rol
                if rol_original == 'Líder de Proyecto':
                    # Si ya era líder, buscar en otros proyectos o usar Instructor/Investigador
                    rol_original = 'Instructor'  # valor temporal
            
            # Verificar si el rol actual es "Líder de Proyecto" antes de cambiarlo
            if usuario_anterior.rol == 'Líder de Proyecto':
                # Buscar otros proyectos donde también sea líder
                otros_proyectos_lider = UsuarioProyecto.objects.filter(
                    cedula=usuario_anterior,
                    es_lider_pro=True
                ).exclude(cod_pro=proyecto).exists()
                
                # Solo cambiar rol si NO es líder en otros proyectos
                if not otros_proyectos_lider:
                    # Buscar el rol original guardado o determinarlo
                    if hasattr(usuario_anterior, 'rol_original') and usuario_anterior.rol_original:
                        usuario_anterior.rol = usuario_anterior.rol_original
                    else:
                        # Determinar basado en el contexto del usuario
                        # Verificar si tiene vinculación laboral como instructor
                        if usuario_anterior.vinculacion_laboral and 'instructor' in usuario_anterior.vinculacion_laboral.lower():
                            usuario_anterior.rol = 'Instructor'
                        else:
                            usuario_anterior.rol = 'Investigador'
                    
                    usuario_anterior.save()
            
            # Quitar liderazgo del proyecto
            relacion_anterior.es_lider_pro = False
            relacion_anterior.save()
            
        except UsuarioProyecto.DoesNotExist:
            pass
        
        # Guardar el rol original del nuevo líder ANTES de cambiarlo
        if miembro.rol != 'Líder de Proyecto':
            if hasattr(miembro, 'rol_original'):
                miembro.rol_original = miembro.rol
        
        # Asignar nuevo líder
        relacion_actual.es_lider_pro = True
        relacion_actual.save()
        
        # Actualizar rol del nuevo líder
        if miembro.rol != 'Líder de Proyecto':
            miembro.rol = 'Líder de Proyecto'
            miembro.save()
        
        return JsonResponse({
            'success': True,
            'es_lider_pro': True,
            'mensaje': f'{miembro.nom_usu} {miembro.ape_usu} ha sido asignado como líder del proyecto'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def alternar_estado_miembro(request, id_sem, cod_pro):
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    proyecto = get_object_or_404(Proyecto, cod_pro=cod_pro)
    cedula_miembro = request.POST.get('cedula_miembro')

    if not cedula_miembro:
        return JsonResponse({'success': False, 'error': 'No se especificÃ³ el miembro'})

    try:
        # Verificar si es usuario o aprendiz
        try:
            relacion = UsuarioProyecto.objects.get(cedula__cedula=cedula_miembro, cod_pro=proyecto)
            
            # ðŸ‘‡ AGREGAR ESTA VALIDACIÃ“N
            # Verificar si es lÃ­der antes de cambiar a inactivo
            if relacion.estado == "activo" and relacion.es_lider_pro:
                return JsonResponse({
                    'success': False, 
                    'error': 'No se puede desactivar al lÃ­der del proyecto. Primero asigne otro lÃ­der.'
                })
            
        except UsuarioProyecto.DoesNotExist:
            relacion = ProyectoAprendiz.objects.get(cedula_apre__cedula_apre=cedula_miembro, cod_pro=proyecto)

        # Cambiar estado
        relacion.estado = "inactivo" if relacion.estado == "activo" else "activo"
        relacion.save()

        return JsonResponse({
            'success': True,
            'nuevo_estado': relacion.estado,
            'mensaje': f'Estado actualizado a {relacion.estado}'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def crear_proyecto(request, id_sem):
    semillero = get_object_or_404(Semillero, id_sem=id_sem)

    # Obtener todos los proyectos del semillero
    proyectos_semillero = SemilleroProyecto.objects.filter(
        id_sem=semillero
    ).values_list('cod_pro', flat=True)

    # Obtener cédulas de usuarios asignados a esos proyectos
    usuarios_asignados = UsuarioProyecto.objects.filter(
        cod_pro__in=proyectos_semillero
    ).values_list('cedula__cedula', flat=True)

    # Obtener cédulas de aprendices asignados a esos proyectos
    aprendices_asignados = ProyectoAprendiz.objects.filter(
        cod_pro__in=proyectos_semillero
    ).values_list('cedula_apre__cedula_apre', flat=True)

    cedulas_asignadas = set(str(c) for c in usuarios_asignados) | set(str(c) for c in aprendices_asignados)

    # Filtrar usuarios del semillero que NO están asignados a ningún proyecto
    usuarios_semillero = SemilleroUsuario.objects.filter(id_sem=semillero).select_related('cedula')

    usuarios_disponibles = [u for u in usuarios_semillero if str(u.cedula.cedula) not in cedulas_asignadas]

    # Filtrar aprendices del semillero que NO están asignados a ningún proyecto
    aprendices_semillero = Aprendiz.objects.filter(id_sem=semillero)

    aprendices_disponibles = [a for a in aprendices_semillero if str(a.cedula_apre) not in cedulas_asignadas]

    # Combinar los miembros disponibles para pasar al template
    miembros_semillero = []

    for u in usuarios_disponibles:
        miembros_semillero.append({
            'cedula': u.cedula.cedula,
            'nombre_completo': f"{u.cedula.nom_usu} {u.cedula.ape_usu}",
            'iniciales': u.cedula.get_iniciales,
            'tipo': 'Usuario',
            'rol': u.cedula.rol
        })

    for a in aprendices_disponibles:
        miembros_semillero.append({
            'cedula': a.cedula_apre,
            'nombre_completo': f"{a.nombre} {a.apellido}",
            'iniciales': a.get_iniciales,
            'tipo': 'Aprendiz',
            'rol': 'Aprendiz'
        })

    # --- Crear proyecto ---
    if request.method == 'POST':
        try:
            cedula_usuario = request.session.get('cedula')
            if not cedula_usuario:
                messages.error(request, 'Debes iniciar sesión para crear un proyecto.')
                return redirect('iniciarsesion')

            usuario_actual = Usuario.objects.get(cedula=cedula_usuario)

            nom_pro = request.POST.get('nom_pro', '').strip()
            tipo = request.POST.get('tipo', '').strip().lower()
            programa_formacion = request.POST.get('programa_formacion', '').strip()

            desc_pro = request.POST.get('desc_pro', '').strip()
            lineas_tec = [l.strip() for l in request.POST.getlist('lineastec[]') if l.strip()]
            lineas_inv = [l.strip() for l in request.POST.getlist('lineasinv[]') if l.strip()]
            lineas_sem = [l.strip() for l in request.POST.getlist('lineassem[]') if l.strip()]
            miembros_seleccionados = request.POST.getlist('miembros_proyecto[]')

            if not all([nom_pro, tipo, desc_pro]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return redirect('resu-proyectos', id_sem=id_sem)
            # Validación específica
            if tipo == "formativo" and not programa_formacion:
                messages.error(request, 'Debe ingresar el programa de formación para un proyecto formativo.')
                return redirect('resu-proyectos', id_sem=id_sem)


            # Crear proyecto
            ultimo_proyecto = Proyecto.objects.order_by('-cod_pro').first()
            nuevo_cod_pro = ultimo_proyecto.cod_pro + 1 if ultimo_proyecto else 1

            proyecto = Proyecto.objects.create(
                cod_pro=nuevo_cod_pro,
                nom_pro=nom_pro,
                tipo=tipo,
                desc_pro=desc_pro,
                linea_tec="\n".join(lineas_tec),
                linea_inv="\n".join(lineas_inv),
                linea_sem="\n".join(lineas_sem),
                estado_pro="Activo",
                programa_formacion=programa_formacion if tipo == "formativo" else None
            )

            # Asociar proyecto al semillero
            SemilleroProyecto.objects.create(id_sem=semillero, cod_pro=proyecto)

            # Asociar usuario creador
            UsuarioProyecto.objects.create(cedula=usuario_actual, cod_pro=proyecto)

            # Asociar miembros seleccionados
            for cedula in miembros_seleccionados:
                if str(cedula) != str(cedula_usuario):
                    try:
                        usuario = Usuario.objects.get(cedula=cedula)
                        UsuarioProyecto.objects.create(cedula=usuario, cod_pro=proyecto)
                    except Usuario.DoesNotExist:
                        try:
                            aprendiz = Aprendiz.objects.get(cedula_apre=cedula)
                            ProyectoAprendiz.objects.create(cedula_apre=aprendiz, cod_pro=proyecto)
                        except Aprendiz.DoesNotExist:
                            pass

            # Crear entregables por defecto con fechas
            entregables_default = [
                {"nombre": "Formalización de Proyecto", "descripcion": "Documento formal del proyecto."},
                {"nombre": "Diagnóstico", "descripcion": "Análisis de la situación actual."},
                {"nombre": "Planeación", "descripcion": "Cronograma y metodología del proyecto."},
                {"nombre": "Ejecución", "descripcion": "Evidencias y resultados del proyecto."},
                {"nombre": "Evaluación", "descripcion": "Cumplimiento e impacto del proyecto."},
                {"nombre": "Conclusiones", "descripcion": "Reflexiones finales y recomendaciones."},
            ]

            ultimo_entregable = Entregable.objects.order_by('-cod_entre').first()
            base_cod = ultimo_entregable.cod_entre if ultimo_entregable else 0

            for i, entregable_data in enumerate(entregables_default, start=1):
                #  Leer fecha_inicio y fecha_fin
                fecha_inicio_str = request.POST.get(f'fecha_inicio_{i}')
                fecha_fin_str = request.POST.get(f'fecha_fin_{i}')

                # Convertir formato de fecha
                fecha_inicio = None
                fecha_fin = None
                
                try:
                    if fecha_inicio_str:
                        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                    if fecha_fin_str:
                        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                except ValueError:
                    # Si el formato es distinto, intenta dd/mm/yyyy
                    try:
                        if fecha_inicio_str:
                            fecha_inicio = datetime.strptime(fecha_inicio_str, '%d/%m/%Y').date()
                        if fecha_fin_str:
                            fecha_fin = datetime.strptime(fecha_fin_str, '%d/%m/%Y').date()
                    except Exception as e:
                        print(f"Error al convertir fechas del entregable {i}: {e}")

                Entregable.objects.create(
                    cod_entre=base_cod + i,
                    nom_entre=entregable_data["nombre"],
                    desc_entre=entregable_data["descripcion"],
                    estado="Pendiente",
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,  
                    cod_pro=proyecto
                )

            messages.success(request, f'Proyecto "{nom_pro}" creado correctamente.')
            return redirect('resu-proyectos', id_sem=id_sem)

        except Exception as e:
            messages.error(request, f'Error al crear proyecto: {str(e)}')
            return redirect('resu-proyectos', id_sem=id_sem)

    # Renderizar formulario
    return render(request, 'crear_proyecto.html', {
        'semillero': semillero,
        'miembros_semillero': miembros_semillero
    })

@login_required
def subir_archivo_entregable(request, id_sem, cod_pro, cod_entre):
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    proyecto = get_object_or_404(Proyecto, cod_pro=cod_pro)
    entregable = get_object_or_404(Entregable, cod_pro=cod_pro, cod_entre=cod_entre)

     # VERIFICAR SI EL PROYECTO ESTÁ ACTIVO
    if not proyecto.estado_pro:
        messages.error(request, "❌ Este proyecto está desactivado. No puedes subir entregables.")
        return redirect('resu-proyectos', id_sem=id_sem)

    if request.method == 'POST':

        # PERMITIMOS VARIOS ARCHIVOS
        archivos = request.FILES.getlist('archivo')

        if not archivos:
            messages.error(request, '⚠️ Debes seleccionar uno o más archivos para subir.')
            return redirect('resu-proyectos', id_sem=id_sem)

        # Guardamos todos los archivos
        for archivo in archivos:
            Archivo.objects.create(
                entregable=entregable,
                archivo=archivo,
                nombre=archivo.name
            )

        fecha_actual = date.today()

        if entregable.fecha_fin:
            if fecha_actual > entregable.fecha_fin:
                entregable.estado = 'Entrega Tardía'
            else:
                entregable.estado = 'Completado'
        else:
            entregable.estado = 'Completado'

        entregable.save()

        # Actualizar progreso del proyecto
        actualizar_progreso_proyecto(request, entregable.cod_pro)

        # Mostrar mensaje con cantidad de archivos
        messages.success(
            request,
            f'✅ {len(archivos)} archivo(s) subido(s) correctamente al entregable "{entregable.nom_entre}".'
        )

        return redirect('resu-proyectos', id_sem=id_sem)

    messages.error(request, 'Método no permitido.')
    return redirect('resu-proyectos', id_sem=id_sem)

def actualizar_progreso_proyecto(proyecto):  
    entregables = Entregable.objects.filter(cod_pro=proyecto)
    total_entregables = entregables.count()

    if total_entregables == 0:
        proyecto.progreso = 0
        proyecto.estado_pro = "pendiente"

    else:
        entregables_completados = entregables.filter(
            estado__in=['Completado', 'Entrega Tardía']
        ).count()

        progreso = round((entregables_completados / total_entregables) * 100)
        proyecto.progreso = progreso

        if entregables_completados == 0:
            proyecto.estado_pro = "pendiente"
        elif entregables_completados == 2:
            proyecto.estado_pro = "diagnostico"
        elif entregables_completados == 3:
            proyecto.estado_pro = "planeacion"
        elif 4 <= entregables_completados <= 5:
            proyecto.estado_pro = "ejecucion"
        elif entregables_completados >= 6:
            proyecto.estado_pro = "completado"

            # ✅ GUARDAR FECHA SOLO LA PRIMERA VEZ
            if not proyecto.fecha_completado:
                proyecto.fecha_completado = date.today()

        else:
            # ✅ Si deja de estar completado, limpiar fecha
            if proyecto.fecha_completado:
                proyecto.fecha_completado = None

    proyecto.save(update_fields=['progreso', 'estado_pro', 'fecha_completado'])
    # Actualizar progreso del semillero asociado
    semilleros = Semillero.objects.filter(semilleroproyecto__cod_pro=proyecto)
    for semillero in semilleros:
        actualizar_progreso_semillero(semillero)

def verificar_y_actualizar_estados_entregables(proyecto):
    from datetime import date
    
    fecha_actual = date.today()
    entregables = Entregable.objects.filter(cod_pro=proyecto)
    
    for entregable in entregables:
        # Verificar si tiene archivos
        tiene_archivos = Archivo.objects.filter(entregable=entregable).exists()
        
        if tiene_archivos:
            # Si tiene archivos, verificar si fue entrega tardía
            if entregable.fecha_fin and fecha_actual > entregable.fecha_fin:
                if entregable.estado not in ['Completado', 'Entrega Tardía']:
                    entregable.estado = 'Entrega Tardía'
                    entregable.save()
            else:
                if entregable.estado != 'Completado':
                    entregable.estado = 'Completado'
                    entregable.save()
        else:
            # No tiene archivos
            if entregable.fecha_fin and fecha_actual > entregable.fecha_fin:
                # Pasó la fecha y no tiene archivos
                if entregable.estado != 'Retrasado':
                    entregable.estado = 'Retrasado'
                    entregable.save()
            else:
                # Aún está en fecha o no hay fecha_fin
                if entregable.estado not in ['Completado', 'Entrega Tardía']:
                    entregable.estado = 'Pendiente'
                    entregable.save()

@login_required
def eliminar_archivo(request, id_sem, cod_pro, cod_entre, id_archivo):
# OBTENER USUARIO DESDE LA SESIÓN
    cedula = request.session.get('cedula')  
    
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    semillero = get_object_or_404(Semillero, id_sem=id_sem)

    proyecto = get_object_or_404(
        Proyecto.objects.filter(
            semilleroproyecto__id_sem=semillero
        ),
        cod_pro=cod_pro
    )

    entregable = get_object_or_404(
        Entregable,
        cod_entre=cod_entre,
        cod_pro=proyecto
    )

    archivo = get_object_or_404(
        Archivo,
        id=id_archivo,
        entregable=entregable
    )

    # VERIFICAR SI EL PROYECTO ESTÁ ACTIVO
    if not proyecto.estado_pro:
        messages.error(request, "❌ El proyecto está desactivado. No puedes eliminar archivos.")
        return redirect('resu-proyectos', id_sem=id_sem)

    try:
        archivo.archivo.delete(save=False)

        archivo.delete()

        if not entregable.archivos.exists():
            entregable.estado = 'Pendiente'
            entregable.save()

        actualizar_progreso_proyecto(request, proyecto)

        messages.success(request, f'Archivo eliminado correctamente')

    except Exception as e:
        messages.error(request, f'Error al eliminar el archivo: {str(e)}')

    return redirect('resu-proyectos', id_sem=id_sem)

@login_required
def eliminar_proyecto_semillero(request, id_sem, cod_pro):
    try:
        semillero = get_object_or_404(Semillero, id_sem=id_sem)
        proyecto = get_object_or_404(Proyecto, cod_pro=cod_pro)

        # Restaurar roles de los líderes de proyecto
        lideres_proyecto = UsuarioProyecto.objects.filter(
            cod_pro=proyecto, es_lider_pro=True
        ).select_related('cedula')

        for relacion_lider in lideres_proyecto:
            usuario_lider = relacion_lider.cedula
            if usuario_lider.rol == 'Líder de Proyecto':
                otros_proyectos_lider = UsuarioProyecto.objects.filter(
                    cedula=usuario_lider, es_lider_pro=True
                ).exclude(cod_pro=proyecto).exists()

                if not otros_proyectos_lider:
                    if hasattr(usuario_lider, 'rol_original') and usuario_lider.rol_original:
                        usuario_lider.rol = usuario_lider.rol_original
                    else:
                        es_lider_semillero = SemilleroUsuario.objects.filter(
                            cedula=usuario_lider, es_lider=True
                        ).exists()
                        if es_lider_semillero:
                            usuario_lider.rol = 'Líder de Semillero'
                        elif usuario_lider.vinculacion_laboral and 'instructor' in usuario_lider.vinculacion_laboral.lower():
                            usuario_lider.rol = 'Instructor'
                        else:
                            usuario_lider.rol = 'Investigador'
                    usuario_lider.save()

        # Eliminar entregables y sus archivos asociados
        entregables = Entregable.objects.filter(cod_pro=proyecto)
        for entregable in entregables:
            archivos = Archivo.objects.filter(entregable=entregable)
            for archivo in archivos:
                archivo.archivo.delete(save=False)  # eliminar físicamente
            archivos.delete()
        entregables.delete()

        # Eliminar relaciones con usuarios y aprendices
        UsuarioProyecto.objects.filter(cod_pro=proyecto).delete()
        ProyectoAprendiz.objects.filter(cod_pro=proyecto).delete()

        # Eliminar relación con semillero
        SemilleroProyecto.objects.filter(id_sem=semillero, cod_pro=proyecto).delete()

        # Eliminar proyecto
        nombre_proyecto = proyecto.nom_pro
        proyecto.delete()

        messages.success(request, f'✅ Proyecto "{nombre_proyecto}" eliminado correctamente y roles restaurados.')

    except Exception as e:
        messages.error(request, f'⚠️ Error al eliminar el proyecto: {e}')

    return redirect('resu-proyectos', id_sem=id_sem)

@login_required
def cambiar_estado_proyecto(request, id_sem, cod_pro):
    # OBTENER USUARIO DESDE LA SESIÓN
    cedula = request.session.get('cedula')  
    
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion') 
    
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    proyecto = get_object_or_404(Proyecto, cod_pro=cod_pro, semilleros=semillero)

    estados_activos = [
        'pendiente',
        'diagnostico',
        'planeacion',
        'ejecucion',
        'completado'
    ]

    if proyecto.estado_pro in estados_activos:
        # DESACTIVAR
        proyecto.estado_original = proyecto.estado_pro
        proyecto.estado_pro = 'desactivado'
        proyecto.activo = False

        messages.warning(
            request,
            f'El proyecto "{proyecto.nom_pro}" fue DESACTIVADO'
        )

    else:
        # ACTIVAR
        proyecto.estado_pro = proyecto.estado_original if proyecto.estado_original else 'pendiente'
        proyecto.estado_anterior = None
        proyecto.activo = True

        messages.success(
            request,
            f'El proyecto "{proyecto.nom_pro}" fue ACTIVADO'
        )

    proyecto.save()
    return redirect('resu-proyectos', semillero.id_sem)

@login_required
def recursos(request, id_sem):
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión para ver tu perfil.")
        return redirect('iniciarsesion')

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion') 
    
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    
    # Obtener todos los documentos del semillero
    relaciones = SemilleroDocumento.objects.filter(id_sem=semillero)
    todos_documentos = Documento.objects.filter(
        cod_doc__in=relaciones.values_list('cod_doc', flat=True)
    )
    
    # Separar por categorías
    documentos = todos_documentos.filter(tipo__in=['Documento', 'Guía'])
    fichas = todos_documentos.filter(tipo__in=['Ficha', 'Acta'])

    # Calcular total de miembros
    cedulas = SemilleroUsuario.objects.filter(
        id_sem=semillero
    ).values_list('cedula', flat=True)
    
    total_usuarios = Usuario.objects.filter(cedula__in=cedulas).count()
    total_aprendices = Aprendiz.objects.filter(id_sem=semillero).count()
    total_miembros = total_usuarios + total_aprendices

    # Proyectos del semillero
    proyectos = SemilleroProyecto.objects.filter(id_sem=semillero)
    total_proyectos = proyectos.count()

    # Entregables asociados a esos proyectos
    total_entregables = Entregable.objects.filter(
        cod_pro__in=proyectos.values('cod_pro')
    ).count()

    
    return render(request, 'paginas/recursos.html', {
        'current_page': 'recursos', 
        'semillero': semillero,
        'current_page_name': 'Semilleros',
        'documentos': documentos,
        'fichas': fichas,
        'total_miembros': total_miembros,
        'total_proyectos': total_proyectos,
        'total_entregables': total_entregables,
        'usuario' : usuario,
    })

@login_required
def agregar_recurso(request, id_sem):
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    
    if request.method == 'POST':
        # Obtener los datos del formulario
        nom_doc = request.POST.get('nom_doc')
        tipo = request.POST.get('tipo')
        archivo = request.FILES.get('archivo')
        
        # Validar que todos los campos requeridos estén presentes
        if not all([nom_doc, tipo, archivo]):
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('recursos', id_sem=id_sem)
                
        # Validar que el archivo sea PDF
        if archivo and not archivo.name.lower().endswith('.pdf'):
            messages.error(request, 'Solo se permiten archivos PDF')
            return redirect('recursos', id_sem=id_sem)
        
        try:
            # Generar el próximo cod_doc
            ultimo_doc = Documento.objects.order_by('-cod_doc').first()
            nuevo_cod_doc = (ultimo_doc.cod_doc + 1) if ultimo_doc else 1
            
            # Obtener fecha actual
            fecha_actual = datetime.now().strftime('%Y-%m-%d')
            
            # Crear el nuevo documento
            documento = Documento(
                cod_doc=nuevo_cod_doc,
                nom_doc=nom_doc,
                tipo=tipo,
                fecha_doc=fecha_actual,
                archivo=archivo
            )
            
            # Guardar el documento en la base de datos
            documento.save()
            
            # Crear la relación con el semillero 
            SemilleroDocumento.objects.create(
                id_sem=semillero,
                cod_doc=documento
            )
            
            messages.success(request, 'Documento guardado exitosamente')
            return redirect('recursos', id_sem=id_sem)
            
        except Exception as e:
            messages.error(request, f'Error al guardar el documento: {str(e)}')
            return redirect('recursos', id_sem=id_sem)
    
    # Si la solicitud no es POST, redirigir a la página de recursos
    return redirect('recursos', id_sem=id_sem)

@login_required
def eliminar_recurso(request, id_sem, cod_doc):
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    documento = get_object_or_404(Documento, cod_doc=cod_doc)

    try:
        # Borrar el archivo físico si existe
        if hasattr(documento, 'archivo') and documento.archivo:
            documento.archivo.delete(save=False)

        # Eliminar la relación con el semillero
        SemilleroDocumento.objects.filter(
            id_sem=semillero, cod_doc=documento
        ).delete()

        # Eliminar el documento en sí
        documento.delete()

        messages.success(request, f'✅ Recurso "{documento.nom_doc}" eliminado correctamente.')
    except Exception as e:
        messages.error(request, f'⚠️ Error al eliminar el recurso: {e}')

    return redirect('recursos', id_sem=id_sem)

# VISTAS DE PROYECTOS
@login_required
def proyectos(request):
    # Obtener cédula de la sesión
    cedula = request.session.get('cedula')
    
    # Obtener el usuario actual
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    # SOLO LOS PROYECTOS EN LOS QUE EL USUARIO ES MIEMBRO O LÍDER
    proyectos = usuario.proyectos.all()
   
    # Obtener la fecha actual y el inicio del mes
    fecha_actual = timezone.now()
    inicio_mes = fecha_actual.replace(day=1)
    
    # Obtener todos los proyectos
    proyectos_list = Proyecto.objects.all().prefetch_related('semilleros')
    
    #  Actualizar estados de TODOS los proyectos primero
    for proyecto in proyectos_list:
        verificar_y_actualizar_estados_entregables(proyecto)
        actualizar_progreso_proyecto(proyecto)
    
    # --- ESTADÍSTICAS GENERALES ---
    total_proyectos = proyectos_list.count()
    
    proyectos_completados = proyectos_list.filter(estado_pro='completado').count()
    proyectos_pendientes = proyectos_list.filter(estado_pro='pendiente').count()
    
    # Proyectos creados este mes
    proyectos_mes = proyectos_list.filter(
        fecha_creacion__gte=inicio_mes
    ).count()
    
    # Proyectos completados este mes
    completados_mes = proyectos_list.filter(
        estado_pro='completado'
    ).count()
    
    # Proyectos pendientes este mes
    pendientes_mes = proyectos_list.filter(
        fecha_creacion__gte=inicio_mes
    ).count()
    
    # --- FILTROS Y BÚSQUEDA ---
    
    # Búsqueda por nombre
    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        proyectos_list = proyectos_list.filter(
            Q(nom_pro__icontains=buscar) |
            Q(semilleros__nombre__icontains=buscar)
        ).distinct()
    
    # Filtro por estado
    estado = request.GET.get('estado', '').strip()
    if estado:
        proyectos_list = proyectos_list.filter(estado_pro=estado)
    
    # Filtro por tipo (si el campo tipo existe en tu modelo)
    tipo = request.GET.get('tipo', '').strip()
    if tipo:
        proyectos_list = proyectos_list.filter(tipo=tipo)
    
    # Ordenamiento
    orden = request.GET.get('orden', '').strip()
    if orden == 'nombre':
        proyectos_list = proyectos_list.order_by('nom_pro')
    elif orden == 'fecha_creacion':
        proyectos_list = proyectos_list.order_by('fecha_creacion')
    else:
        # Ordenamiento por defecto: más recientes primero
        proyectos_list = proyectos_list.order_by('-fecha_creacion')

    # Obtener todos los tipos de proyecto únicos
    tipos_proyecto = Proyecto.objects.values_list('tipo', flat=True).exclude(tipo__isnull=True).exclude(tipo='').distinct().order_by('tipo')
    
    # Contexto para el template
    context = {
        # Estadísticas
        'total_proyectos': total_proyectos,
        'proyectos_completados': proyectos_completados,
        'proyectos_pendientes': proyectos_pendientes,
        'proyectos_mes': proyectos_mes,
        'completados_mes': completados_mes,
        'pendientes_mes': pendientes_mes,
        'proyectos': proyectos_list,
        'tipos_proyecto': tipos_proyecto,
        'usuario' : usuario,
    }
    return render(request, 'paginas/proyectos.html', context)

@login_required
def detalle_proyecto(request, id_sem, cod_pro):
    """Vista para ver los detalles de un proyecto específico"""
    
    # ==================== VALIDACIÓN DE SESIÓN ====================
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión para ver el proyecto.")
        return redirect('iniciarsesion')

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    # ==================== OBTENER SEMILLERO ====================
    semillero = get_object_or_404(Semillero, id_sem=id_sem)
    
    # ==================== OBTENER TODOS LOS PROYECTOS ====================
    proyectos = Proyecto.objects.filter(semilleroproyecto__id_sem=semillero)
    
    tipo_seleccionado = None
    proyecto_sel = None
    
    if cod_pro:
        # Ordenar para que el proyecto seleccionado aparezca primero
        proyectos = proyectos.order_by(
            Case(
                When(cod_pro=cod_pro, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
        
        # Obtener el proyecto seleccionado
        proyecto_sel = proyectos.filter(cod_pro=cod_pro).first()
        
        if proyecto_sel:
            tipo_seleccionado = proyecto_sel.tipo.lower()
    
    # ==================== SEPARAR PROYECTOS POR TIPO ====================
    proyectos_sennova = list(proyectos.filter(tipo__iexact="Sennova"))
    proyectos_capacidad = list(proyectos.filter(tipo__iexact="Capacidad Instalada"))
    proyectos_formativos = list(proyectos.filter(tipo__iexact="Formativo"))
    
    # ==================== PROCESAR MIEMBROS DE CADA PROYECTO ====================
    todos_proyectos = proyectos_sennova + proyectos_capacidad + proyectos_formativos
    
    for proyecto in todos_proyectos:
        usuarios_proyecto = UsuarioProyecto.objects.filter(cod_pro=proyecto).select_related('cedula')
        aprendices_proyecto = ProyectoAprendiz.objects.filter(cod_pro=proyecto).select_related('cedula_apre')
        
        miembros_lista = []
        
        # Agregar usuarios
        for up in usuarios_proyecto:
            miembros_lista.append({
                'cedula': up.cedula.cedula,
                'nombre_completo': f"{up.cedula.nom_usu} {up.cedula.ape_usu}",
                'iniciales': up.cedula.get_iniciales,
                'tipo': 'Usuario',
                'rol': up.cedula.rol,
                'estado': up.estado
            })
        
        # Agregar aprendices
        for ap in aprendices_proyecto:
            miembros_lista.append({
                'cedula': ap.cedula_apre.cedula_apre,
                'nombre_completo': f"{ap.cedula_apre.nombre} {ap.cedula_apre.apellido}",
                'iniciales': ap.cedula_apre.get_iniciales,
                'tipo': 'Aprendiz',
                'rol': 'Aprendiz',
                'estado': ap.estado
            })
        
        proyecto.miembros_lista = miembros_lista
        proyecto.miembros_activos = [m for m in miembros_lista if m['estado'] == 'activo']
        
        # Procesar líneas del proyecto
        proyecto.lineas_tec_lista = [l.strip() for l in proyecto.linea_tec.split('\n') if l.strip()] if proyecto.linea_tec else []
        proyecto.lineas_inv_lista = [l.strip() for l in proyecto.linea_inv.split('\n') if l.strip()] if proyecto.linea_inv else []
        proyecto.lineas_sem_lista = [l.strip() for l in proyecto.linea_sem.split('\n') if l.strip()] if proyecto.linea_sem else []
    
    # ==================== ESTADÍSTICAS DEL SEMILLERO ====================
    proyectos_count = SemilleroProyecto.objects.filter(id_sem=semillero)
    total_proyectos = proyectos_count.count()
    
    usuarios_sem = SemilleroUsuario.objects.filter(id_sem=semillero)
    aprendices_sem = Aprendiz.objects.filter(id_sem=semillero)
    total_miembros = usuarios_sem.count() + aprendices_sem.count()
    miembros = usuarios_sem.select_related('cedula')
    
    total_entregables = Entregable.objects.filter(
        cod_pro__in=proyectos_count.values('cod_pro')
    ).count()
    
    # ==================== CONTEXTO ====================
    context = {
        'current_page': 'detalle_proyectos',
        'current_page_name': 'Detalle Proyecto',
        'usuario': usuario,
        'semillero': semillero,
        'proyectos_sennova': proyectos_sennova,
        'proyectos_capacidad': proyectos_capacidad,
        'proyectos_formativos': proyectos_formativos,
        'proyecto': proyecto_sel,
        'tipo_seleccionado': tipo_seleccionado,
        'cod_pro': cod_pro,
        'total_proyectos': total_proyectos,
        'total_miembros': total_miembros,
        'total_entregables': total_entregables,
        'miembros': miembros,
        'solo_lectura': True,
    }
    
    return render(request, 'paginas/resu-proyectos.html', context)

# VISTAS DE MIEMBROS
@login_required
def miembros(request):
    usuario_id = request.session.get('cedula')
    if not usuario_id:
        messages.error(request, "Debes iniciar sesión para ver tu perfil.")
        return redirect('iniciarsesion')

    # VERIFICAR SI DEBE MOSTRAR EL MODAL
    mostrar_modal_codigo = request.session.get('mostrar_modal_codigo', False)
    aprendiz_verificacion = None

    if mostrar_modal_codigo:
        aprendiz_id = request.session.get('verificacion_aprendiz_id')
        if aprendiz_id:
            aprendiz_verificacion = Aprendiz.objects.filter(cedula_apre=aprendiz_id).first()

    # Procesar verificación si es POST (cuando viene del modal)
    if request.method == 'POST' and 'codigo' in request.POST:
        resultado = verificar_codigo_form(request)
        if resultado:
            return resultado

    try:
        usuario = Usuario.objects.get(cedula=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    vista = request.GET.get('vista', 'tarjeta')
    estado_filtro = request.GET.get('estado', '')
    rol_filtro = request.GET.get('rol', 'todos').strip().lower()
    busqueda = request.GET.get('busqueda', '').strip().lower()
    miembro_id = request.GET.get('miembro_id')

    # VALIDAR QUE miembro_id NO SEA 'None' O VACÍO
    if miembro_id and miembro_id.lower() == 'none':
        miembro_id = None

    if request.method == "POST" and 'verificar_codigo' not in request.POST and 'codigo' not in request.POST:
        cedula = request.POST.get("cedula")
        nuevo_estado = request.POST.get("estado")

        if cedula and nuevo_estado:
            usuario_obj = Usuario.objects.filter(cedula=cedula).first()
            aprendiz_obj = Aprendiz.objects.filter(cedula_apre=cedula).first()

            if usuario_obj:
                usuario_obj.estado = nuevo_estado
                usuario_obj.save()
                messages.success(request, f"Estado de {usuario_obj.nom_usu} actualizado a {nuevo_estado}.")
            elif aprendiz_obj:
                aprendiz_obj.estado_apre = nuevo_estado
                aprendiz_obj.save()
                messages.success(request, f"Estado de {aprendiz_obj.nombre} actualizado a {nuevo_estado}.")

        return redirect(f"{request.path}?miembro_id={cedula}")

    usuarios = Usuario.objects.all()
    aprendices = Aprendiz.objects.all()

    if busqueda:
        usuarios = usuarios.filter(
            Q(nom_usu__icontains=busqueda) |
            Q(ape_usu__icontains=busqueda) |
            Q(cedula__icontains=busqueda)
        )
        aprendices = aprendices.filter(
            Q(nombre__icontains=busqueda) |
            Q(apellido__icontains=busqueda) |
            Q(cedula_apre__icontains=busqueda)
        )

    if estado_filtro and estado_filtro != 'todos':
        usuarios = usuarios.filter(estado=estado_filtro)
        aprendices = aprendices.filter(estado_apre=estado_filtro)

    if rol_filtro == 'investigadores':
        usuarios = usuarios.filter(rol__iexact='Investigador')
        aprendices = aprendices.none()

    elif rol_filtro == 'instructores':
        usuarios = usuarios.filter(rol__iexact='Instructor')
        aprendices = aprendices.none()

    elif rol_filtro == 'lider_semillero':
        usuarios = usuarios.filter(rol__iregex=r'l[ií]der(\s+de)?\s+semillero')
        aprendices = aprendices.none()

    elif rol_filtro == 'dinamizador':
        usuarios = usuarios.filter(rol__iexact='Dinamizador')
        aprendices = aprendices.none()

    elif rol_filtro == 'lider_proyecto':
        usuarios = usuarios.filter(rol__iregex=r'l[ií]der(\s+de)?\s+proyecto')
        aprendices = aprendices.none()

    elif rol_filtro == 'aprendices':
        usuarios = usuarios.none()

    miembros = []

    # ---- Usuarios ----
    for u in usuarios:
        ultimo_acceso = "Sin registro"
        if u.last_login:
            tiempo = now() - u.last_login
            dias, segundos = tiempo.days, tiempo.seconds

            if dias == 0:
                if segundos < 60:
                    ultimo_acceso = "Hace unos segundos"
                elif segundos < 3600:
                    minutos = segundos // 60
                    ultimo_acceso = f"Hace {minutos} minuto{'s' if minutos != 1 else ''}"
                else:
                    horas = segundos // 3600
                    ultimo_acceso = f"Hace {horas} hora{'s' if horas != 1 else ''}"
            elif dias == 1:
                ultimo_acceso = "Ayer"
            elif dias < 7:
                ultimo_acceso = f"Hace {dias} día{'s' if dias != 1 else ''}"
            elif dias < 30:
                semanas = dias // 7
                ultimo_acceso = f"Hace {semanas} semana{'s' if semanas != 1 else ''}"
            elif dias < 365:
                meses = dias // 30
                ultimo_acceso = f"Hace {meses} mes{'es' if meses != 1 else ''}"
            else:
                años = dias // 365
                ultimo_acceso = f"Hace {años} año{'s' if años != 1 else ''}"

        miembros.append({
            'id': u.cedula,
            'nombres': u.nom_usu,
            'apellidos': u.ape_usu,
            'correo_sena': u.correo_ins,
            'celular': u.telefono,
            'rol': u.rol,
            'ultima_sesion': ultimo_acceso,
            'tipo': 'usuario',
            'imagen_perfil': u.imagen_perfil.url if u.imagen_perfil else None,
            'objeto': u
        })

    # ---- Aprendices ----
    for a in aprendices:
        miembros.append({
            'id': a.cedula_apre,
            'nombres': a.nombre,
            'apellidos': a.apellido,
            'correo_sena': a.correo_ins,
            'celular': a.telefono,
            'rol': 'Aprendiz',
            'ultima_sesion': None,
            'tipo': 'aprendiz',
            'objeto': a
        })

    total_instructores = Usuario.objects.filter(rol='Instructor').count()
    total_investigadores = Usuario.objects.filter(rol='Investigador').count()
    total_aprendices = Aprendiz.objects.count()

    miembro_seleccionado = None
    tipo_miembro = None

    # SOLO BUSCAR MIEMBRO SI miembro_id EXISTE Y NO ES 'None'
    if miembro_id:
        usuario_sel = Usuario.objects.filter(cedula=miembro_id).first()
        if usuario_sel:
            tipo_miembro = 'usuario'
            proyectos = SemilleroProyecto.objects.filter(
                cod_pro__usuarioproyecto__cedula=usuario_sel
            ).select_related(
                'id_sem',
                'cod_pro'
            ).distinct()
            miembro_seleccionado = {
                'id': usuario_sel.cedula,
                'cedula': usuario_sel.cedula,
                'nombres': usuario_sel.nom_usu,
                'apellidos': usuario_sel.ape_usu,
                'fecha_nacimiento': usuario_sel.fecha_nacimiento,
                'correo_personal': usuario_sel.correo_per,
                'correo_sena': usuario_sel.correo_ins,
                'celular': usuario_sel.telefono,
                'vinculacion': usuario_sel.vinculacion_laboral,
                'dependencia': usuario_sel.dependencia,
                'rol': usuario_sel.rol,
                'imagen_perfil': usuario_sel.imagen_perfil,
                'estado': usuario_sel.estado,
                'semilleros': usuario_sel.semilleros.all(),
                'proyectos': proyectos,
            }
        else:
            aprendiz_sel = Aprendiz.objects.filter(cedula_apre=miembro_id).first()
            if aprendiz_sel:
                tipo_miembro = 'aprendiz'
                
                #  CAMBIO: Verificar si hay número revelado para ESTE aprendiz específico
                numeros_revelados = request.session.get('numeros_revelados', {})
                info_numero = numeros_revelados.get(str(miembro_id))
                
                if info_numero:
                    # Verificar si han pasado 60 segundos
                    timestamp_revelado = info_numero.get('timestamp', 0)
                    tiempo_transcurrido = timezone.now().timestamp() - timestamp_revelado
                    
                    if tiempo_transcurrido < 60:
                        # Aún no han pasado 60 segundos, mostrar completo
                        numero_visible = info_numero['numero']
                    else:
                        # Ya pasaron 60 segundos, ocultar automáticamente
                        del numeros_revelados[str(miembro_id)]
                        request.session['numeros_revelados'] = numeros_revelados
                        
                        # Mostrar parcialmente
                        numero_descifrado = descifrar_numero(aprendiz_sel.numero_cuenta)
                        if numero_descifrado and numero_descifrado != "****":
                            numero_visible = f"*********{numero_descifrado[-4:]}"
                        else:
                            numero_visible = "**********"
                else:
                    # No hay número revelado, mostrar parcialmente
                    numero_descifrado = descifrar_numero(aprendiz_sel.numero_cuenta)
                    if numero_descifrado and numero_descifrado != "****":
                        numero_visible = f"*********{numero_descifrado[-4:]}"
                    else:
                        numero_visible = "**********"

                    proyectos = SemilleroProyecto.objects.filter(
                        cod_pro__proyectoaprendiz__cedula_apre=aprendiz_sel
                    ).select_related(
                        'id_sem',
                        'cod_pro'
                    ).distinct() 
                miembro_seleccionado = {
                    'id': aprendiz_sel.cedula_apre,
                    'cedula': aprendiz_sel.cedula_apre,
                    'nombres': aprendiz_sel.nombre,
                    'apellidos': aprendiz_sel.apellido,
                    'correo_personal': aprendiz_sel.correo_per,
                    'fecha_nacimiento': aprendiz_sel.fecha_nacimiento,
                    'correo_sena': aprendiz_sel.correo_ins,
                    'medio_bancario': aprendiz_sel.medio_bancario,
                    'numero_cuenta_visible': numero_visible,  
                    'celular': aprendiz_sel.telefono,
                    'ficha': aprendiz_sel.ficha,
                    'programa': aprendiz_sel.programa,
                    'rol': 'Aprendiz',
                    'estado': aprendiz_sel.estado_apre,
                    'semilleros': [aprendiz_sel.id_sem] if aprendiz_sel.id_sem else [],
                    'proyectos': proyectos,
                }

    contexto = {
        'miembros': miembros,
        'total_instructores': total_instructores,
        'total_investigadores': total_investigadores,
        'total_aprendices': total_aprendices,
        'miembro_seleccionado': miembro_seleccionado,
        'tipo_miembro': tipo_miembro,
        'vista': vista,
        'estado_filtro': estado_filtro,
        'rol_filtro': rol_filtro,
        'busqueda': busqueda,
        'usuario': usuario,
        'mostrar_modal_codigo': mostrar_modal_codigo,
        'aprendiz_verificacion': aprendiz_verificacion,
    }

    return render(request, 'paginas/miembros.html', contexto)

def registro_aprendiz(request):  
    if request.method == 'POST':
        form = AprendizForm(request.POST)
        if form.is_valid():
            try:
                aprendiz = form.save(commit=False)
                aprendiz.estado_apre = 'Activo'
                
                # Registrar fecha de aceptación del tratamiento de datos
                if aprendiz.acepta_tratamiento_datos:
                    aprendiz.fecha_aceptacion_datos = timezone.now()
                
                aprendiz.save()
                
                messages.success(request, f'¡Aprendiz {aprendiz.nombre} {aprendiz.apellido} registrado exitosamente!')
                
                # Redirigir después del éxito para evitar re-submit
                return redirect('registro_aprendiz')  # Cambia esto por el nombre de tu URL
                
            except Exception as e:
                messages.error(request, f'Error al guardar: {str(e)}')
        else:
            # Agregar mensaje de error general
            messages.error(request, 'Por favor corrija los errores señalados en el formulario.')
            
            # Crear diccionario con información de errores para JavaScript
            errores_info = {}
            for field_name, errors in form.errors.items():
                if field_name != '__all__':
                    errores_info[field_name] = str(errors[0])
            
            # Pasar errores al contexto
            return render(request, 'paginas/formaprendiz.html', {
                'form': form,
                'errores_info': errores_info
            })
    else:
        form = AprendizForm()
    
    return render(request, 'paginas/formaprendiz.html', {'form': form})

@login_required
def solicitar_codigo_verificacion_form(request, aprendiz_id):
    """Genera y envía un código de verificación con control de límites"""
    try:
        usuario_id = request.session.get('cedula')
        if not usuario_id:
            messages.error(request, 'No hay sesión activa')
            return redirect('miembros')
        
        usuario = Usuario.objects.get(cedula=usuario_id)
        
        # Verificar si puede solicitar código
        puede_solicitar, mensaje_error = usuario.puede_solicitar_codigo()
        
        if not puede_solicitar:
            messages.error(request, mensaje_error)
            return redirect('miembros')
        
        # Verificar que el aprendiz existe
        aprendiz = Aprendiz.objects.filter(cedula_apre=aprendiz_id).first()
        if not aprendiz:
            messages.error(request, 'Aprendiz no encontrado')
            return redirect('miembros')
        
        # Generar código
        codigo = usuario.generar_codigo_verificacion()
        
        # Preparar el correo
        correo_destino = usuario.correo_ins or usuario.correo_per
        
        if not correo_destino:
            messages.error(request, 'No hay correo registrado')
            return redirect('miembros')
        
        # Calcular intentos restantes
        intentos_restantes = max(0, 7 - usuario.intentos_codigo_fallidos)
        
        # Construir mensaje del correo
        asunto = "Código de Verificación - GC"
        mensaje = f"""
        <html>
            <body style="font-family: Arial, sans-serif; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: #f9f9f9; padding: 30px; border-radius: 10px;">
                    <h2 style="color: #2C3E50; text-align: center;">🔐 Código de Verificación</h2>
                    <p style="color: #555; font-size: 16px;">Hola <strong>{usuario.nom_usu}</strong>,</p>
                    <p style="color: #555; font-size: 14px;">
                        Has solicitado ver información sensible. Por seguridad, utiliza el siguiente código:
                    </p>
                    <div style="background-color: #fff; padding: 20px; border-radius: 5px; text-align: center; margin: 20px 0;">
                        <h1 style="color: #27AE60; font-size: 48px; margin: 0; letter-spacing: 10px;">{codigo}</h1>
                    </div>
                    <p style="color: #E74C3C; font-size: 12px; text-align: center;">
                        ⚠️ Este código expirará en 60 segundos
                    </p>
                    <p style="color: #555; font-size: 12px; text-align: center;">
                        Intentos restantes: <strong>{intentos_restantes}</strong>
                    </p>
                    <p style="color: #999; font-size: 12px; text-align: center; margin-top: 30px;">
                        Si no solicitaste este código, ignora este mensaje.
                    </p>
                </div>
            </body>
        </html>
        """
        
        # Enviar correo
        try:
            send_mail(
                asunto,
                '',
                settings.DEFAULT_FROM_EMAIL,
                [correo_destino],
                html_message=mensaje,
                fail_silently=False
            )
            
            # REGISTRAR ENVÍO
            usuario.registrar_codigo_enviado()
            
            # Activar el modal
            request.session['verificacion_aprendiz_id'] = aprendiz_id
            request.session['mostrar_modal_codigo'] = True
            request.session['codigo_enviado'] = True
            
            messages.success(request, f'✅ Código enviado a {correo_destino}.')
            return redirect('miembros')
            
        except Exception as email_error:
            messages.error(request, f'Error al enviar el correo: {str(email_error)}')
            return redirect('miembros')
        
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('miembros')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('miembros')

@login_required
def verificar_codigo_form(request):
    """Verifica el código con control de reintentos"""
    try:
        usuario_id = request.session.get('cedula')
        if not usuario_id:
            messages.error(request, 'No hay sesión activa')
            return redirect('miembros')
        
        usuario = Usuario.objects.get(cedula=usuario_id)
        
        # ✅ Verificar si está bloqueado
        ahora = timezone.now()
        if usuario.bloqueado_hasta and ahora < usuario.bloqueado_hasta:
            tiempo_restante = usuario.bloqueado_hasta - ahora
            minutos = int(tiempo_restante.total_seconds() / 60)
            horas = minutos // 60
            mins = minutos % 60
            
            if horas > 0:
                tiempo_msg = f"{horas} hora(s) y {mins} minuto(s)"
            else:
                tiempo_msg = f"{mins} minuto(s)"
            
            # Cerrar modal y limpiar sesión
            request.session['mostrar_modal_codigo'] = False
            if 'verificacion_aprendiz_id' in request.session:
                del request.session['verificacion_aprendiz_id']
            
            messages.error(request, f'⏱️ Demasiados intentos fallidos. Espera {tiempo_msg} antes de intentar nuevamente.')
            return redirect('miembros')
        
        # Obtener código ingresado
        codigo_ingresado = request.POST.get('codigo', '').strip()
        aprendiz_id = request.session.get('verificacion_aprendiz_id')
        
        if not codigo_ingresado or len(codigo_ingresado) != 6:
            messages.error(request, '❌ Debes ingresar los 6 dígitos del código')
            return redirect(f'/miembros/?miembro_id={aprendiz_id}')
        
        if not aprendiz_id:
            messages.error(request, 'Sesión expirada')
            request.session['mostrar_modal_codigo'] = False
            return redirect('miembros')
        
        # ✅ VERIFICAR CÓDIGO
        if not usuario.verificar_codigo(codigo_ingresado):
            # Código incorrecto
            
            # Incrementar intentos fallidos en el usuario
            usuario.incrementar_intentos_fallidos()
            
            # Calcular intentos restantes
            intentos_restantes = max(0, 7 - usuario.intentos_codigo_fallidos)
            
            # 🆕 Actualizar en sesión (esto hace que aparezca el mensaje en el siguiente intento)
            request.session['intentos_restantes'] = intentos_restantes
            
            # Si se acabaron los intentos, bloquear
            if intentos_restantes == 0:
                tiempo_bloqueo = usuario.bloqueado_hasta - ahora if usuario.bloqueado_hasta else timedelta(minutes=15)
                minutos = int(tiempo_bloqueo.total_seconds() / 60)
                horas = minutos // 60
                mins = minutos % 60
                
                if horas > 0:
                    tiempo_msg = f"{horas} hora(s) y {mins} minuto(s)"
                else:
                    tiempo_msg = f"{mins} minuto(s)"
                
                # Cerrar modal
                request.session['mostrar_modal_codigo'] = False
                if 'verificacion_aprendiz_id' in request.session:
                    del request.session['verificacion_aprendiz_id']
                
                messages.error(request, f'🚫 Demasiados intentos fallidos. Espera {tiempo_msg} antes de intentar nuevamente.')
                return redirect('miembros')
            
            messages.error(request, f'❌ Código incorrecto. Te quedan {intentos_restantes} intentos.')
            return redirect(f'/miembros/?miembro_id={aprendiz_id}')
        
        # ✅ CÓDIGO CORRECTO
        # Obtener aprendiz
        aprendiz = Aprendiz.objects.filter(cedula_apre=aprendiz_id).first()
        if not aprendiz:
            messages.error(request, 'Aprendiz no encontrado')
            return redirect('miembros')
        
        # Descifrar número de cuenta
        numero_completo = descifrar_numero(aprendiz.numero_cuenta)
        
        # Guardar en sesión
        timestamp_actual = timezone.now().timestamp()
        numeros_revelados = request.session.get('numeros_revelados', {})
        numeros_revelados[str(aprendiz_id)] = {
            'numero': numero_completo,
            'timestamp': timestamp_actual
        }
        
        request.session['numeros_revelados'] = numeros_revelados
        request.session['numero_cuenta_aprendiz_id'] = str(aprendiz_id)
        
        # ✅ RESETEAR INTENTOS
        usuario.resetear_intentos_codigo()
        
        # Limpiar sesión del modal
        request.session['mostrar_modal_codigo'] = False
        if 'verificacion_aprendiz_id' in request.session:
            del request.session['verificacion_aprendiz_id']
        if 'codigo_enviado' in request.session:
            del request.session['codigo_enviado']
        if 'intentos_restantes' in request.session:
            del request.session['intentos_restantes']
        
        messages.success(request, '✅ Código verificado correctamente. El número estará visible por 30 segundos.')
        return redirect(f'/miembros/?miembro_id={aprendiz_id}')
        
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('miembros')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('miembros')

@login_required
def cancelar_verificacion(request):
    """Cancela el proceso de verificación"""
    request.session['mostrar_modal_codigo'] = False
    
    if 'verificacion_aprendiz_id' in request.session:
        del request.session['verificacion_aprendiz_id']
    if 'codigo_enviado' in request.session:
        del request.session['codigo_enviado']
    if 'intentos_restantes' in request.session:
        del request.session['intentos_restantes']
    
    if 'numeros_revelados' in request.session:
        del request.session['numeros_revelados']
    if 'numero_cuenta_aprendiz_id' in request.session:
        del request.session['numero_cuenta_aprendiz_id']
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    messages.info(request, 'Verificación cancelada')
    return redirect('miembros')

@login_required
def limpiar_numero_revelado(request):
    """
    Limpia el número de cuenta revelado de la sesión
    """
    if 'numero_cuenta_revelado' in request.session:
        del request.session['numero_cuenta_revelado']
    if 'numero_cuenta_aprendiz_id' in request.session:
        del request.session['numero_cuenta_aprendiz_id']
    
    return redirect('miembros')

# VISTAS DE EVENTOS
@login_required
def eventos(request):
    # Obtener cédula de la sesión
    cedula = request.session.get('cedula')
    
    # Obtener el usuario actual
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    limite = timezone.now() - timedelta(hours=24)
    Evento.objects.filter(
        estado_eve__in=["Cancelado", "Finalizado"],
        fecha_estado__lt=limite
    ).delete()

    ahora = timezone.localtime()
    eventos = Evento.objects.all()

    for evento in eventos:

        if evento.estado_eve.lower() == "cancelado":
            continue

        hora_inicio = datetime.strptime(str(evento.hora_inicio), "%H:%M:%S").time()
        hora_fin = datetime.strptime(str(evento.hora_fin), "%H:%M:%S").time()

        inicio = timezone.make_aware(datetime.combine(evento.fecha_eve, hora_inicio))
        fin = timezone.make_aware(datetime.combine(evento.fecha_eve, hora_fin))

        if ahora > fin:
            if evento.estado_eve != "Finalizado":
                evento.estado_eve = "Finalizado"
                evento.save()
            continue

        if inicio <= ahora <= fin:
            if evento.estado_eve != "En Curso":
                evento.estado_eve = "En Curso"
                evento.save()
            continue

        tiempo_falta = inicio - ahora

        if timedelta(hours=0) < tiempo_falta <= timedelta(hours=24):
            if evento.estado_eve != "Próximo":
                evento.estado_eve = "Próximo"
                evento.save()
            continue

        if tiempo_falta > timedelta(hours=24):
            if evento.estado_eve != "Programado":
                evento.estado_eve = "Programado"
                evento.save()
            continue

    buscar = request.GET.get("buscar")
    if buscar:
        eventos = eventos.filter(nom_eve__icontains=buscar)

    # FILTROS
    estado = request.GET.get("estado")
    modalidad = request.GET.get("modalidad")

    if estado and estado != "Todos":
        eventos = eventos.filter(estado_eve=estado)

    if modalidad and modalidad != "Todos":
        eventos = eventos.filter(modalidad_eve=modalidad)

    # ESTADÍSTICAS
    mes_actual = ahora.month

    total_eventos = Evento.objects.count()
    total_mes = Evento.objects.filter(fecha_eve__month=mes_actual).count()

    presenciales = Evento.objects.filter(modalidad_eve="Presencial").count()
    presenciales_mes = Evento.objects.filter(modalidad_eve="Presencial", fecha_eve__month=mes_actual).count()

    virtuales = Evento.objects.filter(modalidad_eve="Virtual").count()
    virtuales_mes = Evento.objects.filter(modalidad_eve="Virtual", fecha_eve__month=mes_actual).count()

    contexto = {
        'current_page': 'eventos',
        'current_page_name': 'Eventos',
        'eventos': eventos,
        'total_eventos': total_eventos,
        'total_mes': total_mes,
        'presenciales': presenciales,
        'presenciales_mes': presenciales_mes,
        'virtuales': virtuales,
        'virtuales_mes': virtuales_mes,
        'estado_seleccionado': estado,
        'modalidad_seleccionada': modalidad,
        'busqueda': buscar,
        'usuario': usuario
    }

    return render(request, 'paginas/eventos.html', contexto)

@login_required
def crear_evento(request):

    cedula = request.session.get('cedula')
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')

    if request.method == 'POST':

        nom_eve = request.POST.get('nom_eve')
        fecha_eve = request.POST.get('fecha_eve')
        desc_eve = request.POST.get('desc_eve', '')
        modalidad_eve = request.POST.get('modalidad_eve')
        direccion_eve = request.POST.get('direccion_eve', '')
        estado_eve = request.POST.get('estado_eve', 'Programado')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')
        link = request.POST.get('link', '')

        # 🔹 Validación mínima
        if not nom_eve or not fecha_eve:
            messages.error(request, "Faltan campos obligatorios.")
            return redirect('eventos')

        # 🔹 Validación especial: Virtual → requiere link
        if modalidad_eve == "Virtual" and not link.strip():
            messages.error(request, "El link es obligatorio para eventos virtuales.")
            return redirect('eventos')

        # 🔹 Validación especial: Presencial → requiere dirección
        if modalidad_eve == "Presencial" and not direccion_eve.strip():
            messages.error(request, "La dirección es obligatoria para eventos presenciales.")
            return redirect('eventos')

        # 🔹 Crear el evento
        Evento.objects.create(
            nom_eve=nom_eve,
            fecha_eve=fecha_eve,
            desc_eve=desc_eve,
            modalidad_eve=modalidad_eve,
            direccion_eve=direccion_eve,
            link=link,
            estado_eve=estado_eve,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            cedula=usuario
        )

        messages.success(request, f"Evento '{nom_eve}' creado correctamente.")
        return redirect('eventos')

    return redirect('eventos')

@login_required
def editar_evento(request, cod_eve):
    # Obtener el evento
    evento = get_object_or_404(Evento, cod_eve=cod_eve)

    if request.method == "POST":

        evento.nom_eve = request.POST.get("nom_eve")
        evento.fecha_eve = request.POST.get("fecha_eve")
        evento.hora_inicio = request.POST.get("hora_inicio")
        evento.hora_fin = request.POST.get("hora_fin")
        evento.modalidad_eve = request.POST.get("modalidad_eve")
        evento.direccion_eve = request.POST.get("direccion_eve")
        evento.desc_eve = request.POST.get("desc_eve")
        evento.estado_eve = request.POST.get("estado_eve")
        evento.link = request.POST.get("link", "")

        # VALIDACIONES
        if evento.modalidad_eve == "Virtual" and evento.link.strip() == "":
            messages.error(request, "Debe ingresar el link para eventos virtuales.")
            return redirect("eventos")

        if evento.modalidad_eve == "Presencial" and evento.direccion_eve.strip() == "":
            messages.error(request, "Debe ingresar la dirección para eventos presenciales.")
            return redirect("eventos")

        evento.save()
        messages.success(request, f"El evento '{evento.nom_eve}' fue actualizado correctamente.")

        return redirect("eventos")

    return redirect("eventos")

@login_required
def cancelar_evento(request):
    if request.method == "POST":
        cod = request.POST.get("cod_eve")

        # Buscar el evento
        evento = get_object_or_404(Evento, cod_eve=cod)

        # Evitar cancelar un evento ya cancelado
        if evento.estado_eve.lower() == "cancelado":
            messages.warning(request, f"El evento '{evento.nom_eve}' ya está cancelado.")
            return redirect("eventos")

        # Cambiar estado
        evento.estado_eve = "Cancelado"
        evento.save()

        messages.success(request, f"El evento '{evento.nom_eve}' fue cancelado correctamente.")
        return redirect("eventos")

    return redirect("eventos")

# VISTAS DE CENTRO DE AYUDA
@login_required
def centroayuda(request):
    cedula = request.session.get('cedula')
    
    # Obtener el usuario actual
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    contexto = {
        'usuario': usuario
    }
    return render(request, 'paginas/centroayuda.html', contexto)

# VISTAS DE REPORTES
def reportes(request):
    cedula = request.session.get('cedula')
    
    # Obtener el usuario actual
    try:
        usuario = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('iniciarsesion')
    
    contexto = {
        'usuario': usuario
    }
    return render(request, 'paginas/reportes.html', contexto)

def reporte_general_semilleros(request):
    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Semilleros"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Cabecera
    ws.append([
        "Código Semillero",
        "Siglas",
        "Nombre Semillero",
        "Descripción",
        "Objetivos",
        "Fecha Creación",
        "Cantidad de Miembros",
        "Número de Proyectos",
        "Estado",
        "Progreso General",
        "Lider de Semillero"
    ])

    semilleros = Semillero.objects.all()

    for s in semilleros:
        cantidad_miembros = (
            SemilleroUsuario.objects.filter(id_sem=s).count() +
            Aprendiz.objects.filter(id_sem=s).count()
        )
        cantidad_proyectos = s.proyectos.count()

        lider_sem = SemilleroUsuario.objects.filter(id_sem=s, es_lider=True).first()
        nombre_lider = lider_sem.cedula.nom_usu if lider_sem else "Sin líder"

        objetivos = ", ".join(s.objetivo.splitlines()) if s.objetivo else ""

        ws.append([
            s.cod_sem,
            s.sigla,
            s.nombre,
            s.desc_sem,
            objetivos,
            s.fecha_creacion.strftime("%Y-%m-%d %H:%M"),
            cantidad_miembros,
            cantidad_proyectos,
            s.estado,
            s.progreso_sem,
            nombre_lider
        ])

    # Aplicar bordes
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border

    # Preparar respuesta Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_semilleros.xlsx"'

    # --- GRAFICO: Miembros por Semillero ---
    categorias = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # Nombre Semillero
    valores = Reference(ws, min_col=7, min_row=1, max_row=ws.max_row)     # Cantidad Miembros
    chart_miembros = BarChart()
    chart_miembros.title = "Miembros por Semillero"
    chart_miembros.y_axis.title = "Cantidad"
    chart_miembros.x_axis.title = "Semilleros"
    chart_miembros.add_data(valores, titles_from_data=True)
    chart_miembros.set_categories(categorias)
    ws.add_chart(chart_miembros, "L2")

    # --- GRAFICO: Proyectos por Semillero ---
    valores = Reference(ws, min_col=8, min_row=1, max_row=ws.max_row)  # Número de proyectos
    chart_proyectos = BarChart()
    chart_proyectos.title = "Proyectos por Semillero"
    chart_proyectos.y_axis.title = "Cantidad"
    chart_proyectos.x_axis.title = "Semilleros"
    chart_proyectos.add_data(valores, titles_from_data=True)
    chart_proyectos.set_categories(categorias)
    ws.add_chart(chart_proyectos, "L20")

    # --- GRAFICO: Progreso por Semillero ---
    valores = Reference(ws, min_col=10, min_row=1, max_row=ws.max_row)  # Progreso
    chart_progreso = BarChart()
    chart_progreso.title = "Progreso de Semilleros"
    chart_progreso.y_axis.title = "Progreso (%)"
    chart_progreso.x_axis.title = "Semilleros"
    chart_progreso.add_data(valores, titles_from_data=True)
    chart_progreso.set_categories(categorias)
    ws.add_chart(chart_progreso, "L38")

    # --- GRAFICO: Estados (Pastel) ---
    estados_count = {}
    for s in semilleros:
        if s.estado:
            estados_count[s.estado] = estados_count.get(s.estado, 0) + 1

    if estados_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Estado").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for estado, cantidad in estados_count.items():
            fila += 1
            ws.append([estado, cantidad])

        for row in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla+1, max_row=fila)
        chart_estado = PieChart()
        chart_estado.title = "Estados de los Semilleros"
        chart_estado.add_data(data, titles_from_data=True)
        chart_estado.set_categories(labels)
        ws.add_chart(chart_estado, "N55")

    wb.save(response)
    return response

def reporte_general_proyectos(request):
    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # --- Encabezados ---
    ws.append([
        "Nombre Proyecto",
        "Tipo",
        "Descripción",
        "Líneas Tecnologicas",
        "Líneas de Investigación",
        "Líneas de Semillero",
        "Notas Adicionales",
        "Cantidad de Miembros",
        "Cantidad de Entregables",
        "Fecha Creación",
        "Estado Actual",
        "Progreso",
        "Lider de Proyecto"
    ])

    proyectos = Proyecto.objects.all()

    # --- Llenado de datos ---
    for p in proyectos:
        cantidad_miembros = (
            UsuarioProyecto.objects.filter(cod_pro=p).count() +
            ProyectoAprendiz.objects.filter(cod_pro=p).count()
        )
        cantidad_entregables = Entregable.objects.filter(cod_pro=p).count()

        lider_pro = UsuarioProyecto.objects.filter(cod_pro=p, es_lider_pro=True).first()
        nombre_lider = lider_pro.cedula.nom_usu if lider_pro else "Sin líder"

        lineas_tec = ", ".join([l.strip() for l in p.linea_tec.split("\n") if l.strip()]) if p.linea_tec else ""
        lineas_inv = ", ".join([l.strip() for l in p.linea_inv.split("\n") if l.strip()]) if p.linea_inv else ""
        lineas_sem = ", ".join([l.strip() for l in p.linea_sem.split("\n") if l.strip()]) if p.linea_sem else ""
        notas = ", ".join([l.strip() for l in p.notas.split("\n") if l.strip()]) if p.notas else ""

        ws.append([
            p.nom_pro,
            p.tipo,
            p.desc_pro,
            lineas_tec,
            lineas_inv,
            lineas_sem,
            notas,
            cantidad_miembros,
            cantidad_entregables,
            p.fecha_creacion.strftime("%Y-%m-%d %H:%M"),
            p.estado_pro,
            p.progreso,
            nombre_lider
        ])

    # Aplicar bordes
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = thin_border

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_proyectos.xlsx"'

    # --- GRAFICO: Progreso por proyecto (barras) ---
    categorias = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    valores = Reference(ws, min_col=12, min_row=1, max_row=ws.max_row)
    chart_progreso = BarChart()
    chart_progreso.title = "Progreso por Proyecto"
    chart_progreso.y_axis.title = "Progreso (%)"
    chart_progreso.x_axis.title = "Proyectos"
    chart_progreso.add_data(valores, titles_from_data=True)
    chart_progreso.set_categories(categorias)
    ws.add_chart(chart_progreso, "N2")

    # --- GRAFICO: Miembros por proyecto ---
    valores = Reference(ws, min_col=8, min_row=1, max_row=ws.max_row)
    chart_miembros = BarChart()
    chart_miembros.title = "Miembros por Proyecto"
    chart_miembros.y_axis.title = "Cantidad"
    chart_miembros.x_axis.title = "Proyectos"
    chart_miembros.add_data(valores, titles_from_data=True)
    chart_miembros.set_categories(categorias)
    ws.add_chart(chart_miembros, "N20")

    # --- GRAFICO: Entregables por proyecto ---
    valores = Reference(ws, min_col=9, min_row=1, max_row=ws.max_row)
    chart_entregables = BarChart()
    chart_entregables.title = "Entregables por Proyecto"
    chart_entregables.y_axis.title = "Cantidad"
    chart_entregables.x_axis.title = "Proyectos"
    chart_entregables.add_data(valores, titles_from_data=True)
    chart_entregables.set_categories(categorias)
    ws.add_chart(chart_entregables, "N38")

    # --- GRAFICO: Estados de los proyectos (pastel) ---
    estados_count = {}
    for p in proyectos:
        if p.estado_pro:
            estados_count[p.estado_pro] = estados_count.get(p.estado_pro, 0) + 1

    if estados_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Estado").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for estado, cantidad in estados_count.items():
            fila += 1
            ws.append([estado, cantidad])

        for row in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla+1, max_row=fila)
        chart_estado = PieChart()
        chart_estado.title = "Estados de los Proyectos"
        chart_estado.add_data(data, titles_from_data=True)
        chart_estado.set_categories(labels)
        ws.add_chart(chart_estado, "N55")

    # --- GRAFICO: Líneas Tecnológicas (barras) ---
    lineas_tec_count = {}
    for p in proyectos:
        if p.linea_tec:
            for linea in [l.strip() for l in p.linea_tec.split("\n") if l.strip()]:
                lineas_tec_count[linea] = lineas_tec_count.get(linea, 0) + 1

    if lineas_tec_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Línea Tecnológica").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for linea, cantidad in lineas_tec_count.items():
            fila += 1
            ws.append([linea, cantidad])

        for row in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla+1, max_row=fila)
        chart_lineas_tec = BarChart()
        chart_lineas_tec.title = "Proyectos por Línea Tecnológica"
        chart_lineas_tec.add_data(data, titles_from_data=True)
        chart_lineas_tec.set_categories(labels)
        ws.add_chart(chart_lineas_tec, "N72")

    # --- GRAFICO: Líneas de Investigación (pastel) ---
    lineas_inv_count = {}
    for p in proyectos:
        if p.linea_inv:
            for linea in [l.strip() for l in p.linea_inv.split("\n") if l.strip()]:
                lineas_inv_count[linea] = lineas_inv_count.get(linea, 0) + 1

    if lineas_inv_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Línea Investigación").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for linea, cantidad in lineas_inv_count.items():
            fila += 1
            ws.append([linea, cantidad])

        for row in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla+1, max_row=fila)
        chart_lineas_inv = PieChart()
        chart_lineas_inv.title = "Proyectos por Línea de Investigación"
        chart_lineas_inv.add_data(data, titles_from_data=True)
        chart_lineas_inv.set_categories(labels)
        ws.add_chart(chart_lineas_inv, "N90")

    # --- GRAFICO: Líneas de Semillero (barras) ---
    lineas_sem_count = {}
    for p in proyectos:
        if p.linea_sem:
            for linea in [l.strip() for l in p.linea_sem.split("\n") if l.strip()]:
                lineas_sem_count[linea] = lineas_sem_count.get(linea, 0) + 1

    if lineas_sem_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Línea Semillero").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for linea, cantidad in lineas_sem_count.items():
            fila += 1
            ws.append([linea, cantidad])

        for row in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla+1, max_row=fila)
        chart_lineas_sem = BarChart()
        chart_lineas_sem.title = "Proyectos por Línea de Semillero"
        chart_lineas_sem.add_data(data, titles_from_data=True)
        chart_lineas_sem.set_categories(labels)
        ws.add_chart(chart_lineas_sem, "N108")

    wb.save(response)
    return response

def reporte_entregables(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entregables"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # --- Encabezados base ---
    encabezados = [
        "Proyecto",
        "Nombre Entregable",
        "Fecha Inicio",
        "Fecha Fin",
        "Descripción",
        "Estado",
        "Fechas de Subida"
    ]
    ws.append(encabezados)

    entregables = Entregable.objects.all()

    # --- Llenado de datos ---
    for e in entregables:
        archivos = Archivo.objects.filter(entregable=e)
        fechas_subida = ", ".join([a.fecha_subida.strftime("%Y-%m-%d %H:%M") for a in archivos]) if archivos else ""

        fila = [
            e.cod_pro.nom_pro,
            e.nom_entre,
            e.fecha_inicio.strftime("%Y-%m-%d") if e.fecha_inicio else "",
            e.fecha_fin.strftime("%Y-%m-%d") if e.fecha_fin else "",
            e.desc_entre,
            e.estado,
            fechas_subida,
        ]
        ws.append(fila)
        row = ws.max_row

        # Agregar archivos como hipervínculos en columnas adicionales
        col_base = len(encabezados) + 1
        for i, archivo in enumerate(archivos, start=1):
            col = col_base + i - 1
            cell = ws.cell(row=row, column=col)
            cell.value = f"Descargar archivo {i}"
            cell.hyperlink = request.build_absolute_uri(archivo.archivo.url)
            cell.style = "Hyperlink"
            ws.cell(row=1, column=col).value = f"Archivo {i}"

    # Aplicar bordes
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_entregables.xlsx"'

    # --- GRAFICO: Entregables por estado (pastel) ---
    estados_count = {}
    for e in entregables:
        if e.estado:
            estados_count[e.estado] = estados_count.get(e.estado, 0) + 1

    if estados_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Estado").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for estado, cantidad in estados_count.items():
            fila += 1
            ws.append([estado, cantidad])

        for row_cells in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla + 1, max_row=fila)
        chart_estado = PieChart()
        chart_estado.title = "Estados de los Entregables"
        chart_estado.add_data(data, titles_from_data=True)
        chart_estado.set_categories(labels)
        ws.add_chart(chart_estado, "L2")

    # --- GRAFICO: Entregables por proyecto (barras) ---
    proyectos_count = {}
    for e in entregables:
        nom = e.cod_pro.nom_pro
        proyectos_count[nom] = proyectos_count.get(nom, 0) + 1

    if proyectos_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Proyecto").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for proyecto, cantidad in proyectos_count.items():
            fila += 1
            ws.append([proyecto, cantidad])

        for row_cells in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla + 1, max_row=fila)
        chart_proyectos = BarChart()
        chart_proyectos.title = "Entregables por Proyecto"
        chart_proyectos.y_axis.title = "Cantidad"
        chart_proyectos.x_axis.title = "Proyectos"
        chart_proyectos.add_data(data, titles_from_data=True)
        chart_proyectos.set_categories(labels)
        ws.add_chart(chart_proyectos, "L20")

    wb.save(response)
    return response

def reporte_participantes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participantes"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # --- Encabezados ---
    encabezados = [
        "Nombre",
        "Rol",
        "Proyecto",
        "Semillero",
        "Correo Personal",
        "Correo Institucional",
        "Estado"
    ]
    ws.append(encabezados)

    # -------------------------------
    # APRENDICES
    # -------------------------------
    aprendices = Aprendiz.objects.all()
    for a in aprendices:
        semillero_nombre = a.id_sem.nombre if a.id_sem else "Sin semillero"
        proyectos_apre = ProyectoAprendiz.objects.filter(cedula_apre=a)

        if not proyectos_apre.exists():
            fila = [
                f"{a.nombre} {a.apellido}",
                "Aprendiz",
                "Sin proyecto",
                semillero_nombre,
                a.correo_per,
                a.correo_ins,
                a.estado_apre
            ]
            ws.append(fila)
        else:
            for pa in proyectos_apre:
                fila = [
                    f"{a.nombre} {a.apellido}",
                    "Aprendiz",
                    pa.cod_pro.nom_pro,
                    semillero_nombre,
                    a.correo_per,
                    a.correo_ins,
                    a.estado_apre
                ]
                ws.append(fila)

    # -------------------------------
    # USUARIOS (Instructores e Investigadores)
    # -------------------------------
    usuarios = Usuario.objects.filter(rol__in=["Instructor", "Investigador"])
    for u in usuarios:
        proyectos_usuario = UsuarioProyecto.objects.filter(cedula=u, estado="activo")
        semilleros_usuario = SemilleroUsuario.objects.filter(cedula=u)
        semillero_nombre = ", ".join([s.id_sem.nombre for s in semilleros_usuario]) if semilleros_usuario else "Sin semillero"

        if not proyectos_usuario.exists():
            fila = [
                f"{u.nom_usu} {u.ape_usu}",
                u.rol,
                "Sin proyecto",
                semillero_nombre,
                u.correo_per,
                u.correo_ins,
                u.estado or ""
            ]
            ws.append(fila)
        else:
            for up in proyectos_usuario:
                fila = [
                    f"{u.nom_usu} {u.ape_usu}",
                    u.rol,
                    up.cod_pro.nom_pro,
                    semillero_nombre,
                    u.correo_per,
                    u.correo_ins,
                    u.estado or ""
                ]
                ws.append(fila)

    # --- Aplicar bordes a tabla principal ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border

    # --- Preparar respuesta ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_participantes.xlsx"'

    # -------------------------------
    # GRAFICO: Participantes por rol
    # -------------------------------
    roles_count = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2, values_only=True):
        rol = row[0]
        roles_count[rol] = roles_count.get(rol, 0) + 1

    if roles_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Rol").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for rol, cantidad in roles_count.items():
            fila += 1
            ws.append([rol, cantidad])

        # Aplicar bordes
        for row_cells in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla + 1, max_row=fila)
        chart_roles = PieChart()
        chart_roles.title = "Participantes por Rol"
        chart_roles.add_data(data, titles_from_data=True)
        chart_roles.set_categories(labels)
        ws.add_chart(chart_roles, "H2")

    # -------------------------------
    # GRAFICO: Participantes por proyecto
    # -------------------------------
    proyectos_count = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3, values_only=True):
        proyecto = row[0] or "Sin proyecto"
        proyectos_count[proyecto] = proyectos_count.get(proyecto, 0) + 1

    if proyectos_count:
        fila_tabla = ws.max_row + 2
        ws.cell(row=fila_tabla, column=1, value="Proyecto").font = Font(bold=True)
        ws.cell(row=fila_tabla, column=2, value="Cantidad").font = Font(bold=True)

        fila = fila_tabla
        for proyecto, cantidad in proyectos_count.items():
            fila += 1
            ws.append([proyecto, cantidad])

        # Aplicar bordes
        for row_cells in ws.iter_rows(min_row=fila_tabla, max_row=fila, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = thin_border

        data = Reference(ws, min_col=2, min_row=fila_tabla, max_row=fila)
        labels = Reference(ws, min_col=1, min_row=fila_tabla + 1, max_row=fila)
        chart_proyectos = BarChart()
        chart_proyectos.title = "Participantes por Proyecto"
        chart_proyectos.y_axis.title = "Cantidad"
        chart_proyectos.x_axis.title = "Proyectos"
        chart_proyectos.add_data(data, titles_from_data=True)
        chart_proyectos.set_categories(labels)
        ws.add_chart(chart_proyectos, "H20")

    wb.save(response)
    return response

def generar_reporte_excel(request):

    if request.method != "POST":
        return redirect("constructor_reportes")

    # NOMBRE DEL ARCHIVO
    nombre_archivo = request.POST.get("nombre_plantilla", "").strip()
    if not nombre_archivo:
        nombre_archivo = "reporte_personalizado"

    # Crear respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'

    # CATEGORÍAS
    categorias = request.POST.getlist("categoria")

    if not categorias:
        messages.error(request, "Debes seleccionar al menos una categoría.")
        return redirect("constructor_reportes")

    # CAMPOS MARCADOS
    campos = {
        "semilleros": request.POST.getlist("campo_semilleros"),
        "proyectos": request.POST.getlist("campo_proyectos"),
        "miembros": request.POST.getlist("campo_miembros"),
        "entregables": request.POST.getlist("campo_entregables"),
    }

    # CREAR EXCEL
    wb = Workbook()
    hoja_default = wb.active
    wb.remove(hoja_default)
    
    # Crear estilo de borde
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== SEMILLEROS ====================
    if "semilleros" in categorias:
        ws = wb.create_sheet("Semilleros")
        
        # Título en negrita
        ws.cell(row=1, column=1, value="SEMILLEROS").font = Font(bold=True)
        
        # Encabezados en negrita
        fila_encabezado = 2
        for col_idx, campo in enumerate(campos["semilleros"], start=1):
            cell = ws.cell(row=fila_encabezado, column=col_idx, value=campo)
            cell.font = Font(bold=True)
        
        semilleros = Semillero.objects.all()
        fila_inicio_datos = 3

        for s in semilleros:
            fila = []
            for campo in campos["semilleros"]:
                match campo:
                    case "Código de Semillero": 
                        fila.append(s.cod_sem)
                    case "Nombre del Semillero": 
                        fila.append(s.nombre)
                    case "Siglas": 
                        fila.append(s.sigla)
                    case "Descripción": 
                        fila.append(s.desc_sem)
                    case "Progreso": 
                        fila.append(s.progreso_sem)
                    case "Objetivos": 
                        objetivos = s.objetivo.replace('\n', ', ') if s.objetivo else "Sin objetivos"
                        fila.append(objetivos)
                    case "Líder de Semillero":
                        lider = SemilleroUsuario.objects.filter(id_sem=s, es_lider=True).first()
                        fila.append(lider.cedula.nom_usu if lider else "Sin líder")
                    case "Fecha de Creación": 
                        fila.append(s.fecha_creacion.strftime("%Y-%m-%d"))
                    case "Estado Actual": 
                        fila.append(s.estado)
                    case "Número de Integrantes":
                        usuarios_sem = SemilleroUsuario.objects.filter(id_sem=s).select_related('cedula')
                        aprendices_sem = Aprendiz.objects.filter(id_sem=s)
                        roles_count = {}
                        for u_rel in usuarios_sem:
                            rol = u_rel.cedula.rol
                            roles_count[rol] = roles_count.get(rol, 0) + 1
                        cant_aprendices = aprendices_sem.count()
                        if cant_aprendices > 0:
                            roles_count['Aprendiz'] = cant_aprendices
                        total = sum(roles_count.values())
                        fila.append(total)
                    case "Cantidad de Proyectos": 
                        fila.append(s.proyectos.count())
                        
            ws.append(fila)
        
        # APLICAR BORDES A TABLA DE SEMILLEROS
        fila_fin = ws.max_row
        for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_fin, 
                                min_col=1, max_col=len(campos["semilleros"])):
            for cell in row:
                cell.border = thin_border

        # DESGLOSE DE INTEGRANTES POR ROL (en columnas SEPARADAS de Estado)
        if "Número de Integrantes" in campos["semilleros"]:
            col_inicio_roles = len(campos["semilleros"]) + 2
            
            ws.cell(row=fila_encabezado, column=col_inicio_roles, value="Rol").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio_roles + 1, value="Cantidad por Rol").font = Font(bold=True)
            
            todos_roles = {}
            for s in semilleros:
                usuarios_sem = SemilleroUsuario.objects.filter(id_sem=s).select_related('cedula')
                aprendices_sem = Aprendiz.objects.filter(id_sem=s)
                
                for u_rel in usuarios_sem:
                    rol = u_rel.cedula.rol
                    todos_roles[rol] = todos_roles.get(rol, 0) + 1
                
                cant_aprendices = aprendices_sem.count()
                if cant_aprendices > 0:
                    todos_roles['Aprendiz'] = todos_roles.get('Aprendiz', 0) + cant_aprendices
            
            fila_actual = fila_inicio_datos
            for rol, cantidad in todos_roles.items():
                ws.cell(row=fila_actual, column=col_inicio_roles, value=rol)
                ws.cell(row=fila_actual, column=col_inicio_roles + 1, value=cantidad)
                fila_actual += 1
            
            # Bordes para tabla de roles
            for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                    min_col=col_inicio_roles, max_col=col_inicio_roles + 1):
                for cell in row:
                    cell.border = thin_border

        # GRÁFICO DE BARRAS - CANTIDAD DE PROYECTOS POR SEMILLERO
        if "Cantidad de Proyectos" in campos["semilleros"]:
            # Ubicar columna de cantidad de proyectos
            col_proyectos = campos["semilleros"].index("Cantidad de Proyectos") + 1
            col_nombre = campos["semilleros"].index("Nombre del Semillero") + 1 if "Nombre del Semillero" in campos["semilleros"] else 1
            
            # Crear tabla auxiliar para el gráfico - COLUMNA DIFERENTE
            col_inicio_tabla_proy = len(campos["semilleros"]) + 8
            
            ws.cell(row=fila_encabezado, column=col_inicio_tabla_proy, value="Semillero").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio_tabla_proy + 1, value="Cantidad de Proyectos").font = Font(bold=True)
            
            fila_actual = fila_inicio_datos
            
            # Copiar datos de nombre y cantidad
            for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=fila_fin, 
                                min_col=col_nombre, max_col=col_nombre, values_only=True):
                nombre_sem = row[0]
                # Obtener cantidad de proyectos de la columna correspondiente
                cant_proyectos = ws.cell(row=fila_actual, column=col_proyectos).value
                
                ws.cell(row=fila_actual, column=col_inicio_tabla_proy, value=nombre_sem)
                ws.cell(row=fila_actual, column=col_inicio_tabla_proy + 1, value=cant_proyectos)
                fila_actual += 1
            
            # Aplicar bordes a tabla auxiliar
            for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                min_col=col_inicio_tabla_proy, max_col=col_inicio_tabla_proy + 1):
                for cell in row:
                    cell.border = thin_border
            
            # Crear gráfico de barras
            chart_proy = BarChart()
            chart_proy.title = "Cantidad de Proyectos por Semillero"
            chart_proy.y_axis.title = "Número de Proyectos"
            chart_proy.x_axis.title = "Semilleros"
            
            data = Reference(ws, min_col=col_inicio_tabla_proy + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
            labels = Reference(ws, min_col=col_inicio_tabla_proy, min_row=fila_inicio_datos, max_row=fila_actual - 1)
            
            chart_proy.add_data(data, titles_from_data=True)
            chart_proy.set_categories(labels)
            chart_proy.height = 15
            chart_proy.width = 20
            
            ws.add_chart(chart_proy, "N56")

        # GRÁFICO DE BARRAS - PROGRESO DE SEMILLEROS
        if "Progreso" in campos["semilleros"]:
            # Ubicar columnas necesarias
            col_progreso = campos["semilleros"].index("Progreso") + 1
            col_nombre = campos["semilleros"].index("Nombre del Semillero") + 1 if "Nombre del Semillero" in campos["semilleros"] else 1
            
            # Crear tabla auxiliar para el gráfico - COLUMNA MÁS A LA DERECHA
            col_inicio_tabla_prog = len(campos["semilleros"]) + 11
            
            ws.cell(row=fila_encabezado, column=col_inicio_tabla_prog, value="Semillero").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio_tabla_prog + 1, value="Progreso (%)").font = Font(bold=True)
            
            fila_actual = fila_inicio_datos
            
            # Copiar datos de nombre y progreso
            for row_idx in range(fila_inicio_datos, fila_fin + 1):
                nombre_sem = ws.cell(row=row_idx, column=col_nombre).value
                progreso = ws.cell(row=row_idx, column=col_progreso).value
                
                # Asegurar que el progreso sea numérico
                if progreso is None or progreso == "":
                    progreso = 0
                else:
                    try:
                        progreso = float(progreso)
                    except:
                        progreso = 0
                
                ws.cell(row=fila_actual, column=col_inicio_tabla_prog, value=nombre_sem)
                ws.cell(row=fila_actual, column=col_inicio_tabla_prog + 1, value=progreso)
                fila_actual += 1
            
            # Aplicar bordes a tabla auxiliar
            for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                min_col=col_inicio_tabla_prog, max_col=col_inicio_tabla_prog + 1):
                for cell in row:
                    cell.border = thin_border
            
            # Crear gráfico de barras
            chart_prog = BarChart()
            chart_prog.title = "Progreso de Semilleros"
            chart_prog.y_axis.title = "Progreso (%)"
            chart_prog.x_axis.title = "Semilleros"
            
            data = Reference(ws, min_col=col_inicio_tabla_prog + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
            labels = Reference(ws, min_col=col_inicio_tabla_prog, min_row=fila_inicio_datos, max_row=fila_actual - 1)
            
            chart_prog.add_data(data, titles_from_data=True)
            chart_prog.set_categories(labels)
            chart_prog.height = 15
            chart_prog.width = 20
            
            ws.add_chart(chart_prog, "N74")

        # GRÁFICO: Estados de semilleros (EN COLUMNAS DIFERENTES)
        if "Estado Actual" in campos["semilleros"]:
            try:
                col_estado = campos["semilleros"].index("Estado Actual") + 1
                
                estados = {}
                for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row, min_col=col_estado, max_col=col_estado, values_only=True):
                    estado = row[0]
                    if estado:
                        estados[estado] = estados.get(estado, 0) + 1
                
                if estados:
                    col_inicio_estados = len(campos["semilleros"]) + 5
                    
                    ws.cell(row=fila_encabezado, column=col_inicio_estados, value="Estado").font = Font(bold=True)
                    ws.cell(row=fila_encabezado, column=col_inicio_estados + 1, value="Cantidad por Estado").font = Font(bold=True)
                    
                    fila_actual = fila_inicio_datos
                    for estado, cantidad in estados.items():
                        ws.cell(row=fila_actual, column=col_inicio_estados, value=estado)
                        ws.cell(row=fila_actual, column=col_inicio_estados + 1, value=cantidad)
                        fila_actual += 1
                    
                    # Bordes para tabla de estados
                    for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                            min_col=col_inicio_estados, max_col=col_inicio_estados + 1):
                        for cell in row:
                            cell.border = thin_border
                    
                    chart = PieChart()
                    chart.title = "Distribución por Estado"
                    
                    data = Reference(ws, min_col=col_inicio_estados + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                    labels = Reference(ws, min_col=col_inicio_estados, min_row=fila_inicio_datos, max_row=fila_actual - 1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    chart.height = 10
                    chart.width = 15
                    
                    ws.add_chart(chart, "N20")
            except ValueError:
                pass

        # GRÁFICO: Distribución de integrantes por rol
        if "Número de Integrantes" in campos["semilleros"]:
            try:
                col_inicio_roles = len(campos["semilleros"]) + 2
                
                ultima_fila_desglose = fila_encabezado
                for row in ws.iter_rows(min_row=fila_inicio_datos, min_col=col_inicio_roles, max_col=col_inicio_roles, values_only=True):
                    if row[0]:
                        ultima_fila_desglose += 1
                
                if ultima_fila_desglose > fila_encabezado:
                    chart = PieChart()
                    chart.title = "Distribución de Integrantes por Rol"
                    
                    data = Reference(ws, min_col=col_inicio_roles + 1, min_row=fila_encabezado, max_row=ultima_fila_desglose)
                    labels = Reference(ws, min_col=col_inicio_roles, min_row=fila_inicio_datos, max_row=ultima_fila_desglose)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    chart.height = 10
                    chart.width = 15
                    
                    ws.add_chart(chart, "N38")
            except ValueError:
                pass

    # ==================== PROYECTOS ====================
    if "proyectos" in categorias:
        ws = wb.create_sheet("Proyectos")

        # Título en negrita
        ws.cell(row=1, column=1, value="PROYECTOS").font = Font(bold=True)

        # Encabezados en negrita (fila_encabezado = 2)
        fila_encabezado = 2
        num_cols = len(campos["proyectos"])
        for col_idx, campo in enumerate(campos["proyectos"], start=1):
            cell = ws.cell(row=fila_encabezado, column=col_idx, value=campo)
            cell.font = Font(bold=True)

        proyectos = list(Proyecto.objects.all())
        fila_inicio_datos = 3

        # --- Construir filas garantizando la longitud y posición correcta de cada columna ---
        for p_idx, p in enumerate(proyectos):
            # Crear lista con placeholders vacíos (una celda por cada encabezado)
            fila_vals = [""] * num_cols

            for i, campo in enumerate(campos["proyectos"]):
                # columna i -> index i
                match campo:
                    case "Título del Proyecto":
                        fila_vals[i] = p.nom_pro or ""
                    case "Tipo de Proyecto":
                        fila_vals[i] = p.tipo or ""
                    case "Estado":
                        fila_vals[i] = p.estado_pro or ""
                    case "Fecha de Creación":
                        fila_vals[i] = p.fecha_creacion.strftime("%Y-%m-%d") if getattr(p, "fecha_creacion", None) else ""
                    case "Porcentaje de Avance":
                        fila_vals[i] = p.progreso if p.progreso is not None else ""
                    case "Lider":
                        lider = UsuarioProyecto.objects.filter(cod_pro=p, es_lider_pro=True).select_related('cedula').first()
                        fila_vals[i] = lider.cedula.nom_usu if lider else "Sin líder"
                    case "Línea Tecnológica":
                        if getattr(p, "linea_tec", None):
                            lista = [s.strip() for s in p.linea_tec.splitlines() if s.strip()]
                            fila_vals[i] = ", ".join(lista)
                        else:
                            fila_vals[i] = ""
                    case "Línea de Investigación":
                        if getattr(p, "linea_inv", None):
                            lista = [s.strip() for s in p.linea_inv.splitlines() if s.strip()]
                            fila_vals[i] = ", ".join(lista)
                        else:
                            fila_vals[i] = ""
                    case "Línea de Semillero":
                        if getattr(p, "linea_sem", None):
                            lista = [s.strip() for s in p.linea_sem.splitlines() if s.strip()]
                            fila_vals[i] = ", ".join(lista)
                        else:
                            fila_vals[i] = ""
                    case "Participantes De Proyecto":
                        usuarios_pro = UsuarioProyecto.objects.filter(cod_pro=p).select_related('cedula')
                        aprendices_pro = ProyectoAprendiz.objects.filter(cod_pro=p)
                        fila_vals[i] = usuarios_pro.count() + aprendices_pro.count()
                    case "Notas":
                        fila_vals[i] = ", ".join([s.strip() for s in p.notas.splitlines() if s.strip()]) if getattr(p, "notas", None) else ""
                    case "Programa de Formación":
                        fila_vals[i] = p.programa_formacion if getattr(p, "programa_formacion", None) else ""

            # Ahora escribimos la fila completa en la hoja (esto preserva columnas vacías)
            ws.append(fila_vals)

        # APLICAR BORDES A TABLA DE PROYECTOS (uso len(campos["proyectos"]) para evitar desbordes)
        fila_fin = ws.max_row
        for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_fin,
                                min_col=1, max_col=len(campos["proyectos"])):
            for cell in row:
                cell.border = thin_border

        # DESGLOSE DE PARTICIPANTES POR ROL (proyectos)
        if "Participantes De Proyecto" in campos["proyectos"]:

            col_inicio = len(campos["proyectos"]) + 2

            ws.cell(row=fila_encabezado, column=col_inicio, value="Rol").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            roles_count = {}

            for p in proyectos:
                usuarios = UsuarioProyecto.objects.filter(cod_pro=p).select_related('cedula')
                aprendices = ProyectoAprendiz.objects.filter(cod_pro=p)

                for rel in usuarios:
                    rol = rel.cedula.rol
                    roles_count[rol] = roles_count.get(rol, 0) + 1

                count_apre = aprendices.count()
                if count_apre > 0:
                    roles_count["Aprendiz"] = roles_count.get("Aprendiz", 0) + count_apre

            fila_actual = fila_inicio_datos
            for rol, cant in roles_count.items():
                ws.cell(row=fila_actual, column=col_inicio, value=rol)
                ws.cell(row=fila_actual, column=col_inicio + 1, value=cant)
                fila_actual += 1
            
            # Bordes para tabla de roles
            for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border

            # GRÁFICO 
            chart = PieChart()
            chart.title = "Participantes por Rol"

            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15

            ws.add_chart(chart, "N2")
                
        # TABLA Y GRÁFICO – TIPO DE PROYECTO
        if "Tipo de Proyecto" in campos["proyectos"]:

            col_inicio = len(campos["proyectos"]) + 8
            ws.cell(row=fila_encabezado, column=col_inicio, value="Tipo").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            tipos = {}

            for p in proyectos:
                if p.tipo:
                    tipos[p.tipo] = tipos.get(p.tipo, 0) + 1

            fila_actual = fila_inicio_datos
            for tipo, cantidad in tipos.items():
                ws.cell(row=fila_actual, column=col_inicio, value=tipo)
                ws.cell(row=fila_actual, column=col_inicio + 1, value=cantidad)
                fila_actual += 1
            
            # Bordes
            for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border

            chart = PieChart()
            chart.title = "Proyectos por Tipo"
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15

            ws.add_chart(chart, "N38")

        # TABLA Y GRÁFICO – AVANCE
        if "Porcentaje de Avance" in campos["proyectos"]:
            # Ubicar columnas necesarias
            col_avance = campos["proyectos"].index("Porcentaje de Avance") + 1
            col_nombre = campos["proyectos"].index("Título del Proyecto") + 1 if "Título del Proyecto" in campos["proyectos"] else 1
            
            # Crear tabla auxiliar para el gráfico - COLUMNA MÁS A LA DERECHA
            col_inicio = len(campos["proyectos"]) + 11

            # Encabezados
            ws.cell(row=fila_encabezado, column=col_inicio, value="Proyecto").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Avance (%)").font = Font(bold=True)

            fila_actual = fila_inicio_datos

            # Copiar datos de nombre y avance
            for row_idx in range(fila_inicio_datos, fila_fin + 1):
                nombre_pro = ws.cell(row=row_idx, column=col_nombre).value
                avance = ws.cell(row=row_idx, column=col_avance).value
                
                # Asegurar que el avance sea numérico
                if avance is None or avance == "":
                    avance = 0
                else:
                    try:
                        avance = float(avance)
                    except:
                        avance = 0
                
                ws.cell(row=fila_actual, column=col_inicio, value=nombre_pro)
                ws.cell(row=fila_actual, column=col_inicio + 1, value=avance)
                fila_actual += 1

            # Bordes
            for row in ws.iter_rows(
                min_row=fila_encabezado,
                max_row=fila_actual - 1,
                min_col=col_inicio,
                max_col=col_inicio + 1
            ):
                for cell in row:
                    cell.border = thin_border

            # Referencias para el gráfico
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

            # Crear gráfico de barras
            chart = BarChart()
            chart.title = "Porcentaje de Avance por Proyecto"
            chart.y_axis.title = "Avance (%)"
            chart.x_axis.title = "Proyectos"
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 15
            chart.width = 20

            ws.add_chart(chart, "N56")
            
        # TABLA Y GRÁFICO – LÍNEA TECNOLÓGICA 
        if "Línea Tecnológica" in campos["proyectos"]:

            col_inicio = len(campos["proyectos"]) + 14
            ws.cell(row=fila_encabezado, column=col_inicio, value="Línea Tec.").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            lineas_tec = {}

            # Leer directamente desde el objeto proyecto
            for p in proyectos:
                if p.linea_tec and p.linea_tec.strip():
                    # Dividir por saltos de línea y contar cada línea
                    lineas_lista = [l.strip() for l in p.linea_tec.split('\n') if l.strip()]
                    for linea in lineas_lista:
                        lineas_tec[linea] = lineas_tec.get(linea, 0) + 1

            # Solo crear tabla y gráfico si hay datos
            if lineas_tec:
                fila_actual = fila_inicio_datos
                for linea, cant in lineas_tec.items():
                    ws.cell(row=fila_actual, column=col_inicio, value=linea)
                    ws.cell(row=fila_actual, column=col_inicio + 1, value=cant)
                    fila_actual += 1
                
                # Bordes
                for row in ws.iter_rows(
                    min_row=fila_encabezado, max_row=fila_actual - 1, 
                    min_col=col_inicio, max_col=col_inicio + 1
                ):
                    for cell in row:
                        cell.border = thin_border

                chart = PieChart()
                chart.title = "Proyectos por Línea Tecnológica"
                data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.height = 10
                chart.width = 15

                ws.add_chart(chart, "N74")

        # TABLA Y GRÁFICO – LÍNEA DE INVESTIGACIÓN 
        if "Línea de Investigación" in campos["proyectos"]:

            col_inicio = len(campos["proyectos"]) + 17
            ws.cell(row=fila_encabezado, column=col_inicio, value="Línea Inv.").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            lineas_inv = {}

            # Leer directamente desde el objeto proyecto
            for p in proyectos:
                if p.linea_inv and p.linea_inv.strip():
                    # Dividir por saltos de línea y contar cada línea
                    lineas_lista = [l.strip() for l in p.linea_inv.split('\n') if l.strip()]
                    for linea in lineas_lista:
                        lineas_inv[linea] = lineas_inv.get(linea, 0) + 1

            # Solo crear tabla y gráfico si hay datos
            if lineas_inv:
                fila_actual = fila_inicio_datos
                for linea, cant in lineas_inv.items():
                    ws.cell(row=fila_actual, column=col_inicio, value=linea)
                    ws.cell(row=fila_actual, column=col_inicio + 1, value=cant)
                    fila_actual += 1
                
                # Bordes
                for row in ws.iter_rows(
                    min_row=fila_encabezado, max_row=fila_actual - 1, 
                    min_col=col_inicio, max_col=col_inicio + 1
                ):
                    for cell in row:
                        cell.border = thin_border

                chart = PieChart()
                chart.title = "Proyectos por Línea de Investigación"
                data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.height = 10
                chart.width = 15

                ws.add_chart(chart, "N92")

        # TABLA Y GRÁFICO – LÍNEA DE SEMILLERO 
        if "Línea de Semillero" in campos["proyectos"]:

            col_inicio = len(campos["proyectos"]) + 20
            ws.cell(row=fila_encabezado, column=col_inicio, value="Línea Sem.").font = Font(bold=True)
            ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            lineas_sem = {}

            # Leer directamente desde el objeto proyecto
            for p in proyectos:
                if p.linea_sem and p.linea_sem.strip():
                    # Dividir por saltos de línea y contar cada línea
                    lineas_lista = [l.strip() for l in p.linea_sem.split('\n') if l.strip()]
                    for linea in lineas_lista:
                        lineas_sem[linea] = lineas_sem.get(linea, 0) + 1

            # Solo crear tabla y gráfico si hay datos
            if lineas_sem:
                fila_actual = fila_inicio_datos
                for linea, cant in lineas_sem.items():
                    ws.cell(row=fila_actual, column=col_inicio, value=linea)
                    ws.cell(row=fila_actual, column=col_inicio + 1, value=cant)
                    fila_actual += 1
                
                # Bordes
                for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                        min_col=col_inicio, max_col=col_inicio + 1):
                    for cell in row:
                        cell.border = thin_border

                chart = PieChart()
                chart.title = "Proyectos por Línea de Semillero"
                data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.height = 10
                chart.width = 15

                ws.add_chart(chart, "N110")

            # TABLA Y GRÁFICO – PROGRAMA DE FORMACIÓN (PROYECTOS)
            if "Programa de Formación" in campos["proyectos"]:

                col_inicio = len(campos["proyectos"]) + 23

                ws.cell(row=fila_encabezado, column=col_inicio, value="Programa de Formación").font = Font(bold=True)
                ws.cell(row=fila_encabezado, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

                programas_count = {}

                for p in proyectos:
                    programa = p.programa_formacion.strip() if p.programa_formacion else "Sin programa"
                    programas_count[programa] = programas_count.get(programa, 0) + 1

                fila_actual = fila_inicio_datos
                for programa, cantidad in programas_count.items():
                    ws.cell(row=fila_actual, column=col_inicio, value=programa)
                    ws.cell(row=fila_actual, column=col_inicio + 1, value=cantidad)
                    fila_actual += 1
                
                # Bordes
                for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                        min_col=col_inicio, max_col=col_inicio + 1):
                    for cell in row:
                        cell.border = thin_border

                chart = PieChart()
                chart.title = "Proyectos por Programa de Formación"

                data = Reference(ws, min_col=col_inicio + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_datos, max_row=fila_actual - 1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.height = 10
                chart.width = 15

                ws.add_chart(chart, "N128")

            # GRÁFICO: Estados de proyectos (COLUMNAS SEPARADAS)
            if "Estado" in campos["proyectos"]:
                try:
                    col_estado = campos["proyectos"].index("Estado") + 1
                    
                    estados = {}
                    for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row, min_col=col_estado, max_col=col_estado, values_only=True):
                        estado = row[0]
                        if estado:
                            estados[estado] = estados.get(estado, 0) + 1
                    
                    if estados:
                        col_inicio_estados = len(campos["proyectos"]) + 5
                        
                        ws.cell(row=fila_encabezado, column=col_inicio_estados, value="Estado del Proyecto").font = Font(bold=True)
                        ws.cell(row=fila_encabezado, column=col_inicio_estados + 1, value="Cantidad por Estado").font = Font(bold=True)
                        
                        fila_actual = fila_inicio_datos
                        for estado, cantidad in estados.items():
                            ws.cell(row=fila_actual, column=col_inicio_estados, value=estado)
                            ws.cell(row=fila_actual, column=col_inicio_estados + 1, value=cantidad)
                            fila_actual += 1
                        
                        # Bordes
                        for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                                min_col=col_inicio_estados, max_col=col_inicio_estados + 1):
                            for cell in row:
                                cell.border = thin_border
                        
                        chart = PieChart()
                        chart.title = "Distribución por Estado"
                        
                        data = Reference(ws, min_col=col_inicio_estados + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                        labels = Reference(ws, min_col=col_inicio_estados, min_row=fila_inicio_datos, max_row=fila_actual - 1)
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        chart.height = 10
                        chart.width = 15
                        
                        ws.add_chart(chart, "N20")

                except ValueError:
                    pass
    
    # ==================== MIEMBROS ====================
    if "miembros" in categorias:

        ws = wb.create_sheet("Miembros")

        campos_seleccionados = campos["miembros"]

        # === TABLA 1: USUARIOS (NO APRENDICES)
        
        # Filtrar campos para usuarios 
        campos_usuarios = [c for c in campos_seleccionados 
            if c not in ["Programa", "Ficha", "Modalidad", "Programa de Formación"]]
        
        fila_actual = 1
        
        # Solo crear tabla de usuarios si hay campos para mostrar
        if campos_usuarios:
            # TÍTULO
            ws.cell(row=fila_actual, column=1, value="USUARIOS").font = Font(bold=True)
            fila_actual += 1
            
            # ENCABEZADOS en negrita
            fila_inicio_usuarios = fila_actual
            for col_idx, campo in enumerate(campos_usuarios, start=1):
                cell = ws.cell(row=fila_actual, column=col_idx, value=campo)
                cell.font = Font(bold=True)
            fila_actual += 1

            usuarios = Usuario.objects.all()

            for u in usuarios:

                if u.rol.lower() == "aprendiz":
                    continue

                fila = []
                for campo in campos_usuarios:

                    match campo:
                        case "Nombre Completo":
                            fila.append(f"{u.nom_usu} {u.ape_usu}")

                        case "Tipo de Documento":
                            fila.append("Cédula")

                        case "Documento":
                            fila.append(u.cedula)

                        case "Rol":
                            fila.append(u.rol)

                        case "Email":
                            fila.append(u.correo_ins or u.correo_per)

                        case "Teléfono":
                            fila.append(u.telefono)

                ws.append(fila)
            
            # APLICAR BORDES A TABLA DE USUARIOS
            fila_fin_usuarios = ws.max_row
            for row in ws.iter_rows(min_row=fila_inicio_usuarios, max_row=fila_fin_usuarios, 
                                    min_col=1, max_col=len(campos_usuarios)):
                for cell in row:
                    cell.border = thin_border

        # === SEPARADOR VISUAL

        fila_sep = ws.max_row + 2
        ws.cell(row=fila_sep, column=1, value="APRENDICES").font = Font(bold=True)

        # === TABLA 2: APRENDICES

        fila_inicio_ap = fila_sep + 1

        aprendices = Aprendiz.objects.all()

        # Encabezados en negrita
        for col_idx, campo in enumerate(campos_seleccionados, start=1):
            cell = ws.cell(row=fila_inicio_ap, column=col_idx, value=campo)
            cell.font = Font(bold=True)
        
        fila_inicio_ap_datos = fila_inicio_ap + 1

        # CONTENIDO DE APRENDICES
        for a in aprendices:

            fila = []

            for campo in campos_seleccionados:
                match campo:
                    case "Nombre Completo":
                        fila.append(f"{a.nombre} {a.apellido}")

                    case "Tipo de Documento":
                        fila.append(a.tipo_doc)

                    case "Documento":
                        fila.append(a.cedula_apre)

                    case "Rol":
                        fila.append("Aprendiz")

                    case "Email":
                        fila.append(a.correo_ins or a.correo_per)

                    case "Teléfono":
                        fila.append(a.telefono)

                    case "Programa" | "Programa de Formación":
                        fila.append(a.programa if hasattr(a, 'programa') else "")

                    case "Ficha":
                        fila.append(a.ficha if hasattr(a, 'ficha') else "")

                    case "Modalidad":
                        fila.append(a.modalidad if hasattr(a, 'modalidad') else "")

            ws.append(fila)
        
        # APLICAR BORDES A TABLA DE APRENDICES
        fila_fin_aprendices = ws.max_row
        for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_fin_aprendices, 
                                min_col=1, max_col=len(campos_seleccionados)):
            for cell in row:
                cell.border = thin_border

        # GRÁFICOS

        fila_inicio_graficos = fila_inicio_ap + len(aprendices) + 3
        
        # Calcular desplazamiento de columnas según si existe tabla de roles
        desplazamiento_col = 0

        # ----------- TABLA Y GRÁFICO DE ROLES (TODOS LOS MIEMBROS) -----------
        if "Rol" in campos_seleccionados:
            col_inicio_roles = len(campos_seleccionados) + 2
            
            ws.cell(row=fila_inicio_ap, column=col_inicio_roles, value="Rol").font = Font(bold=True)
            ws.cell(row=fila_inicio_ap, column=col_inicio_roles + 1, value="Cantidad").font = Font(bold=True)
            
            # Contar roles de TODOS los usuarios (no aprendices)
            roles_count = {}
            usuarios = Usuario.objects.all()
            
            for u in usuarios:
                if u.rol.lower() != "aprendiz":
                    rol = u.rol
                    roles_count[rol] = roles_count.get(rol, 0) + 1
            
            # Agregar aprendices
            cant_aprendices = aprendices.count()
            if cant_aprendices > 0:
                roles_count["Aprendiz"] = cant_aprendices
            
            # Escribir datos en tabla
            fila_g = fila_inicio_ap + 1
            for rol, cant in roles_count.items():
                ws.cell(row=fila_g, column=col_inicio_roles, value=rol)
                ws.cell(row=fila_g, column=col_inicio_roles + 1, value=cant)
                fila_g += 1
            
            # Bordes para tabla de roles
            for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_g - 1, 
                                    min_col=col_inicio_roles, max_col=col_inicio_roles + 1):
                for cell in row:
                    cell.border = thin_border
            
            # GRÁFICO DE ROLES
            chart = PieChart()
            chart.title = "Distribución de Miembros por Rol"
            data = Reference(ws, min_col=col_inicio_roles + 1, min_row=fila_inicio_ap, max_row=fila_g - 1)
            labels = Reference(ws, min_col=col_inicio_roles, min_row=fila_inicio_ap + 1, max_row=fila_g - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            ws.add_chart(chart, f"N{fila_inicio_graficos}")
            
            desplazamiento_col = 4  # Espacio para tabla de roles

        # ----------- PROGRAMA -----------
        if "Programa" in campos_seleccionados or "Programa de Formación" in campos_seleccionados:

            col_inicio = len(campos_seleccionados) + 2 + desplazamiento_col

            ws.cell(row=fila_inicio_ap, column=col_inicio, value="Programa").font = Font(bold=True)
            ws.cell(row=fila_inicio_ap, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            programas = {}
            for a in aprendices:
                prog = (a.programa if hasattr(a, 'programa') else None) or "Sin programa"
                programas[prog] = programas.get(prog, 0) + 1

            fila_g = fila_inicio_ap + 1
            for prog, cant in programas.items():
                ws.cell(row=fila_g, column=col_inicio, value=prog)
                ws.cell(row=fila_g, column=col_inicio + 1, value=cant)
                fila_g += 1
            
            # Bordes
            for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_g - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border

            chart = PieChart()
            chart.title = "Aprendices por Programa"
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_inicio_ap, max_row=fila_g - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_ap + 1, max_row=fila_g - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            
            # Ajustar posición según si existe tabla de roles
            pos_grafico = fila_inicio_graficos + 18 if desplazamiento_col > 0 else fila_inicio_graficos
            ws.add_chart(chart, f"N{pos_grafico}")

        # ----------- FICHA -----------
        if "Ficha" in campos_seleccionados:

            col_inicio = len(campos_seleccionados) + 6 + desplazamiento_col

            ws.cell(row=fila_inicio_ap, column=col_inicio, value="Ficha").font = Font(bold=True)
            ws.cell(row=fila_inicio_ap, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            fichas = {}
            for a in aprendices:
                f = (a.ficha if hasattr(a, 'ficha') else None) or "Sin ficha"
                fichas[f] = fichas.get(f, 0) + 1

            fila_g = fila_inicio_ap + 1
            for f, cant in fichas.items():
                ws.cell(row=fila_g, column=col_inicio, value=f)
                ws.cell(row=fila_g, column=col_inicio + 1, value=cant)
                fila_g += 1
            
            # Bordes
            for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_g - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border

            chart = PieChart()
            chart.title = "Aprendices por Ficha"
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_inicio_ap, max_row=fila_g - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_ap + 1, max_row=fila_g - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            
            # Ajustar posición según si existe tabla de roles
            pos_grafico = fila_inicio_graficos + 36 if desplazamiento_col > 0 else fila_inicio_graficos + 18
            ws.add_chart(chart, f"N{pos_grafico}")


        # ----------- MODALIDAD -----------
        if "Modalidad" in campos_seleccionados:

            col_inicio = len(campos_seleccionados) + 10 + desplazamiento_col

            ws.cell(row=fila_inicio_ap, column=col_inicio, value="Modalidad").font = Font(bold=True)
            ws.cell(row=fila_inicio_ap, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)

            modalidades = {}
            for a in aprendices:
                mod = (a.modalidad if hasattr(a, 'modalidad') else None) or "Sin modalidad"
                modalidades[mod] = modalidades.get(mod, 0) + 1

            fila_g = fila_inicio_ap + 1
            for mod, cant in modalidades.items():
                ws.cell(row=fila_g, column=col_inicio, value=mod)
                ws.cell(row=fila_g, column=col_inicio + 1, value=cant)
                fila_g += 1
            
            # Bordes
            for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_g - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border

            chart = PieChart()
            chart.title = "Aprendices por Modalidad"
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_inicio_ap, max_row=fila_g - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_ap + 1, max_row=fila_g - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            
            # Ajustar posición según si existe tabla de roles
            pos_grafico = fila_inicio_graficos + 54 if desplazamiento_col > 0 else fila_inicio_graficos + 36
            ws.add_chart(chart, f"N{pos_grafico}")

        # ----------- TIPO DE DOCUMENTO -----------
        if "Tipo de Documento" in campos_seleccionados:
            
            col_inicio = len(campos_seleccionados) + 14 + desplazamiento_col
            
            ws.cell(row=fila_inicio_ap, column=col_inicio, value="Tipo de Documento").font = Font(bold=True)
            ws.cell(row=fila_inicio_ap, column=col_inicio + 1, value="Cantidad").font = Font(bold=True)
            
            tipos_doc = {}
            
            # Contar tipos de documento de USUARIOS (todos tienen Cédula)
            usuarios = Usuario.objects.all()
            for u in usuarios:
                if u.rol.lower() != "aprendiz":
                    tipos_doc["Cédula"] = tipos_doc.get("Cédula", 0) + 1
            
            # Contar tipos de documento de APRENDICES
            for a in aprendices:
                tipo = (a.tipo_doc if hasattr(a, 'tipo_doc') else None) or "Cédula"
                tipos_doc[tipo] = tipos_doc.get(tipo, 0) + 1
            
            # Escribir datos en tabla
            fila_g = fila_inicio_ap + 1
            for tipo, cant in tipos_doc.items():
                ws.cell(row=fila_g, column=col_inicio, value=tipo)
                ws.cell(row=fila_g, column=col_inicio + 1, value=cant)
                fila_g += 1
            
            # Aplicar bordes
            for row in ws.iter_rows(min_row=fila_inicio_ap, max_row=fila_g - 1, 
                                    min_col=col_inicio, max_col=col_inicio + 1):
                for cell in row:
                    cell.border = thin_border
            
            # Crear gráfico
            chart = PieChart()
            chart.title = "Distribución por Tipo de Documento"
            data = Reference(ws, min_col=col_inicio + 1, min_row=fila_inicio_ap, max_row=fila_g - 1)
            labels = Reference(ws, min_col=col_inicio, min_row=fila_inicio_ap + 1, max_row=fila_g - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            
            # Ajustar posición según si existe tabla de roles
            pos_grafico = fila_inicio_graficos + 72 if desplazamiento_col > 0 else fila_inicio_graficos + 54
            ws.add_chart(chart, f"N{pos_grafico}")

    # ==================== ENTREGABLES ====================
    if "entregables" in categorias:
        ws = wb.create_sheet("Entregables")
        
        # Título en negrita
        ws.cell(row=1, column=1, value="ENTREGABLES").font = Font(bold=True)
        
        # Encabezados en negrita
        fila_encabezado = 2
        for col_idx, campo in enumerate(campos["entregables"], start=1):
            cell = ws.cell(row=fila_encabezado, column=col_idx, value=campo)
            cell.font = Font(bold=True)

        entregables = Entregable.objects.all()
        fila_inicio_datos = 3

        for e in entregables:
            fila = []
            for campo in campos["entregables"]:
                match campo:
                    case "Nombre del Entregable":
                        fila.append(e.nom_entre)
                    case "Estado":
                        fila.append(e.estado)
                    case "Fecha de Entrega":
                        fila.append(e.fecha_fin.strftime("%Y-%m-%d") if e.fecha_fin else "")
                    case "Proyecto Asociado":
                        fila.append(e.cod_pro.nom_pro)
                    case "Responsable":
                        resp = UsuarioProyecto.objects.filter(cod_pro=e.cod_pro, es_lider_pro=True).first()
                        fila.append(resp.cedula.nom_usu if resp else "Sin responsable")
            ws.append(fila)
        
        # APLICAR BORDES A TABLA DE ENTREGABLES
        fila_fin = ws.max_row
        for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_fin, 
                                min_col=1, max_col=len(campos["entregables"])):
            for cell in row:
                cell.border = thin_border

        # GRÁFICO: Estados de entregables
        if "Estado" in campos["entregables"]:
            try:
                col_estado = campos["entregables"].index("Estado") + 1
                
                estados = {}
                for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row, min_col=col_estado, max_col=col_estado, values_only=True):
                    estado = row[0]
                    if estado:
                        estados[estado] = estados.get(estado, 0) + 1
                
                if estados:
                    col_inicio_grafico = len(campos["entregables"]) + 2
                    
                    ws.cell(row=fila_encabezado, column=col_inicio_grafico, value="Estado").font = Font(bold=True)
                    ws.cell(row=fila_encabezado, column=col_inicio_grafico + 1, value="Cantidad").font = Font(bold=True)
                    
                    fila_actual = fila_inicio_datos
                    for estado, cantidad in estados.items():
                        ws.cell(row=fila_actual, column=col_inicio_grafico, value=estado)
                        ws.cell(row=fila_actual, column=col_inicio_grafico + 1, value=cantidad)
                        fila_actual += 1
                    
                    # Bordes
                    for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                            min_col=col_inicio_grafico, max_col=col_inicio_grafico + 1):
                        for cell in row:
                            cell.border = thin_border
                    
                    chart = PieChart()
                    chart.title = "Entregables por Estado"
                    
                    data = Reference(ws, min_col=col_inicio_grafico + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                    labels = Reference(ws, min_col=col_inicio_grafico, min_row=fila_inicio_datos, max_row=fila_actual - 1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    chart.height = 10
                    chart.width = 15
                    
                    ws.add_chart(chart, "H2")
            except ValueError:
                pass

        # GRÁFICO: Proyectos Asociados
        if "Proyecto Asociado" in campos["entregables"]:
            try:
                col_proyecto = campos["entregables"].index("Proyecto Asociado") + 1
                
                proyectos = {}
                for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row, min_col=col_proyecto, max_col=col_proyecto, values_only=True):
                    proyecto = row[0]
                    if proyecto:
                        proyectos[proyecto] = proyectos.get(proyecto, 0) + 1
                
                if proyectos:
                    col_inicio_proyectos = len(campos["entregables"]) + 5
                    
                    ws.cell(row=fila_encabezado, column=col_inicio_proyectos, value="Proyecto").font = Font(bold=True)
                    ws.cell(row=fila_encabezado, column=col_inicio_proyectos + 1, value="Cantidad de Entregables").font = Font(bold=True)
                    
                    fila_actual = fila_inicio_datos
                    for proyecto, cantidad in proyectos.items():
                        ws.cell(row=fila_actual, column=col_inicio_proyectos, value=proyecto)
                        ws.cell(row=fila_actual, column=col_inicio_proyectos + 1, value=cantidad)
                        fila_actual += 1
                    
                    # Bordes
                    for row in ws.iter_rows(min_row=fila_encabezado, max_row=fila_actual - 1, 
                                            min_col=col_inicio_proyectos, max_col=col_inicio_proyectos + 1):
                        for cell in row:
                            cell.border = thin_border
                    
                    chart = PieChart()
                    chart.title = "Entregables por Proyecto"
                    
                    data = Reference(ws, min_col=col_inicio_proyectos + 1, min_row=fila_encabezado, max_row=fila_actual - 1)
                    labels = Reference(ws, min_col=col_inicio_proyectos, min_row=fila_inicio_datos, max_row=fila_actual - 1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    chart.height = 10
                    chart.width = 15
                    
                    ws.add_chart(chart, "H20")
            except ValueError:
                pass

    wb.save(response)
    return response

def generar_reporte_dinamico(request):
    if request.method != "POST":
        return redirect("reportes")

    # Obtener formato seleccionado
    formato = request.POST.get("formato", "excel")
    
    if formato == "pdf":
        return generar_reporte_pdf(request)
    else:
        return generar_reporte_excel(request)

def generar_reporte_pdf(request):
    # NOMBRE DEL ARCHIVO
    nombre_archivo = request.POST.get("nombre_plantilla", "").strip()
    if not nombre_archivo:
        nombre_archivo = "reporte_personalizado"

    # Crear respuesta HTTP para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

    # Crear buffer y documento
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=30, 
        leftMargin=30,
        topMargin=30, 
        bottomMargin=18
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495E'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Estilos para las celdas de las tablas
    style_header = ParagraphStyle(
        'HeaderCell',
        fontSize=9,
        textColor=colors.whitesmoke,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        leading=11,
        wordWrap='CJK'
    )
    
    style_celda = ParagraphStyle(
        'DataCell',
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
        wordWrap='CJK'
    )
    
    # Contenedor de elementos
    elements = []
    
    # CATEGORÍAS
    categorias = request.POST.getlist("categoria")
    
    if not categorias:
        messages.error(request, "Debes seleccionar al menos una categoría.")
        return redirect("constructor_reportes")

    # CAMPOS MARCADOS
    campos = {
        "semilleros": request.POST.getlist("campo_semilleros"),
        "proyectos": request.POST.getlist("campo_proyectos"),
        "miembros": request.POST.getlist("campo_miembros"),
        "entregables": request.POST.getlist("campo_entregables"),
    }

    # ==================== SEMILLEROS ====================
    if "semilleros" in categorias:
        elements.append(Paragraph("REPORTE DE SEMILLEROS", title_style))
        elements.append(Spacer(1, 12))
        
        semilleros = Semillero.objects.all()
        
        # Preparar datos para la tabla
        data = [campos["semilleros"]]  # Encabezados
        
        for s in semilleros:
            fila = []
            for campo in campos["semilleros"]:
                valor = ""  # Valor por defecto
                
                if campo == "Código de Semillero":
                    valor = str(s.cod_sem) if s.cod_sem else ""
                elif campo == "Nombre del Semillero":
                    valor = str(s.nombre) if s.nombre else ""
                elif campo == "Siglas":
                    valor = str(s.sigla) if s.sigla else ""
                elif campo == "Descripción":
                    valor = str(s.desc_sem) if s.desc_sem else ""
                elif campo == "Progreso":
                    valor = f"{s.progreso_sem}%" if s.progreso_sem is not None else "0%"
                elif campo == "Objetivos":
                    valor = s.objetivo.replace('\n', ', ') if s.objetivo else ""
                elif campo == "Líder de Semillero":
                    lider = SemilleroUsuario.objects.filter(id_sem=s, es_lider=True).first()
                    valor = lider.cedula.nom_usu if lider else ""
                elif campo == "Fecha de Creación":
                    valor = s.fecha_creacion.strftime("%Y-%m-%d") if s.fecha_creacion else ""
                elif campo == "Estado Actual":
                    valor = str(s.estado) if s.estado else ""
                elif campo == "Número de Integrantes":
                    usuarios_sem = SemilleroUsuario.objects.filter(id_sem=s).count()
                    aprendices_sem = Aprendiz.objects.filter(id_sem=s).count()
                    valor = str(usuarios_sem + aprendices_sem)
                elif campo == "Cantidad de Proyectos":
                    valor = str(s.proyectos.count())
                
                fila.append(valor)  # Siempre agregar valor (aunque sea vacío)
            
            data.append(fila)
        
        # Convertir todos los datos a Paragraphs
        for i in range(len(data)):
            for j in range(len(data[i])):
                if i == 0:  # Encabezado
                    data[i][j] = Paragraph(str(data[i][j]), style_header)
                else:  # Celdas de datos
                    data[i][j] = Paragraph(str(data[i][j]) if data[i][j] else " ", style_celda)  # ✅ Espacio si está vacío
        
        # Calcular anchos de forma inteligente
        ancho_disponible = landscape(A4)[0] - 60
        col_widths = []
        for campo in campos["semilleros"]:
            if campo in ["Código de Semillero", "Siglas", "Progreso", "Estado Actual"]:
                col_widths.append(50)
            elif campo in ["Fecha de Creación", "Número de Integrantes", "Cantidad de Proyectos"]:
                col_widths.append(70)
            elif campo in ["Nombre del Semillero", "Líder de Semillero"]:
                col_widths.append(100)
            else:
                col_widths.append(150)
        
        total_width = sum(col_widths)
        if total_width > ancho_disponible:
            factor = ancho_disponible / total_width
            col_widths = [w * factor for w in col_widths]
        
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # ==================== PROYECTOS ====================
    if "proyectos" in categorias:
        elements.append(Paragraph("REPORTE DE PROYECTOS", title_style))
        elements.append(Spacer(1, 12))
        
        proyectos = Proyecto.objects.all()
        data = [campos["proyectos"]]
        
        for p in proyectos:
            fila = []
            for campo in campos["proyectos"]:
                valor = ""  # Valor por defecto
                
                if campo == "Título del Proyecto":
                    valor = str(p.nom_pro) if p.nom_pro else ""
                elif campo == "Tipo de Proyecto":
                    valor = str(p.tipo) if p.tipo else ""
                elif campo == "Estado":
                    valor = str(p.estado_pro) if p.estado_pro else ""
                elif campo == "Fecha de Creación":
                    valor = p.fecha_creacion.strftime("%Y-%m-%d") if hasattr(p, "fecha_creacion") and p.fecha_creacion else ""
                elif campo == "Porcentaje de Avance":
                    valor = f"{p.progreso}%" if p.progreso is not None else "0%"
                elif campo == "Lider":
                    lider = UsuarioProyecto.objects.filter(cod_pro=p, es_lider_pro=True).first()
                    valor = lider.cedula.nom_usu if lider else ""
                elif campo == "Línea Tecnológica":
                    if p.linea_tec:
                        valor = ", ".join([l.strip() for l in p.linea_tec.splitlines() if l.strip()])
                elif campo == "Línea de Investigación":
                    if p.linea_inv:
                        valor = ", ".join([l.strip() for l in p.linea_inv.splitlines() if l.strip()])
                elif campo == "Línea de Semillero":
                    if p.linea_sem:
                        valor = ", ".join([l.strip() for l in p.linea_sem.splitlines() if l.strip()])
                elif campo == "Participantes De Proyecto":
                    usuarios = UsuarioProyecto.objects.filter(cod_pro=p).count()
                    aprendices = ProyectoAprendiz.objects.filter(cod_pro=p).count()
                    valor = str(usuarios + aprendices)
                elif campo == "Programa de Formación":
                    valor = str(p.programa_formacion) if p.programa_formacion else ""
                elif campo == "Notas":  # AGREGADO
                    valor = ", ".join([s.strip() for s in p.notas.splitlines() if s.strip()]) if hasattr(p, "notas") and p.notas else ""
                
                fila.append(valor)  # Siempre agregar valor
            
            data.append(fila)
        
        # Convertir todos los datos a Paragraphs
        for i in range(len(data)):
            for j in range(len(data[i])):
                if i == 0:
                    data[i][j] = Paragraph(str(data[i][j]), style_header)
                else:
                    data[i][j] = Paragraph(str(data[i][j]) if data[i][j] else " ", style_celda)  # ✅ Espacio si está vacío
        
        ancho_disponible = landscape(A4)[0] - 60
        col_widths = []
        for campo in campos["proyectos"]:
            if campo in ["Estado", "Porcentaje de Avance", "Participantes De Proyecto"]:
                col_widths.append(60)
            elif campo in ["Tipo de Proyecto", "Fecha de Creación"]:
                col_widths.append(80)
            elif campo in ["Título del Proyecto", "Lider", "Programa de Formación"]:
                col_widths.append(100)
            elif campo == "Notas":  # AGREGADO
                col_widths.append(150)
            else:
                col_widths.append(120)
        
        total_width = sum(col_widths)
        if total_width > ancho_disponible:
            factor = ancho_disponible / total_width
            col_widths = [w * factor for w in col_widths]
        
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # ==================== MIEMBROS (CORRECCIÓN) ====================
    if "miembros" in categorias:
        elements.append(Paragraph("REPORTE DE MIEMBROS", title_style))
        elements.append(Spacer(1, 12))
        
        # USUARIOS
        elements.append(Paragraph("Usuarios", heading_style))
        usuarios = Usuario.objects.exclude(rol__iexact="aprendiz")
        
        # Filtrar campos que no aplican a Usuarios
        campos_usuarios = [c for c in campos["miembros"] 
            if c not in ["Programa", "Ficha", "Modalidad", "Programa de Formación"]]
        
        if campos_usuarios and usuarios.exists():  # nVerificar que hay datos
            data = [campos_usuarios]
            
            for u in usuarios:
                fila = []
                for campo in campos_usuarios:
                    valor = ""
                    
                    if campo == "Nombre Completo":
                        valor = f"{u.nom_usu or ''} {u.ape_usu or ''}".strip()
                    elif campo == "Tipo de Documento":
                        valor = "Cédula"
                    elif campo == "Documento":
                        valor = str(u.cedula) if u.cedula else ""
                    elif campo == "Rol":
                        valor = str(u.rol) if u.rol else ""
                    elif campo == "Email":
                        valor = u.correo_ins or u.correo_per or ""
                    elif campo == "Teléfono":
                        valor = str(u.telefono) if u.telefono else ""
                    
                    fila.append(valor)
                
                data.append(fila)
            
            # Convertir a Paragraphs
            for i in range(len(data)):
                for j in range(len(data[i])):
                    if i == 0:
                        data[i][j] = Paragraph(str(data[i][j]), style_header)
                    else:
                        data[i][j] = Paragraph(str(data[i][j]) if data[i][j] else " ", style_celda)
            
            # Calcular anchos
            ancho_disponible = landscape(A4)[0] - 60
            col_widths = []
            for campo in campos_usuarios:
                if campo in ["Tipo de Documento", "Documento", "Rol", "Teléfono"]:
                    col_widths.append(70)
                elif campo == "Nombre Completo":
                    col_widths.append(120)
                elif campo == "Email":
                    col_widths.append(150)
                else:
                    col_widths.append(100)  # ✅ Ancho por defecto
            
            total_width = sum(col_widths)
            if total_width > ancho_disponible:
                factor = ancho_disponible / total_width
                col_widths = [w * factor for w in col_widths]
            
            t = Table(data, repeatRows=1, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))
        
        # APRENDICES
        elements.append(Paragraph("Aprendices", heading_style))
        aprendices = Aprendiz.objects.all()
        
        # ✅ Usar TODOS los campos seleccionados (incluyendo los específicos de aprendices)
        if campos["miembros"] and aprendices.exists():  # ✅ Verificar que hay datos
            data = [campos["miembros"]]
            
            for a in aprendices:
                fila = []
                for campo in campos["miembros"]:
                    valor = ""
                    
                    if campo == "Nombre Completo":
                        valor = f"{a.nombre or ''} {a.apellido or ''}".strip()
                    elif campo == "Tipo de Documento":
                        valor = str(a.tipo_doc) if a.tipo_doc else "Cédula"
                    elif campo == "Documento":
                        valor = str(a.cedula_apre) if a.cedula_apre else ""
                    elif campo == "Rol":
                        valor = "Aprendiz"
                    elif campo == "Email":
                        valor = a.correo_ins or a.correo_per or ""
                    elif campo == "Teléfono":
                        valor = str(a.telefono) if a.telefono else ""
                    elif campo in ["Programa", "Programa de Formación"]:
                        valor = str(a.programa) if hasattr(a, 'programa') and a.programa else ""
                    elif campo == "Ficha":
                        valor = str(a.ficha) if hasattr(a, 'ficha') and a.ficha else ""
                    elif campo == "Modalidad":
                        valor = str(a.modalidad) if hasattr(a, 'modalidad') and a.modalidad else ""
                    
                    fila.append(valor)
                
                data.append(fila)
            
            # Convertir a Paragraphs
            for i in range(len(data)):
                for j in range(len(data[i])):
                    if i == 0:
                        data[i][j] = Paragraph(str(data[i][j]), style_header)
                    else:
                        data[i][j] = Paragraph(str(data[i][j]) if data[i][j] else " ", style_celda)
            
            # Calcular anchos
            ancho_disponible = landscape(A4)[0] - 60
            col_widths = []
            for campo in campos["miembros"]:
                if campo in ["Tipo de Documento", "Documento", "Rol", "Teléfono", "Ficha", "Modalidad"]:
                    col_widths.append(70)
                elif campo == "Nombre Completo":
                    col_widths.append(120)
                elif campo == "Email":
                    col_widths.append(150)
                elif campo in ["Programa", "Programa de Formación"]:
                    col_widths.append(130)
                else:
                    col_widths.append(100)  # ✅ Ancho por defecto
            
            total_width = sum(col_widths)
            if total_width > ancho_disponible:
                factor = ancho_disponible / total_width
                col_widths = [w * factor for w in col_widths]
            
            t = Table(data, repeatRows=1, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
        
        elements.append(PageBreak())

    # ==================== ENTREGABLES (CORRECCIÓN) ====================
    if "entregables" in categorias:
        elements.append(Paragraph("REPORTE DE ENTREGABLES", title_style))
        elements.append(Spacer(1, 12))
        
        entregables = Entregable.objects.all()
        
        if campos["entregables"] and entregables.exists():  # ✅ Verificar que hay datos
            data = [campos["entregables"]]
            
            for e in entregables:
                fila = []
                for campo in campos["entregables"]:
                    valor = ""
                    
                    if campo == "Nombre del Entregable":
                        valor = str(e.nom_entre) if e.nom_entre else ""
                    elif campo == "Estado":
                        valor = str(e.estado) if e.estado else ""
                    elif campo == "Fecha de Entrega":
                        valor = e.fecha_fin.strftime("%Y-%m-%d") if e.fecha_fin else ""
                    elif campo == "Proyecto Asociado":
                        # ✅ Verificar que cod_pro no sea None antes de acceder a nom_pro
                        if e.cod_pro and hasattr(e.cod_pro, 'nom_pro') and e.cod_pro.nom_pro:
                            valor = str(e.cod_pro.nom_pro)
                        else:
                            valor = "Sin proyecto"
                    elif campo == "Responsable":
                        # ✅ Verificar que cod_pro existe antes de buscar el líder
                        if e.cod_pro:
                            resp = UsuarioProyecto.objects.filter(
                                cod_pro=e.cod_pro, 
                                es_lider_pro=True
                            ).first()
                            valor = resp.cedula.nom_usu if resp and resp.cedula else ""
                        else:
                            valor = ""
                    
                    fila.append(valor)
                
                data.append(fila)
            
            # Convertir a Paragraphs
            for i in range(len(data)):
                for j in range(len(data[i])):
                    if i == 0:
                        data[i][j] = Paragraph(str(data[i][j]), style_header)
                    else:
                        data[i][j] = Paragraph(str(data[i][j]) if data[i][j] else " ", style_celda)
            
            # Calcular anchos
            ancho_disponible = landscape(A4)[0] - 60
            col_widths = []
            for campo in campos["entregables"]:
                if campo in ["Estado", "Fecha de Entrega"]:
                    col_widths.append(80)
                elif campo in ["Nombre del Entregable", "Proyecto Asociado", "Responsable"]:
                    col_widths.append(150)
                else:
                    col_widths.append(100)  # ✅ Ancho por defecto
            
            total_width = sum(col_widths)
            if total_width > ancho_disponible:
                factor = ancho_disponible / total_width
                col_widths = [w * factor for w in col_widths]
            
            t = Table(data, repeatRows=1, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

def reporte_tendencias_crecimiento(request):
    from collections import defaultdict
    
    # WORKBOOK
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tendencias de Crecimiento"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # TÍTULO
    ws.merge_cells('A1:E1')
    titulo = ws['A1']
    titulo.value = "TENDENCIAS DE CRECIMIENTO"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')

    # ENCABEZADOS
    ws.append([])
    ws.append(["Período", "Semilleros Creados", "Proyectos Creados", "Participantes Activos", "Entregables Completados"])

    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

    # FECHAS AWARE
    fecha_actual = timezone.now()
    
    fecha_inicio = (fecha_actual - relativedelta(months=11)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Obtener TODOS los registros (no filtrar por fecha aquí)
    semilleros = Semillero.objects.filter(
        fecha_creacion__isnull=False
    ).values_list('fecha_creacion', flat=True)
    
    proyectos = Proyecto.objects.filter(
        fecha_creacion__isnull=False
    ).values_list('fecha_creacion', flat=True)

    # Organizar SEMILLEROS por mes (conteo por mes, no acumulado)
    sem_dict = defaultdict(int)
    for fecha in semilleros:
        if timezone.is_naive(fecha):
            fecha = timezone.make_aware(fecha)
        # Convertir a fecha local y luego crear clave como string YYYY-MM
        fecha_local = timezone.localtime(fecha)
        clave_mes = fecha_local.strftime("%Y-%m")
        sem_dict[clave_mes] += 1

    # Organizar PROYECTOS por mes (conteo por mes, no acumulado)
    proy_dict = defaultdict(int)
    for fecha in proyectos:
        if timezone.is_naive(fecha):
            fecha = timezone.make_aware(fecha)
        # Convertir a fecha local y luego crear clave como string YYYY-MM
        fecha_local = timezone.localtime(fecha)
        clave_mes = fecha_local.strftime("%Y-%m")
        proy_dict[clave_mes] += 1

    # GENERAR FILAS PARA 12 MESES (desde hace 11 meses hasta el mes actual = 12 meses)
    for i in range(12):
        mes_actual = fecha_inicio + relativedelta(months=i)
        mes_siguiente = mes_actual + relativedelta(months=1)
        periodo = mes_actual.strftime("%b %Y")

        # Crear clave como string YYYY-MM para buscar en los diccionarios
        clave_mes = mes_actual.strftime("%Y-%m")

        # Semilleros CREADOS en este mes específico
        semilleros_mes = sem_dict.get(clave_mes, 0)
        
        # Proyectos CREADOS en este mes específico
        proyectos_mes = proy_dict.get(clave_mes, 0)

        participantes = Usuario.objects.filter(
            fecha_registro__lte=mes_siguiente
        ).count() + Aprendiz.objects.filter(
            fecha_registro__lte=mes_siguiente
        ).count()

        # Entregables COMPLETADOS durante este mes específico
        entregables_mes = Entregable.objects.filter(
            estado__in=['Completado', 'Entrega Tardía'],
            fecha_fin__gte=mes_actual,
            fecha_fin__lt=mes_siguiente
        ).count()

        ws.append([periodo, semilleros_mes, proyectos_mes, participantes, entregables_mes])

    # BORDES
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # GRÁFICO
    chart = LineChart()
    chart.title = "Evolución Temporal"
    chart.style = 15
    chart.y_axis.title = 'Cantidad'
    chart.x_axis.title = 'Período'

    data = Reference(ws, min_col=2, min_row=3, max_col=5, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "G3")

    # RESUMEN ESTADÍSTICO
    fila_resumen = ws.max_row + 3
    ws.cell(row=fila_resumen, column=1, value="RESUMEN ESTADÍSTICO").font = Font(bold=True, size=12)
    fila_resumen += 1
    ws.append([])

    ws.append(["Métrica", "Valor"])
    ws.cell(row=fila_resumen + 1, column=1).font = Font(bold=True)
    ws.cell(row=fila_resumen + 1, column=2).font = Font(bold=True)

    # Calcular totales de semilleros y proyectos creados en los últimos 12 meses
    total_sem_12m = sum(sem_dict.values())
    total_proy_12m = sum(proy_dict.values())
    
    # Tasas de crecimiento (comparando primer mes vs último mes)
    if ws.max_row > 4:
        primer_mes_sem = ws.cell(row=4, column=2).value or 0
        ultimo_mes_sem = ws.cell(row=ws.max_row, column=2).value or 0
        
        primer_mes_proy = ws.cell(row=4, column=3).value or 0
        ultimo_mes_proy = ws.cell(row=ws.max_row, column=3).value or 0
        
        # Crecimiento basado en el cambio mensual
        crec_sem = ((ultimo_mes_sem - primer_mes_sem) / primer_mes_sem * 100) if primer_mes_sem > 0 else 0
        crec_proy = ((ultimo_mes_proy - primer_mes_proy) / primer_mes_proy * 100) if primer_mes_proy > 0 else 0
    else:
        crec_sem = crec_proy = 0

    ws.append([f"Total Semilleros Creados (últimos 12 meses)", total_sem_12m])
    ws.append([f"Total Proyectos Creados (últimos 12 meses)", total_proy_12m])
    ws.append([f"Variación Mensual Semilleros", f"{crec_sem:.1f}%"])
    ws.append([f"Variación Mensual Proyectos", f"{crec_proy:.1f}%"])
    ws.append([f"Total Participantes Actuales", Usuario.objects.count() + Aprendiz.objects.count()])
    ws.append([f"Total Entregables Completados (histórico)", Entregable.objects.filter(estado__in=['Completado', 'Entrega Tardía']).count()])

    # RESPUESTA
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tendencias_crecimiento.xlsx"'

    wb.save(response)
    return response

def reporte_productividad_semillero(request):
    """
    Genera reporte Excel con productividad por semillero
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productividad"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # TÍTULO
    ws.merge_cells('A1:F1')
    titulo = ws['A1']
    titulo.value = "PRODUCTIVIDAD POR SEMILLERO"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')

    # ENCABEZADOS
    ws.append([])
    ws.append([
        "Semillero",
        "Proyectos Finalizados",
        "Entregables Completados",
        "Total Participantes",
        "Tasa de Finalización (%)",
        "Productividad Promedio"
    ])
    
    # Aplicar estilo a encabezados
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # OBTENER DATOS
    semilleros = Semillero.objects.all()
    
    for sem in semilleros:
        # Proyectos del semillero
        proyectos = SemilleroProyecto.objects.filter(id_sem=sem)
        total_proyectos = proyectos.count()
        
        # Proyectos finalizados
        proyectos_finalizados = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            estado_pro='completado'
        ).count()
        
        # Entregables completados
        entregables_completados = Entregable.objects.filter(
            cod_pro__in=proyectos.values('cod_pro'),
            estado__in=['Completado', 'Entrega Tardía']
        ).count()
        
        # Total entregables
        total_entregables = Entregable.objects.filter(
            cod_pro__in=proyectos.values('cod_pro')
        ).count()
        
        # Participantes
        usuarios = SemilleroUsuario.objects.filter(id_sem=sem).count()
        aprendices = Aprendiz.objects.filter(id_sem=sem).count()
        total_participantes = usuarios + aprendices
        
        # Tasa de finalización
        tasa_finalizacion = (proyectos_finalizados / total_proyectos * 100) if total_proyectos > 0 else 0
        
        # Productividad promedio (entregables completados / participantes)
        productividad = (entregables_completados / total_participantes) if total_participantes > 0 else 0
        
        ws.append([
            sem.nombre,
            proyectos_finalizados,
            entregables_completados,
            total_participantes,
            round(tasa_finalizacion, 1),
            round(productividad, 2)
        ])

    # APLICAR BORDES
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    # GRÁFICO DE BARRAS - ENTREGABLES COMPLETADOS
    chart1 = BarChart()
    chart1.title = "Entregables Completados por Semillero"
    chart1.style = 10
    chart1.y_axis.title = 'Cantidad'
    chart1.x_axis.title = 'Semillero'
    
    data1 = Reference(ws, min_col=3, min_row=3, max_row=ws.max_row)
    cats1 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    
    ws.add_chart(chart1, "H3")

    # GRÁFICO DE BARRAS - PROYECTOS FINALIZADOS
    chart2 = BarChart()
    chart2.title = "Proyectos Finalizados por Semillero"
    chart2.style = 11
    chart2.y_axis.title = 'Cantidad'
    chart2.x_axis.title = 'Semillero'
    
    data2 = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row)
    cats2 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    
    ws.add_chart(chart2, "H20")

    # GRÁFICO DE DISPERSIÓN - PRODUCTIVIDAD
    chart3 = BarChart()
    chart3.title = "Productividad Promedio por Semillero"
    chart3.style = 12
    chart3.y_axis.title = 'Productividad'
    chart3.x_axis.title = 'Semillero'
    
    data3 = Reference(ws, min_col=6, min_row=3, max_row=ws.max_row)
    cats3 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    
    ws.add_chart(chart3, "H37")

    # RANKING DE SEMILLEROS
    fila_ranking = ws.max_row + 3
    
    ws.cell(row=fila_ranking, column=1, value="TOP 5 SEMILLEROS MÁS PRODUCTIVOS").font = Font(bold=True, size=12)
    fila_ranking += 1
    
    ws.append([])
    ws.append(["Posición", "Semillero", "Productividad"])
    
    for cell in ws[fila_ranking + 1]:
        cell.font = Font(bold=True)
    
    # Ordenar semilleros por productividad
    datos_ranking = []
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:  # Si hay nombre de semillero
            datos_ranking.append({
                'nombre': ws.cell(row=row, column=1).value,
                'productividad': ws.cell(row=row, column=6).value or 0
            })
    
    datos_ranking.sort(key=lambda x: x['productividad'], reverse=True)
    
    for i, dato in enumerate(datos_ranking[:5], 1):
        ws.append([i, dato['nombre'], dato['productividad']])

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="productividad_semillero.xlsx"'
    
    wb.save(response)
    return response

def reporte_mensual_ejecutivo(request):
    """
    Genera reporte consolidado del mes actual: logros, avances y estadísticas clave
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Mensual"
    
    # Obtener fecha actual y rango del mes
    fecha_actual = timezone.now()
    inicio_mes = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calcular fin de mes
    if fecha_actual.month == 12:
        fin_mes = fecha_actual.replace(year=fecha_actual.year + 1, month=1, day=1)
    else:
        fin_mes = fecha_actual.replace(month=fecha_actual.month + 1, day=1)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"REPORTE MENSUAL EJECUTIVO - {fecha_actual.strftime('%B %Y').upper()}"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # SECCIÓN 1: RESUMEN GENERAL
    fila = 3
    ws.cell(row=fila, column=1, value="RESUMEN GENERAL").font = Font(bold=True, size=14)
    fila += 2
    
    # Estadísticas del mes
    semilleros_activos = Semillero.objects.filter(estado='Activo').count()
    proyectos_mes = Proyecto.objects.filter(fecha_creacion__gte=inicio_mes, fecha_creacion__lt=fin_mes).count()
    entregables_completados_mes = Entregable.objects.filter(
        estado__in=['Completado', 'Entrega Tardía'],
        fecha_fin__gte=inicio_mes,
        fecha_fin__lt=fin_mes
    ).count()
    
    # Escribir encabezados
    ws.cell(row=fila, column=1, value="Métrica").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Valor").font = Font(bold=True)
    fila += 1
    
    datos_resumen = [
        ["Semilleros Activos", semilleros_activos],
        ["Proyectos Creados Este Mes", proyectos_mes],
        ["Entregables Completados", entregables_completados_mes],
        ["Total Participantes", Usuario.objects.count() + Aprendiz.objects.count()],
    ]
    
    for dato in datos_resumen:
        ws.cell(row=fila, column=1, value=dato[0])
        ws.cell(row=fila, column=2, value=dato[1])
        fila += 1
    
    inicio_resumen = 5
    # Aplicar bordes
    for row in ws.iter_rows(min_row=inicio_resumen, max_row=fila-1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 2: LOGROS DEL MES
    fila += 2
    ws.cell(row=fila, column=1, value="LOGROS DEL MES").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_logros = fila
    ws.cell(row=fila, column=1, value="Proyecto").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Avance (%)").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Estado").font = Font(bold=True)
    fila += 1
    
    # Proyectos con mayor avance este mes
    proyectos_avance = Proyecto.objects.filter(
        fecha_creacion__lt=fin_mes
    ).order_by('-progreso')[:10]
    
    for p in proyectos_avance:
        ws.cell(row=fila, column=1, value=p.nom_pro)
        ws.cell(row=fila, column=2, value=p.progreso)
        ws.cell(row=fila, column=3, value=p.estado_pro)
        fila += 1
    
    # Guardar la fila final de logros
    fila_final_logros = fila - 1
    
    # Bordes para logros
    for row in ws.iter_rows(min_row=inicio_logros, max_row=fila_final_logros, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 3: PARTICIPACIÓN POR SEMILLERO
    fila += 2
    ws.cell(row=fila, column=1, value="PARTICIPACIÓN POR SEMILLERO").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_participacion = fila
    ws.cell(row=fila, column=1, value="Semillero").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Integrantes").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Proyectos Activos").font = Font(bold=True)
    fila += 1
    
    semilleros = Semillero.objects.all()
    for sem in semilleros:
        integrantes = SemilleroUsuario.objects.filter(id_sem=sem).count() + Aprendiz.objects.filter(id_sem=sem).count()
        proyectos_activos = SemilleroProyecto.objects.filter(id_sem=sem).count()
        ws.cell(row=fila, column=1, value=sem.nombre)
        ws.cell(row=fila, column=2, value=integrantes)
        ws.cell(row=fila, column=3, value=proyectos_activos)
        fila += 1
    
    # Guardar la fila final de participación
    fila_final_participacion = fila - 1
    
    # Bordes para participación
    for row in ws.iter_rows(min_row=inicio_participacion, max_row=fila_final_participacion, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # GRÁFICOS
    # Gráfico 1: Avance de proyectos
    chart1 = BarChart()
    chart1.title = "Top 10 Proyectos por Avance"
    chart1.y_axis.title = "Progreso (%)"
    
    # Datos: solo los valores numéricos (sin encabezado)
    data1 = Reference(ws, min_col=2, min_row=inicio_logros+1, max_row=fila_final_logros)
    # Categorías: nombres de proyectos
    cats1 = Reference(ws, min_col=1, min_row=inicio_logros+1, max_row=fila_final_logros)
    
    chart1.add_data(data1, titles_from_data=False)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "F5")
    
    # Gráfico 2: Participación por semillero
    chart2 = PieChart()
    chart2.title = "Distribución de Integrantes"
    
    # Datos: solo los valores numéricos
    data2 = Reference(ws, min_col=2, min_row=inicio_participacion+1, max_row=fila_final_participacion)
    # Categorías: nombres de semilleros
    cats2 = Reference(ws, min_col=1, min_row=inicio_participacion+1, max_row=fila_final_participacion)
    
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "F25")
    
    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="reporte_mensual_{fecha_actual.strftime("%Y_%m")}.xlsx"'
    
    wb.save(response)
    return response

def informe_trimestral(request):
    """
    Análisis de cumplimiento y resultados por trimestre
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe Trimestral"
    
    fecha_actual = timezone.now()
    
    # Calcular inicio del trimestre actual
    mes_actual = fecha_actual.month
    if mes_actual <= 3:
        inicio_trimestre = fecha_actual.replace(month=1, day=1)
        trimestre = "Q1"
    elif mes_actual <= 6:
        inicio_trimestre = fecha_actual.replace(month=4, day=1)
        trimestre = "Q2"
    elif mes_actual <= 9:
        inicio_trimestre = fecha_actual.replace(month=7, day=1)
        trimestre = "Q3"
    else:
        inicio_trimestre = fecha_actual.replace(month=10, day=1)
        trimestre = "Q4"
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:E1')
    titulo = ws['A1']
    titulo.value = f"INFORME TRIMESTRAL {trimestre} - {fecha_actual.year}"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # SECCIÓN 1: CUMPLIMIENTO DE OBJETIVOS
    fila = 3
    ws.cell(row=fila, column=1, value="CUMPLIMIENTO DE OBJETIVOS").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_cumplimiento = fila
    ws.cell(row=fila, column=1, value="Semillero").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Proyectos Completados").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Proyectos En Curso").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Tasa de Cumplimiento (%)").font = Font(bold=True)
    fila += 1
    
    semilleros = Semillero.objects.all()
    for sem in semilleros:
        proyectos_completados = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            estado_pro='completado',
            fecha_creacion__gte=inicio_trimestre
        ).count()
        
        proyectos_curso = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            estado_pro__in=['planeacion', 'ejecucion']
        ).count()
        
        total = proyectos_completados + proyectos_curso
        tasa = (proyectos_completados / total * 100) if total > 0 else 0
        
        ws.cell(row=fila, column=1, value=sem.nombre)
        ws.cell(row=fila, column=2, value=proyectos_completados)
        ws.cell(row=fila, column=3, value=proyectos_curso)
        ws.cell(row=fila, column=4, value=round(tasa, 1))
        fila += 1
    
    # Guardar la fila final de cumplimiento
    fila_final_cumplimiento = fila - 1
    
    # Bordes para cumplimiento
    for row in ws.iter_rows(min_row=inicio_cumplimiento, max_row=fila_final_cumplimiento, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 2: ENTREGABLES POR MES
    fila += 2
    ws.cell(row=fila, column=1, value="ENTREGABLES COMPLETADOS POR MES").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_entregables = fila
    ws.cell(row=fila, column=1, value="Mes").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Completados").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Tardíos").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Pendientes").font = Font(bold=True)
    fila += 1
    
    # Iterar por los 3 meses del trimestre
    for i in range(3):
        mes_analisis = inicio_trimestre + relativedelta(months=i)
        mes_siguiente = mes_analisis + relativedelta(months=1)
        
        completados = Entregable.objects.filter(
            estado='Completado',
            fecha_fin__gte=mes_analisis,
            fecha_fin__lt=mes_siguiente
        ).count()
        
        tardios = Entregable.objects.filter(
            estado='Entrega Tardía',
            fecha_fin__gte=mes_analisis,
            fecha_fin__lt=mes_siguiente
        ).count()
        
        pendientes = Entregable.objects.filter(
            estado='Pendiente',
            fecha_inicio__gte=mes_analisis,
            fecha_inicio__lt=mes_siguiente
        ).count()
        
        ws.cell(row=fila, column=1, value=mes_analisis.strftime("%B"))
        ws.cell(row=fila, column=2, value=completados)
        ws.cell(row=fila, column=3, value=tardios)
        ws.cell(row=fila, column=4, value=pendientes)
        fila += 1
    
    # Bordes para entregables
    for row in ws.iter_rows(min_row=inicio_entregables, max_row=fila-1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # GRÁFICOS
    # Gráfico 1: Cumplimiento
    chart1 = BarChart()
    chart1.title = "Tasa de Cumplimiento por Semillero"
    chart1.y_axis.title = "Porcentaje (%)"
    
    # Datos: solo porcentajes (sin encabezado)
    data1 = Reference(ws, min_col=4, min_row=inicio_cumplimiento+1, max_row=fila_final_cumplimiento)
    # Categorías: nombres de semilleros
    cats1 = Reference(ws, min_col=1, min_row=inicio_cumplimiento+1, max_row=fila_final_cumplimiento)
    
    chart1.add_data(data1, titles_from_data=False)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "G5")
    
    # Gráfico 2: Entregables
    chart2 = LineChart()
    chart2.title = "Evolución de Entregables"
    chart2.y_axis.title = "Cantidad"
    
    # Guardar la fila final de entregables
    fila_final_entregables = fila - 1
    
    # Datos: columnas 2-4 (Completados, Tardíos, Pendientes) sin encabezados
    data2 = Reference(ws, min_col=2, min_row=inicio_entregables, max_col=4, max_row=fila_final_entregables)
    # Categorías: nombres de meses
    cats2 = Reference(ws, min_col=1, min_row=inicio_entregables+1, max_row=fila_final_entregables)
    
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "G25")
    
    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="informe_trimestral_{trimestre}_{fecha_actual.year}.xlsx"'
    
    wb.save(response)
    return response

def balance_anual(request):
    """
    Resultados del año completo con análisis integral
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Anual"
    
    fecha_actual = timezone.now()
    año_actual = fecha_actual.year
    inicio_año = fecha_actual.replace(month=1, day=1, hour=0, minute=0, second=0)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:F1')
    titulo = ws['A1']
    titulo.value = f"BALANCE ANUAL {año_actual}"
    titulo.font = Font(bold=True, size=18)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # SECCIÓN 1: RESUMEN EJECUTIVO
    fila = 3
    ws.cell(row=fila, column=1, value="RESUMEN EJECUTIVO").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_resumen_anual = fila
    ws.cell(row=fila, column=1, value="Indicador").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Total Año").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Promedio Mensual").font = Font(bold=True)
    fila += 1
    
    # Calcular indicadores
    proyectos_año = Proyecto.objects.filter(fecha_creacion__year=año_actual).count()
    entregables_año = Entregable.objects.filter(
        estado__in=['Completado', 'Entrega Tardía'],
        fecha_fin__year=año_actual
    ).count()
    participantes_nuevos = Usuario.objects.filter(fecha_registro__year=año_actual).count() + \
                          Aprendiz.objects.filter(fecha_registro__year=año_actual).count()
    
    meses_transcurridos = fecha_actual.month
    
    indicadores = [
        ["Proyectos Creados", proyectos_año, round(proyectos_año / meses_transcurridos, 1)],
        ["Entregables Completados", entregables_año, round(entregables_año / meses_transcurridos, 1)],
        ["Participantes Nuevos", participantes_nuevos, round(participantes_nuevos / meses_transcurridos, 1)],
    ]
    
    for ind in indicadores:
        ws.cell(row=fila, column=1, value=ind[0])
        ws.cell(row=fila, column=2, value=ind[1])
        ws.cell(row=fila, column=3, value=ind[2])
        fila += 1
    
    # Bordes para resumen anual
    for row in ws.iter_rows(min_row=inicio_resumen_anual, max_row=fila-1, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 2: EVOLUCIÓN MENSUAL
    fila += 2
    ws.cell(row=fila, column=1, value="EVOLUCIÓN MENSUAL").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_evolucion = fila
    ws.cell(row=fila, column=1, value="Mes").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Proyectos").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Entregables").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Progreso Promedio (%)").font = Font(bold=True)
    fila += 1
    
    # Datos por mes
    for mes in range(1, meses_transcurridos + 1):
        mes_inicio = fecha_actual.replace(month=mes, day=1)
        mes_fin = mes_inicio + relativedelta(months=1)
        
        proyectos_mes = Proyecto.objects.filter(
            fecha_creacion__gte=mes_inicio,
            fecha_creacion__lt=mes_fin
        ).count()
        
        entregables_mes = Entregable.objects.filter(
            estado__in=['Completado', 'Entrega Tardía'],
            fecha_fin__gte=mes_inicio,
            fecha_fin__lt=mes_fin
        ).count()
        
        # Progreso promedio de todos los proyectos
        proyectos_activos = Proyecto.objects.filter(fecha_creacion__lt=mes_fin)
        progreso_promedio = proyectos_activos.aggregate(Avg('progreso'))['progreso__avg'] or 0
        
        ws.cell(row=fila, column=1, value=mes_inicio.strftime("%B"))
        ws.cell(row=fila, column=2, value=proyectos_mes)
        ws.cell(row=fila, column=3, value=entregables_mes)
        ws.cell(row=fila, column=4, value=round(progreso_promedio, 1))
        fila += 1
    
    # Guardar la fila final de evolución
    fila_final_evolucion = fila - 1
    
    # Bordes para evolución
    for row in ws.iter_rows(min_row=inicio_evolucion, max_row=fila_final_evolucion, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 3: RANKING DE SEMILLEROS
    fila += 2
    ws.cell(row=fila, column=1, value="RANKING DE SEMILLEROS").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_ranking = fila
    ws.cell(row=fila, column=1, value="Posición").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Semillero").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Proyectos Completados").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Progreso General (%)").font = Font(bold=True)
    fila += 1
    
    # Calcular ranking
    semilleros_ranking = []
    for sem in Semillero.objects.all():
        proyectos_completados = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            estado_pro='completado'
        ).count()
        
        semilleros_ranking.append({
            'nombre': sem.nombre,
            'completados': proyectos_completados,
            'progreso': sem.progreso_sem
        })
    
    # Ordenar por completados y luego por progreso
    semilleros_ranking.sort(key=lambda x: (x['completados'], x['progreso']), reverse=True)
    
    for i, sem in enumerate(semilleros_ranking, 1):
        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=sem['nombre'])
        ws.cell(row=fila, column=3, value=sem['completados'])
        ws.cell(row=fila, column=4, value=sem['progreso'])
        fila += 1
    
    # Guardar la fila final de ranking
    fila_final_ranking = fila - 1
    
    # Bordes para ranking
    for row in ws.iter_rows(min_row=inicio_ranking, max_row=fila_final_ranking, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # GRÁFICOS
    # Gráfico 1: Evolución mensual
    chart1 = LineChart()
    chart1.title = "Evolución Anual"
    chart1.y_axis.title = "Cantidad"
    
    # Datos: columnas 2-3 (Proyectos, Entregables)
    data1 = Reference(ws, min_col=2, min_row=inicio_evolucion, max_col=3, max_row=fila_final_evolucion)
    # Categorías: nombres de meses
    cats1 = Reference(ws, min_col=1, min_row=inicio_evolucion+1, max_row=fila_final_evolucion)
    
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "H5")
    
    # Gráfico 2: Ranking Top 5
    chart2 = BarChart()
    chart2.title = "Top 5 Semilleros"
    chart2.y_axis.title = "Proyectos Completados"
    
    # Calcular cuántos semilleros mostrar (máximo 5)
    num_semilleros_grafico = min(5, len(semilleros_ranking))
    max_fila_ranking_grafico = inicio_ranking + num_semilleros_grafico
    
    # Datos: solo proyectos completados (columna 3)
    data2 = Reference(ws, min_col=3, min_row=inicio_ranking+1, max_row=max_fila_ranking_grafico)
    # Categorías: nombres de semilleros
    cats2 = Reference(ws, min_col=2, min_row=inicio_ranking+1, max_row=max_fila_ranking_grafico)
    
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "H25")
    
    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="balance_anual_{año_actual}.xlsx"'
    
    wb.save(response)
    return response

def comparativo_anual(request):
    """
    Evolución interanual de indicadores clave de desempeño
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativo Anual"
    
    fecha_actual = timezone.now()
    año_actual = fecha_actual.year
    año_anterior = año_actual - 1
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:E1')
    titulo = ws['A1']
    titulo.value = f"COMPARATIVO {año_anterior} vs {año_actual}"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # SECCIÓN 1: INDICADORES GENERALES
    fila = 3
    ws.cell(row=fila, column=1, value="INDICADORES GENERALES").font = Font(bold=True, size=14)
    fila += 2

    inicio_indicadores = fila
    ws.cell(row=fila, column=1, value="Indicador").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=f"{año_anterior}").font = Font(bold=True)
    ws.cell(row=fila, column=3, value=f"{año_actual}").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Variación (%)").font = Font(bold=True)
    ws.cell(row=fila, column=5, value="Tendencia").font = Font(bold=True)
    fila += 1

    # Calcular indicadores para ambos años
    def calcular_indicadores(año):
        proyectos = Proyecto.objects.filter(fecha_creacion__year=año).count()
        entregables = Entregable.objects.filter(
            estado__in=['Completado', 'Entrega Tardía'],
            fecha_fin__year=año
        ).count()
        participantes = Usuario.objects.filter(fecha_registro__year=año).count() + \
                       Aprendiz.objects.filter(fecha_registro__year=año).count()
        semilleros = Semillero.objects.filter(fecha_creacion__year=año).count()
        
        return {
            'proyectos': proyectos,
            'entregables': entregables,
            'participantes': participantes,
            'semilleros': semilleros
        }
    
    datos_anterior = calcular_indicadores(año_anterior)
    datos_actual = calcular_indicadores(año_actual)
    
    indicadores = [
        ["Proyectos Creados", datos_anterior['proyectos'], datos_actual['proyectos']],
        ["Entregables Completados", datos_anterior['entregables'], datos_actual['entregables']],
        ["Nuevos Participantes", datos_anterior['participantes'], datos_actual['participantes']],
        ["Semilleros Creados", datos_anterior['semilleros'], datos_actual['semilleros']],
    ]
    
    for ind in indicadores:
        anterior = ind[1]
        actual = ind[2]
        
        # Calcular variación y tendencia
        if anterior > 0:
            variacion = ((actual - anterior) / anterior * 100)
            tendencia = "↑" if variacion > 0 else "↓" if variacion < 0 else "="
        elif actual > 0:
            # Si no había datos anteriores pero ahora sí hay, es crecimiento
            variacion = 100
            tendencia = "↑"
        else:
            # Ambos son 0
            variacion = 0
            tendencia = "="
        
        ws.cell(row=fila, column=1, value=ind[0])
        ws.cell(row=fila, column=2, value=anterior)
        ws.cell(row=fila, column=3, value=actual)
        ws.cell(row=fila, column=4, value=round(variacion, 1))
        ws.cell(row=fila, column=5, value=tendencia)
        fila += 1
    
    fin_indicadores = fila - 1
    
    # Bordes para indicadores
    for row in ws.iter_rows(min_row=inicio_indicadores, max_row=fin_indicadores, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 2: COMPARATIVO POR SEMILLERO
    fila += 2
    ws.cell(row=fila, column=1, value="COMPARATIVO POR SEMILLERO").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_semilleros = fila
    ws.cell(row=fila, column=1, value="Semillero").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=f"Proyectos {año_anterior}").font = Font(bold=True)
    ws.cell(row=fila, column=3, value=f"Proyectos {año_actual}").font = Font(bold=True)
    ws.cell(row=fila, column=4, value="Crecimiento (%)").font = Font(bold=True)
    fila += 1
    
    semilleros = Semillero.objects.all()
    for sem in semilleros:
        proyectos_anterior = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            fecha_creacion__year=año_anterior
        ).count()
        
        proyectos_actual = Proyecto.objects.filter(
            semilleroproyecto__id_sem=sem,
            fecha_creacion__year=año_actual
        ).count()
        
        # Calcular crecimiento
        if proyectos_anterior > 0:
            crecimiento = ((proyectos_actual - proyectos_anterior) / proyectos_anterior * 100)
        elif proyectos_actual > 0:
            crecimiento = 100
        else:
            crecimiento = 0
        
        ws.cell(row=fila, column=1, value=sem.nombre)
        ws.cell(row=fila, column=2, value=proyectos_anterior)
        ws.cell(row=fila, column=3, value=proyectos_actual)
        ws.cell(row=fila, column=4, value=round(crecimiento, 1))
        fila += 1
    
    fin_semilleros = fila - 1
    
    # Bordes para semilleros
    for row in ws.iter_rows(min_row=inicio_semilleros, max_row=fin_semilleros, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # SECCIÓN 3: EVOLUCIÓN MENSUAL COMPARADA
    fila += 2
    ws.cell(row=fila, column=1, value="EVOLUCIÓN MENSUAL COMPARADA").font = Font(bold=True, size=14)
    fila += 2
    
    inicio_mensual = fila
    ws.cell(row=fila, column=1, value="Mes").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=f"Proyectos {año_anterior}").font = Font(bold=True)
    ws.cell(row=fila, column=3, value=f"Proyectos {año_actual}").font = Font(bold=True)
    fila += 1
    
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    for mes_num in range(1, 13):
        proyectos_anterior = Proyecto.objects.filter(
            fecha_creacion__year=año_anterior,
            fecha_creacion__month=mes_num
        ).count()
        
        proyectos_actual = Proyecto.objects.filter(
            fecha_creacion__year=año_actual,
            fecha_creacion__month=mes_num
        ).count()
        
        ws.cell(row=fila, column=1, value=meses[mes_num-1])
        ws.cell(row=fila, column=2, value=proyectos_anterior)
        ws.cell(row=fila, column=3, value=proyectos_actual)
        fila += 1
    
    fin_mensual = fila - 1
    
    # Bordes para mensual
    for row in ws.iter_rows(min_row=inicio_mensual, max_row=fin_mensual, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border

    # GRÁFICOS CON LEYENDAS 
    from openpyxl.chart import Series
    
    # Gráfico 1: Comparativo general
    chart1 = BarChart()
    chart1.title = "Comparativo General"
    chart1.y_axis.title = "Cantidad"
    chart1.x_axis.title = "Indicadores"
    
    # Crear series individuales para cada año (columna 2 = año anterior, columna 3 = año actual)
    values_anterior = Reference(ws, min_col=2, min_row=inicio_indicadores+1, max_row=fin_indicadores)
    values_actual = Reference(ws, min_col=3, min_row=inicio_indicadores+1, max_row=fin_indicadores)
    cats1 = Reference(ws, min_col=1, min_row=inicio_indicadores+1, max_row=fin_indicadores)
    
    # Serie 1: Año anterior
    serie1 = Series(values_anterior, title=f"{año_anterior}")
    chart1.series.append(serie1)
    
    # Serie 2: Año actual
    serie2 = Series(values_actual, title=f"{año_actual}")
    chart1.series.append(serie2)
    
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "G3")
    
    # Gráfico 2: Evolución mensual
    chart2 = LineChart()
    chart2.title = "Evolución Mensual"
    chart2.y_axis.title = "Proyectos"
    chart2.x_axis.title = "Mes"
    
    # Crear series individuales para cada año
    values_mensual_anterior = Reference(ws, min_col=2, min_row=inicio_mensual+1, max_row=fin_mensual)
    values_mensual_actual = Reference(ws, min_col=3, min_row=inicio_mensual+1, max_row=fin_mensual)
    cats2 = Reference(ws, min_col=1, min_row=inicio_mensual+1, max_row=fin_mensual)
    
    # Serie 1: Proyectos año anterior
    serie_mensual1 = Series(values_mensual_anterior, title=f"Proyectos {año_anterior}")
    chart2.series.append(serie_mensual1)
    
    # Serie 2: Proyectos año actual
    serie_mensual2 = Series(values_mensual_actual, title=f"Proyectos {año_actual}")
    chart2.series.append(serie_mensual2)
    
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "G20")
    
    # Gráfico 3: Crecimiento por semillero
    chart3 = BarChart()
    chart3.title = "Crecimiento por Semillero"
    chart3.y_axis.title = "Porcentaje (%)"
    chart3.x_axis.title = "Semilleros"
    
    # Datos de crecimiento (columna 4, sin incluir encabezados)
    data3 = Reference(ws, min_col=4, min_row=inicio_semilleros+1, max_row=fin_semilleros)
    cats3 = Reference(ws, min_col=1, min_row=inicio_semilleros+1, max_row=fin_semilleros)
    
    chart3.add_data(data3, titles_from_data=False)
    chart3.set_categories(cats3)
    ws.add_chart(chart3, "G37")
    
    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="comparativo_anual_{año_anterior}_{año_actual}.xlsx"'
    
    wb.save(response)
    return response

def reporte_programa(request):
    """
    Agrupación de proyectos formativos según programa de formación
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos por Programa"
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:F1')
    titulo = ws['A1']
    titulo.value = "PROYECTOS FORMATIVOS POR PROGRAMA"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # ENCABEZADOS
    ws.append([])
    ws.append([
        "Programa de Formación",
        "Cantidad de Proyectos",
        "Proyectos Activos",
        "Proyectos Completados",
        "Total Aprendices",
        "Progreso Promedio (%)"
    ])
    
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    
    # OBTENER DATOS - Versión mejorada
    # Primero verificamos todos los proyectos formativos
    proyectos_formativos = Proyecto.objects.filter(tipo__iexact="formativo")
    
    # Si no hay programa_formacion definido, agrupamos por el campo 'programa' de los aprendices
    if proyectos_formativos.filter(programa_formacion__isnull=False).exclude(programa_formacion="").exists():
        # Opción 1: Agrupar por programa_formacion del proyecto
        programas = proyectos_formativos.filter(
            programa_formacion__isnull=False
        ).exclude(programa_formacion="").values('programa_formacion').distinct()
        
        usar_programa_proyecto = True
    else:
        # Opción 2: Agrupar por programa del aprendiz
        programas = Aprendiz.objects.filter(
            proyectoaprendiz__cod_pro__tipo__iexact="formativo"
        ).values('programa').distinct()
        
        usar_programa_proyecto = False
    
    fila_inicio = 4
    hay_datos = False
    
    for programa_dict in programas:
        if usar_programa_proyecto:
            programa = programa_dict['programa_formacion']
            # Proyectos de este programa
            proyectos = Proyecto.objects.filter(
                tipo__iexact="formativo",
                programa_formacion=programa
            )
        else:
            programa = programa_dict['programa']
            # Obtener aprendices de este programa
            aprendices_programa = Aprendiz.objects.filter(programa=programa)
            # Proyectos donde participan estos aprendices
            proyectos = Proyecto.objects.filter(
                tipo__iexact="formativo",
                proyectoaprendiz__cedula_apre__in=aprendices_programa.values_list('cedula_apre', flat=True)
            ).distinct()
        
        total_proyectos = proyectos.count()
        
        if total_proyectos == 0:
            continue
        
        hay_datos = True
        
        # Proyectos activos (en curso)
        proyectos_activos = proyectos.filter(
            estado_pro__in=['planeacion', 'ejecucion', 'activo', 'en curso']
        ).count()
        
        # Proyectos completados
        proyectos_completados = proyectos.filter(
            estado_pro__in=['completado', 'finalizado', 'terminado']
        ).count()
        
        # Total aprendices en estos proyectos
        total_aprendices = ProyectoAprendiz.objects.filter(
            cod_pro__in=proyectos.values_list('cod_pro', flat=True)
        ).values('cedula_apre').distinct().count()
        
        # Progreso promedio
        progreso_promedio = proyectos.aggregate(Avg('progreso'))['progreso__avg'] or 0
        
        ws.append([
            programa or "Sin Programa",
            total_proyectos,
            proyectos_activos,
            proyectos_completados,
            total_aprendices,
            round(progreso_promedio, 1)
        ])
    
    # Si no hay datos, agregar una fila informativa
    if not hay_datos:
        ws.append([
            "No hay proyectos formativos registrados",
            0,
            0,
            0,
            0,
            0
        ])
    
    # BORDES
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
    
    # Solo crear gráficos si hay datos reales
    if hay_datos and ws.max_row > 3:
        # GRÁFICO 1: Proyectos por Programa
        chart1 = BarChart()
        chart1.title = "Cantidad de Proyectos por Programa"
        chart1.y_axis.title = "Número de Proyectos"
        chart1.x_axis.title = "Programa de Formación"
        
        data1 = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row)
        cats1 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.height = 15
        chart1.width = 20
        
        ws.add_chart(chart1, "H3")
        
        # GRÁFICO 2: Progreso Promedio
        chart2 = LineChart()
        chart2.title = "Progreso Promedio por Programa"
        chart2.y_axis.title = "Progreso (%)"
        chart2.x_axis.title = "Programa de Formación"
        
        data2 = Reference(ws, min_col=6, min_row=3, max_row=ws.max_row)
        cats2 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.height = 15
        chart2.width = 20
        
        ws.add_chart(chart2, "H21")
        
        # GRÁFICO 3: Estado de Proyectos
        chart3 = PieChart()
        chart3.title = "Distribución de Estados"
        
        # Calcular totales para el gráfico
        fila_totales = ws.max_row + 3
        ws.cell(row=fila_totales, column=1, value="Estado").font = Font(bold=True)
        ws.cell(row=fila_totales, column=2, value="Cantidad").font = Font(bold=True)
        
        # CORRECCIÓN: Manejar valores None en las celdas
        total_activos = sum([ws.cell(row=i, column=3).value or 0 for i in range(4, ws.max_row + 1)])
        total_completados = sum([ws.cell(row=i, column=4).value or 0 for i in range(4, ws.max_row + 1)])
        
        ws.append(["Activos", total_activos])
        ws.append(["Completados", total_completados])
        
        data3 = Reference(ws, min_col=2, min_row=fila_totales, max_row=ws.max_row)
        labels3 = Reference(ws, min_col=1, min_row=fila_totales + 1, max_row=ws.max_row)
        
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(labels3)
        
        ws.add_chart(chart3, "H39")
    
    # RESPUESTA
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="proyectos_formativos_programa.xlsx"'
    
    wb.save(response)
    return response

def reporte_fichas(request):
    """
    Fichas participantes y sus proyectos de investigación asociados
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vinculación Fichas"
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # TÍTULO
    ws.merge_cells('A1:G1')
    titulo = ws['A1']
    titulo.value = "VINCULACIÓN DE FICHAS A PROYECTOS"
    titulo.font = Font(bold=True, size=16)
    titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # ENCABEZADOS
    ws.append([])
    ws.append([
        "Ficha",
        "Programa",
        "Total Aprendices",
        "Aprendices en Proyectos",
        "Proyectos Asociados",
        "Semilleros Vinculados",
        "Tasa de Participación (%)"
    ])
    
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    
    # OBTENER DATOS
    # Agrupar aprendices por ficha
    fichas = Aprendiz.objects.values('ficha', 'programa').distinct().order_by('ficha')
    
    for ficha_dict in fichas:
        ficha = ficha_dict['ficha']
        programa = ficha_dict['programa']
        
        # Total aprendices en esta ficha
        aprendices_ficha = Aprendiz.objects.filter(ficha=ficha)
        total_aprendices = aprendices_ficha.count()
        
        # Aprendices que están en proyectos
        aprendices_en_proyectos = ProyectoAprendiz.objects.filter(
            cedula_apre__in=aprendices_ficha.values_list('cedula_apre', flat=True)
        ).values('cedula_apre').distinct().count()
        
        # Proyectos asociados
        proyectos_asociados = Proyecto.objects.filter(
            proyectoaprendiz__cedula_apre__in=aprendices_ficha.values_list('cedula_apre', flat=True)
        ).distinct().count()
        
        # Semilleros vinculados
        semilleros_vinculados = Semillero.objects.filter(
            id_sem__in=aprendices_ficha.values_list('id_sem', flat=True)
        ).distinct().count()
        
        # Tasa de participación
        tasa_participacion = (aprendices_en_proyectos / total_aprendices * 100) if total_aprendices > 0 else 0
        
        ws.append([
            ficha,
            programa,
            total_aprendices,
            aprendices_en_proyectos,
            proyectos_asociados,
            semilleros_vinculados,
            round(tasa_participacion, 1)
        ])
    
    # BORDES
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
    
    # GRÁFICO 1: Aprendices por Ficha
    chart1 = BarChart()
    chart1.title = "Aprendices por Ficha"
    chart1.y_axis.title = "Cantidad de Aprendices"
    chart1.x_axis.title = "Ficha"
    
    data1 = Reference(ws, min_col=3, min_row=3, max_row=ws.max_row)
    cats1 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.height = 15
    chart1.width = 20
    
    ws.add_chart(chart1, "I3")
    
    # GRÁFICO 2: Tasa de Participación
    chart2 = LineChart()
    chart2.title = "Tasa de Participación por Ficha"
    chart2.y_axis.title = "Porcentaje (%)"
    chart2.x_axis.title = "Ficha"
    
    data2 = Reference(ws, min_col=7, min_row=3, max_row=ws.max_row)
    cats2 = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 15
    chart2.width = 20
    
    ws.add_chart(chart2, "I21")
    
    # SECCIÓN DETALLADA: Proyectos por Ficha
    fila_detalle = ws.max_row + 3
    ws.cell(row=fila_detalle, column=1, value="DETALLE DE PROYECTOS POR FICHA").font = Font(bold=True, size=14)
    fila_detalle += 2
    
    ws.cell(row=fila_detalle, column=1, value="Ficha").font = Font(bold=True)
    ws.cell(row=fila_detalle, column=2, value="Proyecto").font = Font(bold=True)
    ws.cell(row=fila_detalle, column=3, value="Tipo").font = Font(bold=True)
    ws.cell(row=fila_detalle, column=4, value="Estado").font = Font(bold=True)
    ws.cell(row=fila_detalle, column=5, value="Aprendices Participantes").font = Font(bold=True)
    fila_detalle += 1
    
    inicio_detalle = fila_detalle
    
    for ficha_dict in fichas:
        ficha = ficha_dict['ficha']
        
        # Obtener aprendices de esta ficha
        aprendices_ids = Aprendiz.objects.filter(ficha=ficha).values_list('cedula_apre', flat=True)
        
        # Obtener proyectos donde participan estos aprendices
        proyectos = Proyecto.objects.filter(
            proyectoaprendiz__cedula_apre__in=aprendices_ids
        ).distinct()
        
        for proyecto in proyectos:
            # Contar aprendices de esta ficha en este proyecto
            aprendices_proyecto = ProyectoAprendiz.objects.filter(
                cod_pro=proyecto,
                cedula_apre__in=aprendices_ids
            ).count()
            
            ws.append([
                ficha,
                proyecto.nom_pro,
                proyecto.tipo,
                proyecto.estado_pro,
                aprendices_proyecto
            ])
            fila_detalle += 1
    
    # Bordes para detalle
    for row in ws.iter_rows(min_row=inicio_detalle-1, max_row=fila_detalle-1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
    
    # RESPUESTA
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="vinculacion_fichas.xlsx"'
    
    wb.save(response)
    return response

# VISTA DE LOGOUT
def logout(request):
    # Limpiar toda la sesión
    request.session.flush()
    messages.success(request, "Has cerrado sesión correctamente")
    return redirect('iniciarsesion')